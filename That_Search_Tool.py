# -- coding: utf-8 --
""" 
Created: 2025-07-11 by TBuchanan
Revised: 2025-11-05 by SEberhard + AI

Tool Objectives:
    + Capable of searching PSFT for parts and related BU inventories
        - Provides BU based inventories (OnHand, Reserved, Available). Accuracy up to ~24hr.
        - Provides MFG and MFG PN if recorded in PSFT
        - Provides Lot/Serial Control designation along with Quality Code assignments
        - Shows if a part is Active within the BU in PSFT
        - Shows which vault controls the part
        - Provides the perpetual average cost for a part as recorded in PSFT
        - (Not Yet Implemented) Allows for Where-Used part searches
        - (Not Yet Implemented) Allows for Item-Contains (BOM Review) part searches
        - (Not Yet Implemented) Allows for Last-Purchased/Ordered part searches
        - (Not Yet Implemented) Allows for NC/Deviation part searches
        - (Not Yet Implemented) Allows for rental-fleet part searches
    + Capable of searching imported csv boms for assembly wide inventory searches
    + Capable of Indexing the various user root folders to perform quick file searches
    + Capable of switching between psft search and file search as needed
"""

from urllib3.util.retry import Retry
from urllib3.exceptions import InsecureRequestWarning
from requests.adapters import HTTPAdapter
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMenu,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QShortcut,
    QSizePolicy,
    QStackedWidget,
    QTableView,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWidget,
)
from PyQt5.QtGui import (
    QBrush,
    QColor,
    QFont,
    QKeySequence,
    QPalette,
    QSyntaxHighlighter,
    QTextCharFormat,
)
from PyQt5.QtCore import (
    QAbstractTableModel,
    QByteArray,
    QDateTime,
    QModelIndex,
    QSortFilterProxyModel,
    Qt,
    QThread,
    QTimer,
    pyqtSignal,
    pyqtSlot,
)
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl import load_workbook
import win32com.client as win32
import urllib3
import requests
import pandas as pd
import keyring
from typing import Dict, List, Optional
from pathlib import Path
from datetime import datetime
import webbrowser
import uuid
import time
import threading
import sys
import subprocess
import sqlite3
import shutil
import re
import os
import json
import io
import hashlib
import glob
import fnmatch
import ctypes
import copy
import base64
import random

APP_VERSION = "Baseline-2025-08-27-a"


PN_REGEXES = [
    r"\b[A-Z]{2,}-\d{3,}\b",
    r"\b\d{5,}[A-Z]?\b",
    r"\b[A-Z0-9]{4,}[-_\.][A-Z0-9]{2,}\b",
]

# --- Environment quirks (ZScaler MITM) ---
urllib3.disable_warnings(InsecureRequestWarning)

# ----------------------------------
# Config
# ----------------------------------
BASE_URL = "https://denododev.oii.oceaneering.com/denodo-restfulws/finance/views"
ITEMS_VIEW = "item_master"
INV_VIEW = "inventory_search_all"
ITEM_CONTROLS = "bv_psfinance_ps_master_item_tbl"

# Only show these if present
DESIRED_HEADERS = [
    "BU",
    "part_number",
    "part_description",
    "revision",
    "UOM",
    "available_quantity",
    "q_codes",
    "lot",
    "serial",
    "onhand_quantity",
    "reserved_quantity",
    "mfg_id",
    "mfg_part",
    "perpetual_avg_cost",
    "family",
    "source",
    "item_status",
    "ctrl_in",
    "group_description",
]
POSSIBLE_BUS = {
    "BR001": "BR001 - Boris 01",
    "BUABO": "Oceaneering Intl Services Ltd",
    "BUACE": "Oceaneering Mobile Robotics BV",
    "BUAIN": "Oceaneering Asset Integrity AS",
    "BUAIR": "AIRSIS, Inc",
    "BUANG": "Oceaneering Angola SA",
    "BUAOS": "Oceaneering Intl Services Ltd.",
    "BUAPR": "Oceaneering International GmbH",
    "BUAUS": "Oceaneering Australia Pty Ltd",
    "BUBAK": "OISL Azerbaijan Branch",
    "BUBOP": "Oceaneering International Inc.",
    "BUBRO": "Brazil ROV",
    "BUBTM": "Oceaneering Intl GmbH - Batam",
    "BUBXO": "Brazil Non-Owned ROV Inventory",
    "BUCCT": "Oceaneering Survey Services",
    "BUCCU": "Oceaneering Intl Services Ltd.",
    "BUCHD": "OISL India Branch",
    "BUCLO": "Oceaneering International Inc.",
    "BUDSO": "Oceaneering International Inc.",
    "BUDTS": "Oceaneering International, Inc",
    "BUEGO": "Oceaneering Int'l AG - EGuinea",
    "BUEGY": "SOSI - Egypt Branch",
    "BUFTW": "Oceaneering International Inc.",
    "BUGHA": "Oceaneering Intl Serv Ghana",
    "BUGHL": "Oceaneering Ghana Limited",
    "BUGHN": "Oceaneering Intl Svc Ltd Ghana",
    "BUGPV": "Oceaneering International, Inc",
    "BUGUY": "OSOL - Guyana Branch",
    "BUHFO": "Oceaneering International Inc.",
    "BUHPC": "OI High Performance Cables Inc",
    "BUHSP": "OI HPC Spare Parts",
    "BUIEH": "Oceaneering International Inc.",
    "BUIER": "OIE Reflange",
    "BUKAO": "Oceaneering International GmbH",
    "BUKKD": "OIGmbH Kakinada Branch",
    "BULUL": "Oceaneering Angola SA",
    "BUMAC": "BUMAC",
    "BUMCO": "Oceaneering International Inc.",
    "BUMSP": "R&M Inventory - BUMUS",
    "BUMTN": "OI GmbH",
    "BUMUK": "Oceaneering Intl Services Ltd",
    "BUMUS": "Oceaneering International Inc.",
    "BUNCH": "NCA, Inc.- An Oceaneering Co.",
    "BUNCN": "Oceaneering NCA AS",
    "BUOAD": "Oceaneering OIS Co WLL",
    "BUOAS": "Oceaneering International Inc.",
    "BUOCS": "OI Communications Solutions",
    "BUOES": "Oceaneering International Inc",
    "BUOTM": "Oceaneering International Inc.",
    "BUOVN": "Oceaneering International Inc.",
    "BUPHL": "Solus Schall Nigeria Ltd",
    "BUPHN": "OI GmbH Nigeria PE",
    "BUPHO": "Oceaneering International AG",
    "BUPHT": "OceaneeringServicesNigeriaLTD",
    "BUPRO": "OII Australia Branch",
    "BUROT": "Oceaneering Rotator A/S",
    "BUSFR": "Oceaneering Intl Services Ltd",
    "BUSGO": "Oceaneering International GmbH",
    "BUSTO": "Stavanger",
    "BUTDL": "Technology Design Ltd",
    "CM025": "CM025 - Comache 025",
    "EMG01": "eMagnum 01",
    "EMG02": "E Magnum 02",
    "EN001": "EN001 - Enovus 001",
    "EN002": "EN002 - Enovus 002",
    "EN003": "EN003 – Enovus 003",
    "EN004": "EN004 – Enovus 004",
    "FADR1": "FADR1 - ROV Falcon DR1",
    "FL001": "FL001 - Falcon 1",
    "FR001": "FR001 - Freedom 001",
    "GC043": "GC043",
    "GE001": "GE001 - Global Explorer 01",
    "HL001": "HL001 - High Speed Loader 01",
    "HL002": "HL002 - High Speed Loader 02",
    "IR001": "IR001 - Iris 01",
    "IS001": "IS001 - Isurus 001",
    "IS002": "IS002 - Isurus 002",
    "IS003": "IS003 - Isurus 003",
    "IS004": "IS004 – ROV Isurus 004",
    "IS005": "IS005 - ROV Isurus 005",
    "IS006": "IS006 - ROV Isurus 006",
    "IS007": "IS007 - Isurus 007",
    "LP014": "LP014 - ROV Saab Leopard 14",
    "MG001": "Magnum 001",
    "MG002": "MG002 - Magnum 02",
    "MG007": "Magnum 7",
    "MG009": "Magnum 009",
    "MG012": "Magnum 12",
    "MG014": "Magnum 014",
    "MG016": "Magnum 016",
    "MG018": "Magnum 18",
    "MG020": "Magnum 020",
    "MG021": "MG021 - Magnum 21",
    "MG022": "Magnum 022",
    "MG023": "Magnum 023",
    "MG024": "Magnum 24",
    "MG028": "Magnum 028",
    "MG029": "Magnum 29",
    "MG030": "Magnum 030",
    "MG031": "Magnum 031",
}
DEFAULT_BUS = ["BUIEH", "BUMCO", "BUMAC"]

# ----------------------------------
# Utilities
# ----------------------------------


def ensure_dir(d: str) -> str:
    os.makedirs(d, exist_ok=True)
    return d


def _win_path(env_name: str, fallback: Path) -> Path:
    p = os.environ.get(env_name, "") or ""
    try:
        p = Path(p)
    except Exception:
        p = fallback
    return p


def app_paths() -> dict:
    """Centralized app paths, using per-user AppData (writeable, no admin)."""
    base_user = _win_path("APPDATA", Path.home() / "AppData" / "Roaming") / "PartSearch"
    base_shared = _win_path("PROGRAMDATA", Path("C:/ProgramData")) / "PartSearch"
    base_user.mkdir(parents=True, exist_ok=True)
    (base_user / "index").mkdir(parents=True, exist_ok=True)
    return {
        "base_user": base_user,
        "base_shared": base_shared,
        "prefs": base_user / "user_prefs.json",
        "locations": base_user / "search_locations.json",
        "index_roots": base_user / "index_roots.json",
        "enovia_policy": base_user / "enovia_policy.json",
        "enovia_cache": base_user / "enovia_cache.json",
        "index_dir": base_user / "index",
    }


def _read_user_prefs() -> dict:
    try:
        with open(str(app_paths()["prefs"]), "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}


def _write_user_prefs(updates: dict) -> None:
    """
    Merge 'updates' into existing prefs, optionally filtering to PREF_WHITELIST.
    Writes to %APPDATA%/PartSearch/user_prefs.json via app_paths()["prefs"].
    """
    path = str(app_paths()["prefs"])
    cur = _read_user_prefs()

    # Merge (shallow). If you want nested merges, add a tiny deep-merge helper.
    for k, v in (updates or {}).items():
        if PREF_WHITELIST and k not in PREF_WHITELIST:
            # either skip, or include if you don't want a whitelist
            continue
        cur[k] = v

    _atomic_write_json(path, cur)


# --- Paths and Directories ---
P = app_paths()
PROGRAMDATA_DIR = ensure_dir(
    os.path.join(os.environ.get("PROGRAMDATA", r"C:\ProgramData"), "PartSearch")
)
LOCALAPPDATA_DIR = ensure_dir(
    os.path.join(
        os.environ.get("LOCALAPPDATA", os.path.expanduser(r"~\AppData\Local")),
        "PartSearch",
    )
)
INDEX_ROOTS_JSON = str(P["index_roots"])
SEARCH_LOCATIONS_JSON = str(P["locations"])

PREF_WHITELIST = {
    # examples — include your actual keys here
    "dark_mode",
    "theme",
    "window_geom",
    "last_tab",
    "search_order_matters",
    "bom_mode_last",
    "limit_default",
    "psft_defaults",
    "bus_defaults",
    "file_inputs",
}

if "DEFAULT_PREFS" not in globals():
    DEFAULT_PREFS = {
        "default_bus": [],
        "filter_defaults": {
            "min_qty": "",
            "and_terms": [],
            "or_terms": [],
            "either_or_groups": [],
            "include_unassigned": True,
            "preferred_bu": "",
        },
        "ui": {"geometry": None, "last_tab": "psft"},
    }

DEFAULT_LOCATIONS = {"checked_roots": [], "last_updated": None}

DEFAULT_INDEX_ROOTS = {
    "roots": [
        # {"path": "C:\\", "last_full_scan": None, "files_count": 0, "fts_ok": False}
    ]
}

DEFAULT_ENOVIA_POLICY = {
    "enabled": False,
    "base_url": "",
    "auto_import_csv": False,
    "last_csv_path": "",
    "collab_spaces": [],
}

DEFAULT_ENOVIA_CACHE = {"last_sync": None, "records": []}

# --- Formatting ---


def _win_apps_uses_light() -> bool:
    """True if Windows 'AppsUseLightTheme' is 1 (light), False if 0 (dark)."""
    try:
        import winreg

        with winreg.OpenKey(
            winreg.HKEY_CURRENT_USER,
            r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize",
        ) as k:
            val, _ = winreg.QueryValueEx(k, "AppsUseLightTheme")
            return bool(val)
    except Exception:
        # Default to light if we cannot determine
        return True


def _apply_dark_palette(app: QApplication):
    app.setStyle("Fusion")  # Fusion respects palette
    p = QPalette()

    # Base colors
    p.setColor(QPalette.Window, QColor(53, 53, 53))
    p.setColor(QPalette.WindowText, Qt.white)
    p.setColor(QPalette.Base, QColor(35, 35, 35))
    p.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
    p.setColor(QPalette.ToolTipBase, Qt.white)
    p.setColor(QPalette.ToolTipText, Qt.white)
    p.setColor(QPalette.Text, Qt.white)
    p.setColor(QPalette.Button, QColor(53, 53, 53))
    p.setColor(QPalette.ButtonText, Qt.white)
    p.setColor(QPalette.BrightText, Qt.red)

    # Links / highlights
    p.setColor(QPalette.Highlight, QColor(42, 130, 218))
    p.setColor(QPalette.HighlightedText, Qt.black)

    # Disabled
    p.setColor(QPalette.Disabled, QPalette.Text, QColor(164, 164, 164))
    p.setColor(QPalette.Disabled, QPalette.ButtonText, QColor(164, 164, 164))

    app.setPalette(p)


def _apply_light_palette(app: QApplication):
    # Reset to Fusion light
    app.setStyle("Fusion")
    app.setPalette(QApplication.style().standardPalette())


def apply_theme(app: QApplication, mode: str = "system"):
    """
    mode: "system" | "light" | "dark"
    """
    m = (mode or "system").lower()
    if m == "dark":
        _apply_dark_palette(app)
    elif m == "light":
        _apply_light_palette(app)
    else:
        # system
        if sys.platform.startswith("win"):
            is_light = _win_apps_uses_light()
            _apply_light_palette(app) if is_light else _apply_dark_palette(app)
        else:
            _apply_light_palette(app)  # sensible default on non-Windows


def get_basic_auth_header(
    service: str, username: Optional[str] = None
) -> Dict[str, str]:
    """Pull credentials from keyring and return headers with Basic auth. Store with keyring.set_password(service, username, password)."""
    if username:
        pwd = keyring.get_password(service, username)
        if pwd is None:
            raise RuntimeError(
                f"No password found in keyring for service={service} user={username}"
            )
        creds = f"{username}:{pwd}"
    else:
        cred = keyring.get_credential(service, None)
        if cred is None:
            raise RuntimeError(
                "No credentials found in keyring. Click 'Generate Credentials' first."
            )
        creds = f"{cred.username}:{cred.password}"
    tok = base64.b64encode(creds.encode()).decode()
    return {"Authorization": f"Basic {tok}", "Accept": "application/json"}


# ----- BOOTSTRAP: paths + json helpers + defaults -----


def _ensure_dirs():
    os.makedirs(PROGRAMDATA_DIR, exist_ok=True)


def _atomic_write_json(path: str, data: dict) -> None:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = f"{path}.{uuid.uuid4().hex}.tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    # atomic on same volume; also avoids WinError 5 races with AV/EDR
    os.replace(tmp, path)


def _canon_key(p: str) -> str:
    if not p:
        return ""
    k = os.path.normcase(os.path.normpath(p)).rstrip("\\/")
    return k


def _load_json_or_default(path: str, default: dict) -> dict:
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return default
        return data
    except Exception:
        return default


def _normkey(p: str) -> str:
    # normalize & case-fold so dedupe works across OS case handling
    try:
        return os.path.normcase(os.path.normpath(p))
    except Exception:
        return p.strip()


def read_index_roots_json() -> dict:
    """
    Unified source of truth for IndexBuilderPane and FileSearchPane.
    Shape: {"roots":[{"path": "...", "checked": true, "files_count": 0, "updated_count": 0,
                      "last_full_scan": "YYYY-MM-DD HH:MM", ...}]}
    """
    data = _load_json_or_default(INDEX_ROOTS_JSON, {"roots": []})
    roots = []
    seen = set()
    for r in data.get("roots", []):
        p = (r.get("path") or "").strip()
        if not p:
            continue
        k = _normkey(p)
        if k in seen:
            continue
        seen.add(k)
        roots.append(
            {
                "path": p,
                "checked": bool(r.get("checked", True)),
                "files_count": int(r.get("files_count", 0) or 0),
                "updated_count": int(r.get("updated_count", 0) or 0),
                "last_full_scan": r.get("last_full_scan"),
            }
        )
    return {"roots": roots}


def write_index_roots_json(data: dict) -> None:
    """
    Merge incoming roots into existing index_roots.json (per-user).
    Preserves: checked, files_count, updated_count, last_full_scan.
    """
    # read current
    try:
        with open(INDEX_ROOTS_JSON, "r", encoding="utf-8") as f:
            current = json.load(f) or {}
    except Exception:
        current = {}

    cur_roots = current.get("roots", [])
    by_key = {
        _canon_key(r.get("path", "")): dict(r) for r in cur_roots if r.get("path")
    }

    for r in (data or {}).get("roots", []):
        p = (r.get("path") or "").strip()
        if not p:
            continue
        k = _canon_key(p)
        cur = by_key.get(k, {"path": p})

        # preserve/merge your known fields
        cur["checked"] = bool(r.get("checked", cur.get("checked", True)))
        cur["files_count"] = int(r.get("files_count", cur.get("files_count", 0)) or 0)
        cur["updated_count"] = int(
            r.get("updated_count", cur.get("updated_count", 0)) or 0
        )
        cur["last_full_scan"] = r.get("last_full_scan", cur.get("last_full_scan"))

        by_key[k] = cur

    out = {"roots": list(by_key.values())}
    _atomic_write_json(INDEX_ROOTS_JSON, out)


def read_search_locations() -> dict:
    try:
        with open(SEARCH_LOCATIONS_JSON, "r", encoding="utf-8") as f:
            obj = json.load(f)
    except Exception:
        return {"checked_roots": []}

    # coerce odd shapes back to the expected schema
    if isinstance(obj, dict):
        roots = obj.get("checked_roots") or obj.get("roots") or []
        if isinstance(roots, str):
            roots = [roots]
        obj["checked_roots"] = list(roots or [])
        return obj

    if isinstance(obj, (list, tuple, set)):
        return {"checked_roots": list(obj)}

    if isinstance(obj, str):
        return {"checked_roots": [obj]}

    return {"checked_roots": []}


def write_search_locations(data) -> None:
    """
    Accepts:
      - dict like {"checked_roots": [...], ...}
      - list/tuple/set of paths
      - single path string
    Writes to %APPDATA%/PartSearch/search_locations.json atomically.
    """
    # normalize to a dict with checked_roots
    if isinstance(data, dict):
        payload = dict(data)  # shallow copy
        roots = payload.get("checked_roots") or payload.get("roots") or []
        # if they accidentally passed a bare string in the dict, normalize it too
        if isinstance(roots, str):
            roots = [roots]
    elif isinstance(data, (list, tuple, set)):
        roots = list(data)
        payload = {"checked_roots": roots}
    elif isinstance(data, str):
        roots = [data]
        payload = {"checked_roots": roots}
    else:
        roots = []
        payload = {"checked_roots": roots}

    # de-dupe while preserving display strings
    seen, dedup = set(), []
    for p in roots:
        k = _canon_key(p)
        if not k or k in seen:
            continue
        seen.add(k)
        dedup.append(p)
    payload["checked_roots"] = dedup

    _atomic_write_json(SEARCH_LOCATIONS_JSON, payload)


def save_json_atomic(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = Path(str(path) + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, ensure_ascii=False)
    os.replace(tmp, path)  # atomic on Win11


def load_json_or_default(path: Path, default_obj):
    try:
        if path.exists() and path.stat().st_size > 0:
            with path.open("r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    save_json_atomic(path, default_obj)
    return default_obj


def maybe_migrate_from_programdata(src_path: Path, dst_path: Path):
    """If legacy ProgramData file exists and user file is missing, copy once."""
    try:
        if src_path.exists() and not dst_path.exists():
            dst_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src_path, dst_path)
    except Exception:
        pass  # non-fatal


def _tz_abbrev(dt: datetime) -> str:
    name = (dt.tzname() or "").strip()
    if not name:
        return "UTC"
    if " " in name:  # e.g., "Central Daylight Time" -> "CDT"
        return "".join(w[0] for w in name.split()).upper()
    return name


def local_now_short() -> str:
    dt = datetime.now().astimezone()
    return dt.strftime("%Y-%m-%d %H:%M:%S ") + _tz_abbrev(dt)


# ---- default JSON contents ----
def bootstrap_files():
    """Create any missing JSONs in %APPDATA%\\PartSearch and migrate old ones."""
    p = app_paths()
    # migrate once from ProgramData (legacy installs)
    maybe_migrate_from_programdata(p["base_shared"] / "user_prefs.json", p["prefs"])
    maybe_migrate_from_programdata(
        p["base_shared"] / "search_locations.json", p["locations"]
    )
    maybe_migrate_from_programdata(
        p["base_shared"] / "index_roots.json", p["index_roots"]
    )
    maybe_migrate_from_programdata(
        p["base_shared"] / "enovia_policy.json", p["enovia_policy"]
    )
    # ensure files exist
    load_json_or_default(p["prefs"], DEFAULT_PREFS)
    load_json_or_default(p["locations"], DEFAULT_LOCATIONS)
    load_json_or_default(p["index_roots"], DEFAULT_INDEX_ROOTS)
    load_json_or_default(p["enovia_policy"], DEFAULT_ENOVIA_POLICY)
    load_json_or_default(p["enovia_cache"], DEFAULT_ENOVIA_CACHE)


# --- Policy helpers ---


def enovia_config():
    """Read ENOVIA Online policy. Safe defaults if policy.json is missing."""
    try:
        pdir = os.environ.get("PROGRAMDATA", r"C:\ProgramData")
        with open(
            os.path.join(pdir, "PartSearch", "policy.json"), "r", encoding="utf-8"
        ) as f:
            pol = json.load(f)
    except Exception:
        pol = {}
    cfg = pol.get("enovia_online", {})
    return {
        "enabled": bool(cfg.get("enabled", True)),
        "search_url": cfg.get("search_url_template", ""),
        "downloads": os.path.expandvars(
            cfg.get("downloads_dir", os.path.join(os.path.expanduser("~"), "Downloads"))
        ),
        "csv_glob": cfg.get("csv_glob", "ENOVIA*Export*.csv"),
        "index_enabled": bool(cfg.get("index_enabled", True)),
        "include_default": bool(cfg.get("include_in_index_default", True)),
        "automation_enabled": bool(cfg.get("automation_enabled", False)),
    }


# ---------- fallback user-prefs helpers (used by panes if main window not available) ----------


def _local_data_dir():
    base = os.environ.get("LOCALAPPDATA", os.path.expanduser(r"~\\AppData\\Local"))
    path = os.path.join(base, "PartSearch")
    os.makedirs(path, exist_ok=True)
    return path


def _user_prefs_path():
    return os.path.join(_local_data_dir(), "user_prefs.json")


# ---------- Index DB helpers (single source of truth) ----------


def index_db_path(root: str) -> str:
    """Per-root DB path under %LOCALAPPDATA%\\PartSearch\\index\\<hash>\\index_db.sqlite (canonicalized)."""
    idx_dir = os.path.join(_local_data_dir(), "index")
    # normalize the key so C:\, c:\, c:\\, C:/ all hash the same
    try:
        key = os.path.normcase(os.path.normpath((root or "").strip()))
        # keep directory roots with a trailing separator so 'C:' != 'C:\'
        if key and not key.endswith(os.sep):
            # add a sep only for drive roots like 'C:' or top-level dirs
            if len(key) == 2 and key[1] == ":":
                key = key + os.sep
    except Exception:
        key = (root or "").strip()
    h = hashlib.sha1(key.encode("utf-8", "ignore")).hexdigest()[:12]
    dbp = os.path.join(idx_dir, h, "index_db.sqlite")
    os.makedirs(os.path.dirname(dbp), exist_ok=True)
    return dbp


# Back-compat alias
get_index_db_path = index_db_path


def _ensure_meta_schema(db_path: str):
    """
    Ensure a meta table exists and has a 'value' column.
    If only 'val' exists, add 'value' and copy over.
    """
    with open_sqlite(db_path) as c:
        cur = c.cursor()
        cur.execute("CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY)")
        # inspect columns
        cols = {row[1] for row in cur.execute("PRAGMA table_info(meta)")}
        changed = False
        if "val" in cols and "value" not in cols:
            cur.execute("ALTER TABLE meta ADD COLUMN value TEXT")
            # copy existing values if present
            cur.execute("UPDATE meta SET value = val WHERE value IS NULL OR value = ''")
            changed = True
        if "value" not in cols and "val" not in cols:
            # fresh table: add value column
            cur.execute("ALTER TABLE meta ADD COLUMN value TEXT")
            changed = True
        if changed:
            c.commit()


def set_last_indexed_now_for_root(db_path: str, root: str) -> str:
    """Write/overwrite last_full_scan in meta, regardless of meta schema variant."""
    dbp = index_db_path(root)
    _ensure_meta_schema(dbp)
    with open_sqlite(dbp) as c:
        cur = c.cursor()
        # prefer 'value'; keep 'val' updated too if that column exists
        cols = {row[1] for row in cur.execute("PRAGMA table_info(meta)")}
        if "value" in cols:
            cur.execute(
                "INSERT INTO meta(key,value) VALUES('last_full_scan', datetime('now','localtime'))"
            )
        if "val" in cols:
            cur.execute(
                "INSERT INTO meta(key,val) VALUES('last_full_scan', datetime('now',localtime))"
            )
        c.commit()


"""def set_last_indexed_now_for_root(db_path: str, root: str):
    with open_sqlite(db_path) as con:
        con.execute(
            "INSERT INTO meta(key,val) VALUES('last_full_scan', datetime('now','localtime')) "
            "ON CONFLICT(key) DO UPDATE SET val=excluded.val"
        )
        con.commit()"""


def get_last_indexed_text_for_root(root: str) -> str:
    """Read a friendly last-indexed string (supports 'value' or 'val' columns, and two keys)."""
    dbp = index_db_path(root)
    if not os.path.exists(dbp):
        return "Never"
    try:
        with open_sqlite_ro(dbp) as c:
            cur = c.cursor()
            # try value first
            row = cur.execute(
                "SELECT value FROM meta WHERE key IN ('last_full_scan','last_indexed') LIMIT 1"
            ).fetchone()
            if row and row[0]:
                return row[0]
            # fallback to val
            row = cur.execute(
                "SELECT val FROM meta WHERE key IN ('last_full_scan','last_indexed') LIMIT 1"
            ).fetchone()
            return row[0] if row and row[0] else "Never"
    except Exception:
        return "Never"


def quickindex_rebuild_fts(db_path: str) -> None:
    """Rebuild FTS index in the given DB; safe no-op if FTS not present."""
    try:
        with open_sqlite(db_path) as c:
            c.execute("INSERT INTO files_fts(files_fts) VALUES('rebuild')")
            c.commit()
    except sqlite3.OperationalError:
        pass  # FTS table not present yet — ignore


def _qi_path():
    return os.path.join(
        os.environ.get("LOCALAPPDATA", os.path.expanduser(r"~\AppData\Local")),
        "PartSearch",
        "quick_index.sqlite",
    )


def open_sqlite(db_path: str) -> sqlite3.Connection:
    con = sqlite3.connect(db_path, timeout=30)  # wait up to 30s if locked
    con.execute("PRAGMA busy_timeout = 10000")  # 10s internal wait
    # fewer locks; safe for readers
    con.execute("PRAGMA journal_mode = WAL")
    con.execute("PRAGMA synchronous = NORMAL")
    con.execute("PRAGMA temp_store = MEMORY")
    return con


def open_sqlite_ro(db_path: str) -> sqlite3.Connection:
    """
    Read-only connection – safe while indexer is writing.
    """
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True, timeout=30)
    con.execute("PRAGMA query_only = 1")
    con.execute("PRAGMA busy_timeout = 10000")
    return con


def ensure_quick_index_db(db_path: str) -> sqlite3.Connection:
    """
    Ensure per-root QuickIndex DB exists and has the expected schema.
    Auto-migrates older 'files' tables that lack newer columns (root, parent, ctime, is_dir).
    Returns an OPEN connection with WAL + busy timeouts (via open_sqlite).
    """
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    con = open_sqlite(db_path)
    cur = con.cursor()

    # 1) meta table: keep both 'val' and 'value' for legacy code
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS meta(
          key   TEXT PRIMARY KEY,
          val   TEXT,
          value TEXT
        )
    """
    )

    # 2) files table (new schema)
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS files(
          id       INTEGER PRIMARY KEY,
          root     TEXT,               -- drive or logical root label
          path     TEXT NOT NULL,      -- full path
          name     TEXT,               -- file/folder name
          ext      TEXT,               -- lowercased extension (no dot)
          size     INTEGER,            -- bytes (NULL for dirs)
          mtime    REAL,               -- last write time (epoch)
          ctime    REAL,               -- creation time (epoch)
          is_dir   INTEGER NOT NULL DEFAULT 0,
          parent   TEXT,               -- parent directory
          pass_id  INTEGER
        )
    """
    )

    # 2a) MIGRATION: add missing columns if coming from older schema
    cur.execute("PRAGMA table_info(files)")
    cols = {r[1] for r in cur.fetchall()}
    to_add = []
    if "root" not in cols:
        to_add.append(("root", "TEXT"))
    if "parent" not in cols:
        to_add.append(("parent", "TEXT"))
    if "ctime" not in cols:
        to_add.append(("ctime", "REAL"))
    # legacy used 'dir' instead of 'is_dir'
    if "is_dir" not in cols and "dir" not in cols:
        to_add.append(("is_dir", "INTEGER NOT NULL DEFAULT 0"))
    for col, ddl in to_add:
        try:
            cur.execute(f"ALTER TABLE files ADD COLUMN {col} {ddl}")
        except sqlite3.OperationalError:
            pass  # already added or locked; continue

    # If legacy 'dir' column exists but 'is_dir' doesn't, mirror it once
    if "dir" in cols and "is_dir" not in cols:
        try:
            cur.execute(
                "ALTER TABLE files ADD COLUMN is_dir INTEGER NOT NULL DEFAULT 0"
            )
            cur.execute(
                "UPDATE files SET is_dir=CASE WHEN dir IS NULL THEN 0 ELSE dir END"
            )
        except sqlite3.OperationalError:
            pass

    # 3) indices — prefer uniqueness by (root, path) so one DB can hold multiple logical roots if desired
    cur.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS ux_files_root_path ON files(root, path)"
    )
    cur.execute("CREATE INDEX IF NOT EXISTS ix_files_name   ON files(name)")
    cur.execute("CREATE INDEX IF NOT EXISTS ix_files_ext    ON files(ext)")
    cur.execute("CREATE INDEX IF NOT EXISTS ix_files_parent ON files(parent)")
    cur.execute("CREATE INDEX IF NOT EXISTS ix_files_isdir  ON files(is_dir)")
    cur.execute("CREATE INDEX IF NOT EXISTS ix_files_pass   ON files(pass_id)")
    cur.execute(
        "CREATE INDEX IF NOT EXISTS ix_files_name_nocase ON files(name COLLATE NOCASE)"
    )

    # 4) optional FTS (best-effort)
    try:
        cur.execute(
            """
            CREATE VIRTUAL TABLE IF NOT EXISTS files_fts
            USING fts5(name, path, content='files', content_rowid='id', tokenize='unicode61');
        """
        )
        cur.execute(
            """
            CREATE TRIGGER IF NOT EXISTS files_ai AFTER INSERT ON files
            BEGIN
              INSERT INTO files_fts(rowid, name, path) VALUES (new.id, new.name, new.path);
            END;
        """
        )
        cur.execute(
            """
            CREATE TRIGGER IF NOT EXISTS files_au AFTER UPDATE ON files
            BEGIN
              INSERT INTO files_fts(files_fts, rowid, name, path) VALUES ('delete', old.id, old.name, old.path);
              INSERT INTO files_fts(rowid, name, path) VALUES (new.id, new.name, new.path);
            END;
        """
        )
        cur.execute(
            """
            CREATE TRIGGER IF NOT EXISTS files_ad AFTER DELETE ON files
            BEGIN
              INSERT INTO files_fts(files_fts, rowid, name, path) VALUES ('delete', old.id, old.name, old.path);
            END;
        """
        )
    except sqlite3.OperationalError:
        # FTS5 not compiled; that's fine
        pass

    con.commit()
    return con


def _fts_build_query(user_q: str, ordered: bool = False) -> str:
    """Turn 'coating* removal*' into 'coating* AND removal*' (FTS5 MATCH string).
    If ordered=True, treat as a phrase: '"coating removal"'.
    """
    q = (user_q or "").strip().lower()
    if not q:
        return ""
    tokens = [t for t in re.split(r"\s+", q) if t]
    if not tokens:
        return ""
    if ordered:
        phrase = " ".join([t.rstrip("*") for t in tokens])
        return f'"{phrase}"'
    parts = []
    for t in tokens:
        t = t.strip('"')
        if not t or t == "*":
            continue
        parts.append(t)  # FTS5 supports trailing * for prefix
    return " AND ".join(parts) if parts else ""


def get_index_db_path_for_root(root: str) -> str:
    """
    Return the per-root DB path if your app uses one-DB-per-root; otherwise fall back
    to the single quick_index.sqlite if that’s your current layout.
    """
    # Try per-root layout
    idx_dir = os.path.join(
        os.environ.get("LOCALAPPDATA", os.path.expanduser(r"~\AppData\Local")),
        "PartSearch",
        "index",
    )
    os.makedirs(idx_dir, exist_ok=True)
    h = hashlib.sha1(root.encode("utf-8", errors="ignore")).hexdigest()[:12]
    per_root = os.path.join(idx_dir, h, "index_db.sqlite")
    if os.path.exists(per_root):
        return per_root

    # Fallback: single quick index (if that’s what you built)
    qi = os.path.join(
        os.environ.get("LOCALAPPDATA", os.path.expanduser(r"~\AppData\Local")),
        "PartSearch",
        "quick_index.sqlite",
    )
    return qi  # may or may not exist yet


def _find_row_for_root(self, root: str):
    for r in range(self.table.rowCount()):
        item = self.table.item(r, 0)
        if item and item.text() == root:
            return r
    return None


# --- PN Search helpers ---


class DenodoTransientError(RuntimeError):
    pass


def denodo_fetch_all_safe(
    url,
    headers,
    params,
    *,
    max_retries=3,
    timeout=(5, 60),  # (connect, read) seconds
    url_len_limit=7000,
):
    """
    Robust GET against Denodo:
      - guards against monster URLs (use chunking for huge IN(...))
      - retries transient 5xx / Oracle init / connection reset with backoff+jitter
      - re-creates the session once if the socket is closed remotely
      - returns a normalized DataFrame (or empty DataFrame on odd payloads)
    """
    sess = _denodo_session()
    req = requests.Request("GET", url, headers=headers, params=params)
    prepped = sess.prepare_request(req)

    # Large-URL guard (typical failure mode with big BOM IN-lists)
    if len(prepped.url or "") > url_len_limit:
        raise RuntimeError(
            _with_tip(
                "Query URL is very large; split your IN(...) into chunks (see _fetch_*_chunked)."
            )
        )

    # Retry loop
    for attempt in range(max_retries):
        try:
            resp = sess.send(prepped, verify=False, timeout=timeout)
            resp.raise_for_status()
            data = resp.json()
            # normalize payload -> DataFrame
            try:
                df = pd.json_normalize(pd.json_normalize(data)["elements"][0])
            except Exception:
                df = pd.json_normalize(data)
            return normalize_keys(
                df if isinstance(df, pd.DataFrame) else pd.DataFrame()
            )

        except requests.exceptions.ConnectionError as e:
            # recreate the session once if remote closed the connection
            if attempt == 0 and (
                "RemoteDisconnected" in str(e) or "closed connection" in str(e)
            ):
                try:
                    global _D_SESSION
                    _D_SESSION = None
                    sess = _denodo_session()
                    prepped = sess.prepare_request(req)
                    continue  # try again immediately with new session
                except Exception:
                    pass
            if attempt < max_retries - 1:
                time.sleep((2**attempt) + random.random())
                continue
            raise RuntimeError(_with_tip(f"Denodo connection error: {e}")) from e

        except requests.HTTPError as e:
            body = ""
            try:
                body = e.response.text
            except Exception:
                pass
            status = e.response.status_code if e.response is not None else None
            # treat 5xx / Oracle init / connection errors as transient
            transient = (
                (status and 500 <= status <= 599)
                or ("ORA-01033" in body)
                or ("CONNECTION_ERROR" in body)
            )
            if transient and attempt < max_retries - 1:
                time.sleep((2**attempt) + random.random())
                continue
            # non-transient (or retries exhausted)
            raise RuntimeError(
                _with_tip(
                    f"Denodo request failed: {e}\nURL: {url}\nParams: {params}\nBody: {body}"
                )
            ) from e

        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep((2**attempt) + random.random())
                continue
            raise RuntimeError(_with_tip(f"Denodo request error: {e}")) from e


def denodo_fetch_all(url, headers, params):
    # legacy alias; keep for backward-compat
    return denodo_fetch_all_safe(url, headers, params)


def _fetch_in_chunks(view_name, id_col, ids, headers, select_cols, chunk_size=150):
    frames = []
    for i in range(0, len(ids), chunk_size):
        batch = ids[i : i + chunk_size]
        ids_in = ", ".join("'" + x.replace("'", "''") + "'" for x in batch)
        params = {
            "$select": ",".join(select_cols),
            "$filter": f"({id_col} IN ({ids_in}))",
        }
        df = denodo_fetch_all_safe(f"{BASE_URL}/{view_name}", headers, params)
        df = normalize_keys(df if isinstance(df, pd.DataFrame) else pd.DataFrame())
        frames.append(df)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def extract_pn_candidates(text: str):
    c = set()
    for pat in PN_REGEXES:
        for m in re.findall(pat, text or "", flags=re.I):
            c.add(m.upper())
    for tok in re.split(r"[^\w\-\.]+", text or ""):
        t = tok.strip().upper()
        if t and any(ch.isdigit() for ch in t) and len(t) >= 4:
            c.add(t)
    return sorted(c, key=lambda s: (-len(s), sum(ch.isdigit() for ch in s)))


def choose_one_pn(parent, candidates):
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    pn, ok = QInputDialog.getItem(
        parent, "Pick PN", "Multiple PN candidates:", candidates, 0, False
    )
    return pn if ok else None


def file_search_enabled() -> bool:
    """Read policy gate; if anything fails, default to enabled."""
    try:
        pdir = os.environ.get("PROGRAMDATA", r"C:\ProgramData")
        p = os.path.join(pdir, "PartSearch", "policy.json")
        with open(p, "r", encoding="utf-8") as f:
            pol = json.load(f)
        return bool(pol.get("file_search_enabled", True))
    except Exception:
        return True


# ------------ Location picking helpers (module level) ------------


def _list_candidate_roots():
    """
    Discover useful starting points: fixed drives, mapped network drives,
    and common OneDrive/SharePoint sync folders.
    """
    roots = []

    # 1) Logical Windows drives (fixed + network)
    # DRIVE types
    DRIVE_UNKNOWN = 0
    DRIVE_NO_ROOT_DIR = 1
    DRIVE_REMOVABLE = 2
    DRIVE_FIXED = 3
    DRIVE_REMOTE = 4
    kernel32 = ctypes.windll.kernel32
    buf = ctypes.create_unicode_buffer(256)
    kernel32.GetLogicalDriveStringsW(ctypes.sizeof(buf) // 2, buf)
    for d in buf.value.split("\x00"):
        if not d:
            continue
        try:
            dtype = kernel32.GetDriveTypeW(d)
        except Exception:
            dtype = DRIVE_UNKNOWN
        if dtype in (DRIVE_FIXED, DRIVE_REMOTE):
            roots.append(d)

    # 2) OneDrive / SharePoint synced folders under user profile
    home = os.path.expanduser("~")
    candidates = []
    # Typical OneDrive env
    if os.environ.get("OneDrive"):
        candidates.append(os.environ["OneDrive"])
    # Common OneDrive org folders: "OneDrive - OrgName"
    candidates += glob.glob(os.path.join(home, "OneDrive*"))
    # SharePoint libraries via OneDrive client tend to live under the org folder too
    candidates += glob.glob(os.path.join(home, "*SharePoint*"))
    for c in candidates:
        if os.path.isdir(c):
            roots.append(c)

    # 3) De-dup + keep order
    seen = set()
    uniq = []
    for r in roots:
        r = os.path.normpath(r)
        if r not in seen:
            seen.add(r)
            uniq.append(r)
    return uniq


# ===== FileSearch: Locations filter dialog =====
class SearchLocationsDialog(QDialog):
    """
    Filter dialog for FileSearchPane. You can only pick from roots that are
    already registered in QuickIndex (IndexBuilderPane). No adding here.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Choose locations to search")
        self.resize(700, 420)
        self._build_ui()
        self._load_roots()

    def _json_dir(self):
        return os.path.dirname(self._locations_path())

    def _index_roots_path(self):
        return str(app_paths()["index_roots"])

    def _locations_path(self):
        return str(app_paths()["locations"])

    def _load_json(self, path, default):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return default

    def _save_json(self, path, obj):
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(obj, f, indent=2)
        os.replace(tmp, path)

    def _build_ui(self):
        layout = QVBoxLayout(self)

        self.table = QTableWidget(0, 4, self)
        self.table.setHorizontalHeaderLabels(["", "Path", "Files", "Last Indexed"])
        self.table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeToContents
        )
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeToContents
        )
        self.table.horizontalHeader().setSectionResizeMode(
            3, QHeaderView.ResizeToContents
        )
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table)

        buttons = QHBoxLayout()
        btn_all = QPushButton("Check All")
        btn_none = QPushButton("Uncheck All")
        btn_all.clicked.connect(self._check_all)
        btn_none.clicked.connect(self._uncheck_all)
        buttons.addWidget(btn_all)
        buttons.addWidget(btn_none)
        buttons.addStretch(1)
        layout.addLayout(buttons)

        bb = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel, parent=self
        )
        bb.accepted.connect(self._accept)
        bb.rejected.connect(self.reject)
        layout.addWidget(bb)

    def _load_roots(self):
        self.table.setRowCount(0)
        index_roots = self._load_json(self._index_roots_path(), {"roots": []}).get(
            "roots", []
        )
        checked = set(
            self._load_json(self._locations_path(), {"checked_roots": []}).get(
                "checked_roots", []
            )
        )

        def add_row(path, files, last):
            r = self.table.rowCount()
            self.table.insertRow(r)
            chk = QTableWidgetItem()
            chk.setFlags(chk.flags() | Qt.ItemIsUserCheckable)
            chk.setCheckState(Qt.Checked if path in checked else Qt.Unchecked)
            self.table.setItem(r, 0, chk)
            self.table.setItem(r, 1, QTableWidgetItem(path))
            self.table.setItem(r, 2, QTableWidgetItem(f"{int(files):,}"))
            self.table.setItem(r, 3, QTableWidgetItem(last or ""))

        for rec in index_roots:
            add_row(
                rec.get("path", ""),
                rec.get("files_count", 0),
                rec.get("last_full_scan", ""),
            )

    def _check_all(self):
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            if it:
                it.setCheckState(Qt.Checked)

    def _uncheck_all(self):
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            if it:
                it.setCheckState(Qt.Unchecked)

    def _accept(self):
        chosen = []
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            if it and it.checkState() == Qt.Checked:
                path = self.table.item(r, 1).text().strip()
                if path:
                    chosen.append(path)
        self._save_json(self._locations_path(), {"checked_roots": chosen})
        self.accept()


class LocationPickerDialog(QDialog):
    def __init__(self, parent=None, initial=None):
        super().__init__(parent)
        self.setWindowTitle("Choose locations to search")
        self.resize(600, 420)
        self._initial = set(os.path.normpath(p) for p in (initial or []))

        v = QVBoxLayout(self)

        v.addWidget(QLabel("Select the drives/folders to include:", self))
        self.listw = QListWidget(self)
        self.listw.setSelectionMode(QListWidget.NoSelection)
        v.addWidget(self.listw)

        # Buttons row
        h = QHBoxLayout()
        self.btn_all = QPushButton("Check all", self)
        self.btn_none = QPushButton("Uncheck all", self)
        h.addStretch(1)
        h.addWidget(self.btn_all)
        h.addWidget(self.btn_none)
        v.addLayout(h)

        # OK/Cancel
        h2 = QHBoxLayout()
        h2.addStretch(1)
        self.btn_ok = QPushButton("OK", self)
        self.btn_cancel = QPushButton("Cancel", self)
        h2.addWidget(self.btn_ok)
        h2.addWidget(self.btn_cancel)
        v.addLayout(h2)

        # Populate
        for root in _list_candidate_roots():
            self._add_item(root, checked=(os.path.normpath(root) in self._initial))

        # Wire
        self.btn_all.clicked.connect(lambda: self._set_all(True))
        self.btn_none.clicked.connect(lambda: self._set_all(False))
        self.btn_ok.clicked.connect(self.accept)
        self.btn_cancel.clicked.connect(self.reject)

    def _add_item(self, path, checked=False):
        it = QListWidgetItem(path)
        it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
        it.setCheckState(Qt.Checked if checked else Qt.Unchecked)
        self.listw.addItem(it)

    def _set_all(self, checked: bool):
        for i in range(self.listw.count()):
            it = self.listw.item(i)
            it.setCheckState(Qt.Checked if checked else Qt.Unchecked)

    def _on_add(self):
        p = QFileDialog.getExistingDirectory(self, "Add folder")
        if p:
            p = os.path.normpath(p)
            # Avoid duplicates
            for i in range(self.listw.count()):
                if os.path.normpath(self.listw.item(i).text()) == p:
                    self.listw.item(i).setCheckState(Qt.Checked)
                    return
            self._add_item(p, checked=True)

    def selected_paths(self):
        out = []
        for i in range(self.listw.count()):
            it = self.listw.item(i)
            if it.checkState() == Qt.Checked:
                out.append(os.path.normpath(it.text()))
        return out


def choose_search_locations(parent=None, initial=None):
    """
    Show the picker and return a list of selected roots.
    If the user cancels, return the 'initial' list unchanged.
    """
    dlg = LocationPickerDialog(parent=parent, initial=initial)
    if dlg.exec_() == QDialog.Accepted:
        sel = dlg.selected_paths()
        if not sel:
            # Be kind: confirm empty selection so user doesn't think it's broken
            QMessageBox.information(
                parent,
                "Locations",
                "No locations selected; all locations will be searched.",
            )
        return sel
    return initial or []


# --- File Search Helpers ---


def _index_base_dir():
    p = os.path.join(
        os.environ.get("LOCALAPPDATA", os.path.expanduser("~\\AppData\\Local")),
        "PartSearch",
        "index",
    )
    os.makedirs(p, exist_ok=True)
    return p


def ensure_index_db(dbpath: str):
    """
    Ensure the per-root quick-index database exists with the right schema.
    Also makes the 'meta' table compatible with both 'val' and 'value' writes.
    """
    os.makedirs(os.path.dirname(dbpath), exist_ok=True)
    conn = open_sqlite_ro(dbpath)  # <-- use the actual parameter name
    try:
        cur = conn.cursor()
        cur.executescript(
            """
            CREATE TABLE IF NOT EXISTS files(
                id INTEGER PRIMARY KEY,
                path TEXT UNIQUE,
                name TEXT,
                ext TEXT,
                size INTEGER,
                mtime REAL,
                dir INTEGER DEFAULT 0,
                pass_id INTEGER
            );
            CREATE VIRTUAL TABLE IF NOT EXISTS files_fts USING fts5(
                name, path, ext, content='files', content_rowid='id'
            );
            /* include both val and value to avoid 'no column named value' */
            CREATE TABLE IF NOT EXISTS meta(
                key TEXT PRIMARY KEY,
                val   TEXT,
                value TEXT
            );
        """
        )
        cur.execute("CREATE INDEX IF NOT EXISTS idx_files_name ON files(name);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_files_ext  ON files(ext);")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_files_dir  ON files(dir);")
        conn.commit()
    finally:
        conn.close()


def get_last_indexed_text(root):
    db = get_index_db_path(root)  # your helper
    if not os.path.exists(db):
        return "Never"
    with open_sqlite_ro(db) as c:
        val = c.execute("SELECT value FROM meta WHERE key='last_indexed'").fetchone()
    return val[0] if val and val[0] else "Unknown"


def classify_ext(path: str) -> str:
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    if not ext:
        return "File"
    cad = {
        "sldprt": "CAD Part",
        "sldasm": "CAD Assembly",
        "slddrw": "CAD Drawing",
        "step": "STEP",
        "stp": "STEP",
        "iges": "IGES",
        "igs": "IGES",
        "dwg": "DWG",
        "dxf": "DXF",
    }
    docs = {
        "pdf": "PDF",
        "doc": "DOC",
        "docx": "DOCX",
        "xls": "XLS",
        "xlsx": "XLSX",
        "csv": "CSV",
        "txt": "Text",
    }
    pics = {
        "png": "Image",
        "jpg": "Image",
        "jpeg": "Image",
        "tif": "Image",
        "tiff": "Image",
        "bmp": "Image",
    }
    return cad.get(ext) or docs.get(ext) or pics.get(ext) or ext.upper()


# ===== Shared prefs + index helpers (kept consistent with your code) =====
# Keys:
#   "index_roots"       -> canonical set of indexed roots (QuickIndex pane owns this)
#   "file_search_roots" -> filter set to search (FileSearch pane owns this)
APP_NAME = "PartSearch"


def _appdata_dir() -> str:
    base = os.environ.get("APPDATA") or os.path.expanduser("~")
    p = os.path.join(base, APP_NAME)
    os.makedirs(p, exist_ok=True)
    return p


def _prefs_paths():
    """Return (machine_defaults_path, user_overrides_path)."""
    return (
        os.path.join(PROGRAMDATA_DIR, "user_prefs.json"),
        os.path.join(LOCALAPPDATA_DIR, "user_prefs.json"),
    )


def _read_json(path: str):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _deep_merge(a: dict, b: dict) -> dict:
    """Shallow for lists, recursive for dicts."""
    out = dict(a or {})
    for k, v in (b or {}).items():
        if isinstance(v, dict) and isinstance(out.get(k), dict):
            out[k] = _deep_merge(out[k], v)
        else:
            out[k] = v
    return out


def _prefs_read() -> dict:
    """
    Load preferences with layered fallback:
      1) DEFAULT_PREFS (built-in)
      2) machine defaults:  %PROGRAMDATA%\\PartSearch\\user_prefs.json  (optional)
      3) user overrides:    %LOCALAPPDATA%\\PartSearch\\user_prefs.json
    User overrides win.
    """
    machine_path, user_path = _prefs_paths()

    base = copy.deepcopy(DEFAULT_PREFS)
    machine = _read_json(machine_path) or {}
    user = _read_json(user_path) or {}

    prefs = _deep_merge(base, _deep_merge(machine, user))

    # normalize presence/types
    prefs.setdefault("default_bus", [])
    prefs.setdefault("filter_defaults", {})
    fd = prefs["filter_defaults"]
    fd.setdefault("min_qty", "")
    fd.setdefault("and_terms", [])
    fd.setdefault("or_terms", [])
    fd.setdefault("either_or_groups", [])
    fd.setdefault("include_unassigned", True)
    fd.setdefault("preferred_bu", "")

    prefs.setdefault("ui", {})
    prefs["ui"].setdefault("geometry", None)
    prefs["ui"].setdefault("last_tab", "psft")

    return prefs


def _prefs_write(prefs: dict) -> bool:
    """
    Write per-user preferences to %LOCALAPPDATA%\\PartSearch\\user_prefs.json atomically.
    Returns True on success.
    """
    _, user_path = _prefs_paths()
    ensure_dir(os.path.dirname(user_path))
    try:
        # keep only known sections to avoid dumping large runtime state
        to_store = {
            "default_bus": prefs.get("default_bus", []),
            "filter_defaults": prefs.get("filter_defaults", {}),
            "ui": prefs.get("ui", {}),
        }
        tmp = user_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(to_store, f, indent=2)
        os.replace(tmp, user_path)
        return True
    except Exception as e:
        print("prefs write error:", e)
        return False


def get_saved_roots(key: str) -> list:
    prefs = _prefs_read()
    roots = prefs.get(key, [])
    out, seen = [], set()
    for r in roots:
        n = os.path.normpath(r)
        if n not in seen:
            seen.add(n)
            out.append(n)
    return out


def set_saved_roots(key: str, roots: list):
    prefs = _prefs_read()
    out, seen = [], set()
    for r in roots or []:
        n = os.path.normpath(r)
        if n not in seen:
            seen.add(n)
            out.append(n)
    prefs[key] = out
    _prefs_write(prefs)


# ---- per-root index helpers (unified) ----
# Use the *canonical* index_db_path defined earlier in the file:
#   def index_db_path(root: str) -> str  (under %LOCALAPPDATA%\PartSearch\index\<hash>\index_db.sqlite)
# plus the meta helpers you already have:
#   set_last_indexed_now_for_root(root), get_last_indexed_text_for_root(root)


def indexed_file_count_for_root(root: str) -> int:
    try:
        dbp = index_db_path(root)
        if not os.path.exists(dbp):
            return 0
        con = open_sqlite_ro(dbp)
        try:
            row = con.execute("SELECT COUNT(*) FROM files").fetchone()
            return int(row[0] if row else 0)
        finally:
            con.close()
    except Exception:
        return 0


def last_full_scan_for_root(root: str) -> str | None:
    try:
        dbp = index_db_path(root)
        if not os.path.exists(dbp):
            return ""
        con = open_sqlite_ro(dbp)
        try:
            row = con.execute(
                "SELECT val FROM meta WHERE key='last_full_scan'"
            ).fetchone()
            if row and row[0]:
                return str(row[0])
            row = con.execute(
                "SELECT value FROM meta WHERE key='last_full_scan'"
            ).fetchone()
            return str(row[0]) if row and row[0] else ""
        finally:
            con.close()
    except Exception:
        return ""


# --- Enovia Search helpers ---


def enovia_db_path():
    base = os.path.join(
        os.environ.get("LOCALAPPDATA", os.path.expanduser("~\\AppData\\Local")),
        "PartSearch",
        "index",
    )
    os.makedirs(base, exist_ok=True)
    return os.path.join(base, "enovia_index.sqlite")


def ensure_enovia_db():
    p = enovia_db_path()
    conn = open_sqlite(p)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute(
        """
      CREATE TABLE IF NOT EXISTS enovia (
        id       TEXT,
        name     TEXT,
        type     TEXT,
        rev      TEXT,
        state    TEXT,
        modified TEXT,
        url      TEXT,
        last_seen REAL,
        PRIMARY KEY(url, name, type, rev)
      )
    """
    )
    conn.execute("""CREATE TABLE IF NOT EXISTS meta (key TEXT PRIMARY KEY, val TEXT)""")
    return conn


def enovia_ingest_csv(path: str):
    """Read an ENOVIA export CSV and upsert rows into the local ENOVIA index."""
    conn = ensure_enovia_db()
    now = time.time()
    try:
        try:
            df = pd.read_csv(path, encoding="utf-8", engine="python")
        except Exception:
            df = pd.read_csv(path, encoding="utf-8", engine="python", sep=";")

        def pick(*names):
            for n in names:
                if n in df.columns:
                    return df[n]
            return pd.Series([""] * len(df))

        recs = zip(
            pick("ID", "Object Id", "ObjectId", "OID").astype(str),
            pick(
                "Name", "Title", "File Name", "FileName", "DrawingNo", "Drawing No"
            ).astype(str),
            pick("Type", "Format", "Object Type").astype(str),
            pick("Revision", "Rev").astype(str),
            pick("State").astype(str),
            pick("Modified", "Last Modified", "LastModified").astype(str),
            pick("URL", "Link", "Href", "Download URL").astype(str),
        )
        cur = conn.cursor()
        for oid, name, typ, rev, state, mod, url in recs:
            cur.execute(
                """
              INSERT OR REPLACE INTO enovia(id,name,type,rev,state,modified,url,last_seen)
              VALUES (?,?,?,?,?,?,?,?)
            """,
                (
                    oid.strip(),
                    name.strip(),
                    typ.strip(),
                    rev.strip(),
                    state.strip(),
                    mod.strip(),
                    url.strip(),
                    now,
                ),
            )
        conn.execute(
            "INSERT OR REPLACE INTO meta(key,val) VALUES ('last_full_scan', ?)",
            (time.strftime("%Y-%m-%d %H:%M:%S"),),
        )
        conn.commit()
    finally:
        conn.close()


# --- Normalize incoming part numbers from CSV ---
SOLIDWORKS_EXTS = (".sldprt", ".sldasm", ".slddrw")


def normalize_part(p: str) -> str:
    """Uppercase PN and strip soldiworks extensions/whitespace"""
    s = str(p).strip()
    # remove SolidWorks file extensions (case-insensitive)
    s = re.sub(r"\.(sldprt|sldasm|slddrw)$", "", s, flags=re.IGNORECASE)
    # uppercase because the backend and joins are normalized to upper
    return s.upper()


def normalize_status_key(s) -> str:
    t = str(s or "").strip().upper()
    return "OK" if t.startswith("OK") else t


def normalize_bu(df):
    """Ensure a 'BU' string column is present; do NOT drop NULLs."""
    if "BU" not in df.columns and "business_unit" in df.columns:
        df = df.rename(columns={"business_unit": "BU"})
    df["BU"] = df.get("BU", "").astype(str)
    # keep blanks as blanks for filtering; you can label for display later
    df["BU"] = df["BU"].replace({"nan": "", "None": ""}).fillna("")
    return df


def apply_min_qty_filter(df, min_qty):
    """Blank => no filter; 0 => include NULL; >=1 => numeric cutoff."""
    q = pd.to_numeric(df.get("available_quantity"), errors="coerce")
    if min_qty is None:
        return df
    if min_qty == 0:
        return df[(q >= 0) | (q.isna())]
    return df[q >= int(min_qty)]


def apply_bu_filter(df, bus, include_unassigned=True):
    """If no BUs chosen => no filter; else allow blanks when requested."""
    if not bus:
        return df
    mask = df["BU"].str.upper().isin([b.upper() for b in bus])
    if include_unassigned:
        mask = mask | df["BU"].isna() | (df["BU"].astype(str).str.strip() == "")
    return df[mask]


def apply_qcode_filter(df: pd.DataFrame, qcodes: list[str] | None):
    if not qcodes:
        return df
    import re

    # normalize input & column
    qcodes = [q.strip().upper() for q in qcodes if str(q).strip()]
    series = df.get("q_codes", "").fillna("").astype(str).str.upper()
    # each requested code must appear (token-aware)
    mask = pd.Series(True, index=df.index)
    for qc in qcodes:
        pat = rf"(^|[ ,;|]){re.escape(qc)}([ ,;|]|$)"
        mask &= series.str.contains(pat, regex=True)
    return df[mask]


def q_like(field: str, value: str) -> str:
    v = "" if value is None else str(value)
    v = v.replace("'", "''")
    return f"{field} like '%{v}%'"


def sql_literal(v) -> str:
    s = "" if v is None else str(v)
    return "'" + s.replace("'", "''") + "'"


def _q_escape(s: str) -> str:
    return str(s).replace("'", "''")


def q_like_ci(field: str, term: str) -> str:
    """
    Case-insensitive LIKE with user wildcards:
    * -> %   and   ? -> _
    Always does a contains match unless the user already supplied %/_.
    """
    raw = (_q_escape(term or "")).strip()
    # translate UI wildcards to SQL LIKE
    pat = raw.replace("*", "%").replace("?", "_")
    # if the user didn't include any %/_ at all, do a contains search
    if "%" not in pat and "_" not in pat:
        pat = f"%{pat}%"
    return f"upper({field}) LIKE '{pat.upper()}'"


def q_or_like_ci(field: str, terms) -> str:
    terms = [t for t in (terms or []) if str(t).strip()]
    if not terms:
        return ""
    return "(" + " OR ".join(q_like_ci(field, t) for t in terms) + ")"


def q_and_like_ci(field: str, terms) -> str:
    terms = [t for t in (terms or []) if str(t).strip()]
    if not terms:
        return ""
    return " AND ".join(q_like_ci(field, t) for t in terms)


def q_either_or_ci(field: str, groups) -> str:
    parts = []
    for grp in groups or []:
        grp = [t for t in grp if str(t).strip()]
        if grp:
            parts.append("(" + " OR ".join(q_like_ci(field, t) for t in grp) + ")")
    return " AND ".join(parts)


def q_in_ci(field: str, seq) -> str:
    vals = [f"'{_q_escape(x).upper()}'" for x in (seq or []) if str(x).strip()]
    return f"upper({field}) in ({', '.join(vals)})" if vals else ""


def _ensure_tuple3(result):
    """Normalize items_lookup outputs to a (main, per_loc, inv_raw) triple."""
    import pandas as pd

    empty = pd.DataFrame()
    if isinstance(result, tuple):
        # pad or trim to exactly 3
        r = list(result)[:3]
        while len(r) < 3:
            r.append(empty)
        return tuple(r)
    elif isinstance(result, pd.DataFrame):
        return (result, empty, empty)
    else:
        return (empty, empty, empty)


def build_itemmaster_filter_parts(
    item_ids=None,
    mfg_parts=None,
    and_wild=None,
    either_or=None,
    or_wild=None,  # repurposed in UI for QC; ignore here on purpose
    **_compat,  # swallow legacy kw names
):
    """
    Back-compat: returns THREE lists of SQL predicate strings:
      (desc_bits, id_bits, mfg_bits)
    Older code expects this exact tuple and unpacks it.
    """

    # ---- legacy kw aliases ----
    if and_wild is None:
        and_wild = _compat.pop("and_wildcards", None)
    if or_wild is None:
        or_wild = _compat.pop("or_wildcards", None)  # intentionally unused
    if either_or is None:
        either_or = _compat.pop("either_or_wildcards", None)

    # ---- sanitize helpers ----
    def _nz_list(x):
        return [t for t in (x or []) if str(t).strip()]

    item_ids = _nz_list(item_ids)
    mfg_parts = _nz_list(mfg_parts)
    and_wild = _nz_list(and_wild)
    either_or = [
        [y for y in (grp or []) if str(y).strip()] for grp in (either_or or [])
    ]

    # ---- build the three groups ----
    # A) ID predicates
    id_bits = []
    if item_ids:
        likes = [x for x in item_ids if any(ch in x for ch in ("%", "_", "*", "?"))]
        eqs = [x for x in item_ids if not any(ch in x for ch in ("%", "_", "*", "?"))]
        if eqs:
            if len(eqs) == 1:
                id_bits.append(f"UPPER(item_id) = '{_q_escape(eqs[0]).upper()}'")
            else:
                id_bits.append(q_in_ci("item_id", eqs))
        if likes:
            id_bits.append(
                "(" + " OR ".join(q_like_ci("item_id", t) for t in likes) + ")"
            )

    # B) Manufacturer-part predicates
    mfg_bits = []
    if mfg_parts:
        if len(mfg_parts) == 1:
            mfg_bits.append(
                f"UPPER(manufacturer_part) = '{_q_escape(mfg_parts[0]).upper()}'"
            )
        else:
            mfg_bits.append(q_in_ci("manufacturer_part", mfg_parts))

    # C) Description predicates
    desc_bits = []
    if and_wild:
        desc_bits.append(q_and_like_ci("item_description", and_wild))
    if either_or:
        grp = q_either_or_ci("item_description", either_or)
        if grp:
            desc_bits.append(grp)

    # NOTE: 'or_wild' (old OR wildcards) is now repurposed for Quality Codes,
    # so we deliberately do NOT add description OR-clauses here.

    return desc_bits, id_bits, mfg_bits


def build_itemmaster_filter(
    item_ids=None,
    mfg_parts=None,
    and_wild=None,
    either_or=None,
    or_wild=None,
    **_compat,
) -> str:
    """
    Returns a SINGLE string: the combined WHERE clause for item_master.
    Kept in sync with *_parts above.
    Accepts both modern and legacy kw names.
    """
    desc_bits, id_bits, mfg_bits = build_itemmaster_filter_parts(
        item_ids=item_ids,
        mfg_parts=mfg_parts,
        and_wild=and_wild,
        either_or=either_or,
        or_wild=or_wild,
        **_compat,
    )
    where = [*desc_bits, *id_bits, *mfg_bits]
    return " AND ".join(w for w in where if w) or ""


# ---- Currency helpers ----


def fx_convert(
    amount: float, from_ccy: str, to_ccy: str, rates: dict[str, float]
) -> float:
    """
    rates are USD-per-unit, e.g., {'USD':1.0,'EUR':1.08,'NOK':0.094}.
    amount_in_to = (amount * USD_per_from) / USD_per_to
    """
    if amount is None:
        return 0.0
    from_ccy = (from_ccy or "USD").upper()
    to_ccy = (to_ccy or "USD").upper()
    usd_per_from = float(rates.get(from_ccy, 1.0) or 1.0)
    usd_per_to = float(rates.get(to_ccy, 1.0) or 1.0)
    if usd_per_to <= 0:
        usd_per_to = 1.0
    return float(amount) * usd_per_from / usd_per_to


# ---- Build Filters & Equivalent SQL ----


def build_item_filters(self) -> str:
    """
    Build WHERE clause fragments for PSFT/Denodo.
    - No BU selected => no BU predicate (NULL/blank BUs are included)
    - BU selected + 'Include unassigned' checked => selected BUs OR NULL/blank
    - Min Qty blank => no qty filter
    - Min Qty = 0   => include NULL qty (unknown) as well
    """
    parts = []

    # BU predicate
    bu_list = self._get_checked_bus() if hasattr(self, "_get_checked_bus") else []
    include_unassigned = bool(
        getattr(self, "chk_include_unassigned", None)
        and self.chk_include_unassigned.isChecked()
    )
    if bu_list:
        in_list = ", ".join("'" + b.replace("'", "''").upper() + "'" for b in bu_list)
        clause = f"UPPER(business_unit) IN ({in_list})"
        if include_unassigned:
            clause = f"({clause} OR business_unit IS NULL OR TRIM(business_unit)='')"
        parts.append(clause)

    # Min qty (blank = ignore; 0 = include NULLs; N => >= N)
    txt = (self.qty.text() if hasattr(self, "qty") else "").strip()
    if txt != "":
        try:
            m = int(txt)
            if m == 0:
                parts.append("(available_quantity >= 0 OR available_quantity IS NULL)")
            else:
                parts.append(f"available_quantity >= {m}")
        except ValueError:
            pass

    return " AND ".join(parts)


def build_inv_filters(
    item_ids: list[str], bus: list[str], include_unassigned: bool = False
) -> str:
    parts = q_in_ci("inventory_item_id", item_ids)

    if bus:
        bu_clause = q_in_ci("business_unit", bus)
        if include_unassigned:
            bu_clause = (
                f"({bu_clause} OR business_unit IS NULL OR TRIM(business_unit) = '')"
            )
        bits = [x for x in (parts, bu_clause) if x]
    else:
        # No BU predicate at all lets NULL/blank BUs flow through
        bits = [x for x in (parts,) if x]

    return " AND ".join(bits)


def build_equivalent_sql(
    item_ids, mfg_parts, and_wild, either_or, or_wild, bus, min_qty
):
    def esc(x):
        return str(x).replace("'", "''")

    where = []

    # IM predicates
    if item_ids:
        like_items = [x for x in item_ids if any(ch in x for ch in ["%", "_"])]
        eq_items = [x for x in item_ids if not any(ch in x for ch in ["%", "_"])]
        if len(eq_items) == 1:
            where.append(f"IM.ITEM_ID = '{esc(eq_items[0])}'")
        elif eq_items:
            where.append(
                "IM.ITEM_ID IN (" + ", ".join(f"'{esc(x)}'" for x in eq_items) + ")"
            )
        if like_items:
            where.append(
                "("
                + " OR ".join(f"IM.ITEM_ID LIKE '{esc(x)}'" for x in like_items)
                + ")"
            )

    if mfg_parts:
        if len(mfg_parts) == 1:
            where.append(f"IM.manufacturer_part = '{esc(mfg_parts[0])}'")
        else:
            where.append(
                "IM.manufacturer_part IN ("
                + ", ".join(f"'{esc(x)}'" for x in mfg_parts)
                + ")"
            )

    if and_wild:
        where += [f"IM.ITEM_DESCRIPTION LIKE '%{esc(t)}%'" for t in and_wild if t]

    if either_or:
        for group in either_or:
            opts = [f"IM.ITEM_DESCRIPTION LIKE '%{esc(t)}%'" for t in group if t]
            if opts:
                where.append("(" + " OR ".join(opts) + ")")

    if or_wild:
        where += [
            f"UPPER(PI.item_field_c30_b) LIKE '%{esc(t).upper()}%'"
            for t in or_wild
            if t
        ]

    if bus:
        where.append(
            "INV.BUSINESS_UNIT IN (" + ", ".join(f"'{esc(b)}'" for b in bus) + ")"
        )

    # Min qty (None => no predicate)
    if min_qty is not None:
        where.append(f"INV.available_quantity >= {int(min_qty)}")

    where_sql = (" WHERE " + " AND ".join(where)) if where else ""

    select_sql = """
    SELECT DISTINCT
      INV.business_unit     AS BU,
      IM.item_id            AS part_number,
      IM.item_description   AS part_description,
      IM.revision           AS revision,
      IM.unit_of_measure    AS UOM,
      PI.item_field_c30_b   AS Q_CODES,
      IM.lot_control        AS LOT,
      IM.serial_control     AS SERIAL,
      INV.available_quantity,
      INV.onhand_quantity,
      INV.reserved_quantity,
      IM.manufacturer_id    AS mfg_id,
      IM.manufacturer_part  AS mfg_part,
      INV.unit_cost         AS perpetual_avg_cost,
      INV.currency_cd,
      IM.family             AS family,
      IM.source             AS source,
      IM.item_status,
      PI.item_field_c10_c   AS ctrl_in,
      IM.group_description  AS group_description
    FROM finance.item_master AS IM
    LEFT JOIN (
        SELECT DISTINCT
          INVENTORY_ITEM_ID,
          BUSINESS_UNIT,
          SUM(INS.onhand_quantity)
            - SUM(INS.open_po_qty + INS.invenotry_demand + INS.regional_demand
                  + INS.wo_demand + INS.pid_qty + INS.pid_demand + INS.qty_in_transit)
            AS available_quantity,
          SUM(INS.onhand_quantity) AS onhand_quantity,
          SUM(INS.open_po_qty + INS.invenotry_demand + INS.regional_demand
              + INS.wo_demand + INS.pid_qty + INS.pid_demand + INS.qty_in_transit)
            AS reserved_quantity,
          unit_cost,
          currency_cd
        FROM finance.inventory_search_all AS INS
        GROUP BY INVENTORY_ITEM_ID, BUSINESS_UNIT, UNIT_COST, CURRENCY_CD
    ) AS INV
      ON IM.item_id = INV.inventory_item_id
    LEFT JOIN finance.bv_psfinance_ps_master_item_tbl AS PI
      ON IM.item_id = PI.inv_item_id
    """
    order_sql = " ORDER BY IM.ITEM_ID, INV.BUSINESS_UNIT"
    return select_sql + where_sql + order_sql


def _with_tip(msg: str) -> str:
    m = str(msg)
    needle = ("RemoteDisconnected", "closed connection", "ChunkedEncodingError")
    if any(n in m for n in needle):
        m += (
            "\n\nTip: The server likely closed a very large response. "
            "Add one more wildcard (size/material/standard) or reduce selected BUs; "
            "you can also run the BOM path which is naturally narrower."
        )
    return m


# ---- Reach out to Denodo ----
_D_SESSION = None


def _denodo_session():
    global _D_SESSION
    if _D_SESSION is not None:
        return _D_SESSION
    s = requests.Session()
    retry = Retry(
        total=5,
        connect=5,
        read=5,
        backoff_factor=0.6,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(["GET", "POST"]),
        raise_on_status=False,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    s.headers.update(
        {
            "Accept": "application/json",
            "Connection": "keep-alive",
            "Accept-Encoding": "gzip, deflate",
        }
    )
    _D_SESSION = s
    return s


def normalize_keys(df: pd.DataFrame) -> pd.DataFrame:
    for c in df.columns:
        if df[c].dtype == object:
            # strip + uppercase keys we join on
            if c in (
                "item_id",
                "inventory_item_id",
                "business_unit",
                "manufacturer_part",
            ):
                df[c] = df[c].astype(str).str.strip().str.upper()
    return df


def aggregate_inventory(df_inv: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate inventory per (inventory_item_id, business_unit).
    Adds: min_qty (min_qty_reorder_point), unit_cost (perpetual_avg_cost), currency_cd.
    """
    cols_out = [
        "inventory_item_id",
        "business_unit",
        "available_quantity",
        "onhand_quantity",
        "reserved_quantity",
        "unit_cost",
        "currency_cd",
    ]

    if df_inv is None or df_inv.empty:
        return pd.DataFrame(columns=cols_out)

    df = df_inv.copy()

    for c in ("inventory_item_id", "business_unit"):
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip().str.upper()

    # Numerics
    num_cols = [
        "onhand_quantity",
        "open_po_qty",
        "invenotry_demand",
        "regional_demand",
        "wo_demand",
        "pid_qty",
        "pid_demand",
        "qty_in_transit",
        "min_qty",
        "unit_cost",
    ]
    for c in num_cols:
        if c not in df.columns:
            df[c] = 0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    if "currency_cd" not in df.columns:
        df["currency_cd"] = "USD"
    df["currency_cd"] = df["currency_cd"].astype(str).str.upper().fillna("USD")

    g = (
        df.groupby(["inventory_item_id", "business_unit"], dropna=False)
        .agg(
            {
                "onhand_quantity": "sum",
                "open_po_qty": "sum",
                "invenotry_demand": "sum",
                "regional_demand": "sum",
                "wo_demand": "sum",
                "pid_qty": "sum",
                "pid_demand": "sum",
                "qty_in_transit": "sum",
                "min_qty": "max",  # typical reorder point choice
                # simple mean (we’ll choose cheapest at view-time)
                "unit_cost": "mean",
                "currency_cd": "first",  # assume stable per BU
            }
        )
        .reset_index()
    )

    demand = g[["invenotry_demand", "invenotry_demand"]].max(axis=1)
    g["reserved_quantity"] = (
        g["open_po_qty"]
        + demand
        + g["regional_demand"]
        + g["wo_demand"]
        + g["pid_qty"]
        + g["pid_demand"]
        + g["qty_in_transit"]
    ).clip(lower=0)
    g["onhand_quantity"] = g["onhand_quantity"].clip(lower=0)
    g["available_quantity"] = (g["onhand_quantity"] - g["reserved_quantity"]).clip(
        lower=0
    )

    return g[
        [
            "inventory_item_id",
            "business_unit",
            "available_quantity",
            "onhand_quantity",
            "reserved_quantity",
            "min_qty",
            "unit_cost",
            "currency_cd",
        ]
    ]


def aggregate_inventory_per_loc(df_inv: pd.DataFrame) -> pd.DataFrame:
    """
    Optional per-location rollup. Returns empty if no 'location' column present.
    Returns columns:
      ['inventory_item_id','business_unit','location','available_quantity','onhand_quantity','reserved_quantity']
    """
    cols_out = [
        "inventory_item_id",
        "business_unit",
        "location",
        "available_quantity",
        "onhand_quantity",
        "reserved_quantity",
    ]

    if df_inv is None or df_inv.empty or "location" not in df_inv.columns:
        return pd.DataFrame(columns=cols_out)

    df = df_inv.copy()

    # Normalize key fields
    for c in ("inventory_item_id", "business_unit", "location"):
        if c in df.columns:
            df[c] = df[c].astype(str).str.strip().str.upper()

    # Numeric set (same as per-BU)
    num_cols = [
        "onhand_quantity",
        "open_po_qty",
        "invenotry_demand",
        "regional_demand",
        "wo_demand",
        "pid_qty",
        "pid_demand",
        "qty_in_transit",
    ]
    for c in num_cols:
        if c not in df.columns:
            df[c] = 0
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    g = (
        df.groupby(["inventory_item_id", "business_unit", "location"], dropna=False)[
            num_cols
        ]
        .sum()
        .reset_index()
    )

    demand = g[["inventory_demand", "inventory_demand"]].max(axis=1)

    g["reserved_quantity"] = (
        g["open_po_qty"]
        + demand
        + g["regional_demand"]
        + g["wo_demand"]
        + g["pid_qty"]
        + g["pid_demand"]
        + g["qty_in_transit"]
    ).clip(lower=0)

    g["onhand_quantity"] = g["onhand_quantity"].clip(lower=0)
    g["available_quantity"] = (g["onhand_quantity"] - g["reserved_quantity"]).clip(
        lower=0
    )

    return g[
        [
            "inventory_item_id",
            "business_unit",
            "location",
            "available_quantity",
            "onhand_quantity",
            "reserved_quantity",
        ]
    ]


def _fetch_inventory_chunked(ids_for_inv, headers, chunk_size=150):
    frames = []
    for i in range(0, len(ids_for_inv), chunk_size):
        batch = ids_for_inv[i : i + chunk_size]
        ids_in = ", ".join("'" + x.replace("'", "''") + "'" for x in batch)
        inv_params = {
            "$select": ",".join(
                [
                    "inventory_item_id",
                    "business_unit",
                    "onhand_quantity",
                    "open_po_qty",
                    "invenotry_demand",
                    "regional_demand",
                    "wo_demand",
                    "pid_qty",
                    "pid_demand",
                    "qty_in_transit",
                    "unit_cost",
                    "currency_cd",
                ]
            ),
            "$filter": f"(inventory_item_id IN ({ids_in}))",
        }
        tmp = denodo_fetch_all_safe(f"{BASE_URL}/{INV_VIEW}", headers, inv_params)
        frames.append(
            normalize_keys(tmp if isinstance(tmp, pd.DataFrame) else pd.DataFrame())
        )
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def items_lookup(
    BUs: list = None,
    min_qty: int | None = None,
    include_unassigned: bool = True,
    item_id: Optional[List[str]] = None,
    mfg_part: Optional[List[str]] = None,
    and_wildcards: Optional[List[str]] = None,
    or_wildcards: Optional[List[str]] = None,  # used as Q-CODE AND terms
    either_or_wildcards: Optional[List[List[str]]] = None,
    qty: int | None = None,  # kept for backward-compat; prefer min_qty
    keyring_service: str = "Denodo",
    keyring_username: Optional[str] = None,
    filter_positive_only: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    headers = get_basic_auth_header(keyring_service, keyring_username)

    # ---- sanitize inputs ----
    item_ids = [x.strip() for x in (item_id or []) if x and x.strip()]
    mfg_parts = [x.strip() for x in (mfg_part or []) if x and x.strip()]
    and_wild = [x.strip() for x in (and_wildcards or []) if x and x.strip()]
    either_or = [
        [y.strip() for y in grp if y and y.strip()]
        for grp in (either_or_wildcards or [])
    ]

    q_and_terms = [
        x.strip() for x in (or_wildcards or []) if x and x.strip()
    ]  # repurposed for Q-codes

    bus = [b.strip().upper() for b in (BUs or []) if b and b.strip()]

    eff_min_qty = min_qty if min_qty is not None else qty

    # ---- 1) build item_master WHERE from UI (no Q-codes here) ----
    im_where = build_itemmaster_filter(
        item_ids=item_ids,
        mfg_parts=mfg_parts,
        and_wild=and_wild,
        either_or=either_or,
        or_wild=None,
    )

    # ---- 2) Q-CODE AND prefilter -> list of matching item_ids (ctrl_ids) ----
    ctrl_ids: list[str] = []
    if q_and_terms:

        def esc(s: str) -> str:
            # escape single quotes and force upper for case-insensitive match
            return s.replace("'", "''").upper()

        # choose one of these patterns:
        # contains-match:
        clauses = [f"(UPPER(item_field_c30_b) LIKE '%{esc(t)}%')" for t in q_and_terms]
        # or, starts-with match:
        # clauses = [f"(UPPER(item_field_c30_b) LIKE '{esc(t)}%')" for t in q_and_terms]

        qfilter = " AND ".join(clauses)

        ctrl_params = {
            "$select": "inv_item_id,item_field_c30_b,item_field_c10_c",
            "$filter": qfilter,  # <-- no extra quotes around the whole clause
        }

        df_ctrl_prefilter = denodo_fetch_all_safe(
            f"{BASE_URL}/{ITEM_CONTROLS}", headers, ctrl_params
        )
        df_ctrl_prefilter = normalize_keys(
            df_ctrl_prefilter
            if isinstance(df_ctrl_prefilter, pd.DataFrame)
            else pd.DataFrame()
        )

        if df_ctrl_prefilter.empty or "inv_item_id" not in df_ctrl_prefilter.columns:
            # bail out cleanly (prevents IN ())
            return (
                (
                    pd.DataFrame(columns=DESIRED_HEADERS)
                    if "DESIRED_HEADERS" in globals()
                    else pd.DataFrame()
                ),
                pd.DataFrame(
                    columns=[
                        "inventory_item_id",
                        "business_unit",
                        "location",
                        "available_quantity",
                        "onhand_quantity",
                        "reserved_quantity",
                    ]
                ),
                pd.DataFrame(),
            )

        ctrl_ids = (
            df_ctrl_prefilter["inv_item_id"]
            .astype(str)
            .str.strip()
            .str.upper()
            .dropna()
            .unique()
            .tolist()
        )

        # If caller also supplied explicit item_ids (e.g., BOM import), intersect
        if item_ids:
            s = set(ctrl_ids)
            item_ids = [x for x in item_ids if x.strip().upper() in s]

        # Constrain item master with the ctrl_ids
        if ctrl_ids:
            in_list = ", ".join("'" + x.replace("'", "''") + "'" for x in ctrl_ids)
            im_where = (
                f"({im_where}) AND (item_id IN ({in_list}))"
                if im_where
                else f"(item_id IN ({in_list}))"
            )

    # ---- 3) fetch ITEM_MASTER (correct columns!) ----
    im_params = {
        "$select": ",".join(
            [
                "item_id",
                "item_description",
                "revision",
                "unit_of_measure",
                "manufacturer_id",
                "manufacturer_part",
                "family",
                "source",
                "item_status",
                "group_description",
                "lot_control",
                "serial_control",
            ]
        )
    }
    if im_where:
        im_params["$filter"] = im_where

    df_items = denodo_fetch_all_safe(f"{BASE_URL}/{ITEMS_VIEW}", headers, im_params)
    df_items = normalize_keys(
        df_items if isinstance(df_items, pd.DataFrame) else pd.DataFrame()
    )

    if df_items.empty or "item_id" not in df_items.columns:
        empty_main = (
            pd.DataFrame(columns=DESIRED_HEADERS)
            if "DESIRED_HEADERS" in globals()
            else pd.DataFrame()
        )
        empty_inv = pd.DataFrame(
            columns=[
                "inventory_item_id",
                "business_unit",
                "available_quantity",
                "onhand_quantity",
                "reserved_quantity",
                "unit_cost",
                "currency_cd",
            ]
        )
        return empty_main, empty_inv, pd.DataFrame()

    df_items["item_id"] = df_items["item_id"].astype(str).str.strip().str.upper()
    ids_from_items = df_items["item_id"].dropna().unique().tolist()
    ids_for_inv = ctrl_ids if ctrl_ids else ids_from_items
    if not ids_for_inv:  # avoid IN ()
        empty_main = (
            pd.DataFrame(columns=DESIRED_HEADERS)
            if "DESIRED_HEADERS" in globals()
            else pd.DataFrame()
        )
        empty_inv = pd.DataFrame(
            columns=[
                "inventory_item_id",
                "business_unit",
                "available_quantity",
                "onhand_quantity",
                "reserved_quantity",
                "min_qty_reorder_point",
                "unit_cost",
                "currency_cd",
            ]
        )
        return empty_main, empty_inv, pd.DataFrame()

    # ---- 4) fetch INVENTORY (no 'available_quantity' in $select) ----
    ids_in = ", ".join("'" + x.replace("'", "''") + "'" for x in ids_for_inv)
    tmp = _fetch_inventory_chunked(ids_for_inv, headers, chunk_size=150)
    df_inv_raw = normalize_keys(
        tmp if isinstance(tmp, pd.DataFrame) else pd.DataFrame()
    )
    df_inv_agg = aggregate_inventory(df_inv_raw)  # computes reserved & available

    # ---- 5) fetch CONTROLS for same ids (if not already taken from prefilter) ----
    if q_and_terms:
        df_ctrl = df_ctrl_prefilter[
            df_ctrl_prefilter["inv_item_id"].isin(ids_for_inv)
        ].copy()
    else:
        ctrl_params = {
            "$select": "inv_item_id,item_field_c30_b,item_field_c10_c",
            "$filter": f"(inv_item_id IN ({ids_in}))",
        }
        df_ctrl = denodo_fetch_all_safe(
            f"{BASE_URL}/{ITEM_CONTROLS}", headers, ctrl_params
        )
        df_ctrl = normalize_keys(
            df_ctrl if isinstance(df_ctrl, pd.DataFrame) else pd.DataFrame()
        )

    if not df_ctrl.empty and "inv_item_id" in df_ctrl.columns:
        df_ctrl = df_ctrl.rename(
            columns={
                "inv_item_id": "item_id",
                "item_field_c30_b": "q_codes",
                "item_field_c10_c": "ctrl_in",
            }
        )
        df_items = df_items.merge(df_ctrl, how="left", on="item_id")

    # ---- 6) merge + shape ----
    df = df_items.merge(
        df_inv_agg,
        how="left",
        left_on=["item_id"],
        right_on=["inventory_item_id"],
        suffixes=("", "_inv"),
    )
    if "inventory_item_id" in df.columns:
        df = df.drop(columns=["inventory_item_id"])

    if filter_positive_only and "available_quantity" in df.columns:
        df = df[df["available_quantity"].fillna(0) > 0]

    out = df.rename(
        columns={
            "business_unit": "BU",
            "item_id": "part_number",
            "item_description": "part_description",
            "unit_of_measure": "UOM",
            "manufacturer_id": "mfg_id",
            "manufacturer_part": "mfg_part",
            "unit_cost": "perpetual_avg_cost",
            "lot_control": "lot",
            "serial_control": "serial",
            "item_field_c30_b": "q_codes",
            "item_field_c10_c": "ctrl_in",
        }
    )

    # Normalize & apply GUI filters (reuse your helpers)
    out = normalize_bu(out)
    out = apply_qcode_filter(out, q_and_terms)
    out = apply_min_qty_filter(out, eff_min_qty)
    out = apply_bu_filter(out, bus, include_unassigned=bool(include_unassigned))

    # Final order if you maintain DESIRED_HEADERS
    if "DESIRED_HEADERS" in globals():
        out = out[[c for c in DESIRED_HEADERS if c in out.columns]]

    per_loc = aggregate_inventory_per_loc(df_inv_raw)
    return out.reset_index(drop=True), per_loc, df_inv_raw


def safe_reset_index(df: pd.DataFrame, name: str = "part_number") -> pd.DataFrame:
    """Bring the current index out as 'name' without colliding if it already exists."""
    idx_vals = df.index.to_numpy()
    df2 = df.reset_index(drop=True)
    if name not in df2.columns:
        df2.insert(0, name, idx_vals)
    return df2


# ---- UI & Excel colors for status ----


def pretty_sql(sql: str) -> str:
    """Lightweight SQL pretty-printer:
    - Newlines before major clauses
    - Indent AND/OR
    - Uppercase common keywords"""
    s = (sql or "").strip()
    if not s:
        return s

    # collapse whitespace
    s = re.sub(r"\s+", " ", s)

    # break before main clauses
    for pat in [
        r"SELECT\b",
        r"FROM\b",
        r"LEFT JOIN\b",
        r"RIGHT JOIN\b",
        r"INNER JOIN\b",
        r"FULL JOIN\b",
        r"WHERE\b",
        r"GROUP BY\b",
        r"HAVING\b",
        r"ORDER BY\b",
    ]:
        s = re.sub(pat, lambda m: "\n" + m.group(0).upper(), s, flags=re.IGNORECASE)

    # indent logical operators
    s = re.sub(r"\sAND\s", "\n AND ", s, flags=re.IGNORECASE)
    s = re.sub(r"\sOR\s", "\n OR ", s, flags=re.IGNORECASE)

    # uppercase common keywords (simple pass)
    for kw in sorted(
        [
            "select",
            "distinct",
            "from",
            "left join",
            "right join",
            "inner join",
            "full join",
            "on",
            "where",
            "group by",
            "having",
            "order by",
            "and",
            "or",
            "as",
        ],
        key=len,
        reverse=True,
    ):
        pattern = r"\b" + re.escape(kw) + r"\b"
        s = re.sub(pattern, kw.upper(), s, flags=re.IGNORECASE)

    return s.lstrip()


STATUS_BG_HEX = {
    "OK": "#e8f5e9",  # light green
    "SPLIT": "#fff8e1",  # light yellow
    "SHORT": "#ffe0b2",  # light orange
    "OOS": "#ffebee",  # light red
}
PRIME_BG_HEX = "#e3f2fd"  # light blue for prime BU cells
SPLIT_BG_HEX = "#e0f7fa"  # light cyan for split contributors

# Brushes for QT Table coloring
STATUS_BG_BRUSH = {k: QBrush(QColor(v)) for k, v in STATUS_BG_HEX.items()}


def _hex_to_argb(hex_code: str) -> str:
    h = (hex_code or "").lstrip("#").upper()
    return ("FF" + h) if len(h) == 6 else "FFFFFFFF"


def _merge_border(cell, *, left=None, right=None, top=None, bottom=None):
    """Set specific border sides without wiping the others."""
    b = cell.border
    cell.border = Border(
        left=left or b.left,
        right=right or b.right,
        top=top or b.top,
        bottom=bottom or b.bottom,
    )


def apply_excel_status_colors(
    path: str,
    status_header: str = "status",
    preferred_bu: str = "",
    bu_headers: list[str] | None = None,
    status_palette: dict[str, str] | None = None,
    # e.g. ['part_number','part_description','requested_qty']
    import_fields: list[str] | None = None,
    # e.g. ['perpetual_avg_cost_used','unit_currency_used','est_cost','est_currency']
    cost_fields: list[str] | None = None,
) -> tuple[bool, str]:
    try:
        wb = load_workbook(path)
        ws = wb.active

        # Headers
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [
            str(c.value).strip() if c.value is not None else "" for c in header_row
        ]
        lower_map = {h.lower(): i + 1 for i, h in enumerate(headers) if h}

        def find_col(names):
            for n in names:
                c = lower_map.get(n.lower())
                if c:
                    return c
            return None

        status_col = find_col([status_header, "Status"])
        rq_col = find_col(
            ["requested_qty", "requestedqty", "req_qty", "quantityrequired"]
        )

        # BU columns
        bu_cols = []
        if bu_headers:
            wanted = {b.upper() for b in bu_headers}
            for idx, h in enumerate(headers, start=1):
                if h.upper() in wanted:
                    bu_cols.append(idx)

        # Section borders (thick right border at group ends)
        """thin = Side(style='thin', color='FFB0B0B0')"""
        thick = Side(style="thick", color="FF808080")
        for c in range(1, ws.max_column + 1):
            _merge_border(ws.cell(row=1, column=c), bottom=thick)
        border_right_thick = Border(right=thick)
        border_bottom_thick = Border(bottom=thick)
        """border_right_thin = Border(right=thin)"""

        # header bottom divider
        for c in range(1, ws.max_column + 1):
            ws.cell(row=1, column=c).border = border_bottom_thick

        # Optional group borders if we can infer groups
        def mark_group_right(edge_col):
            if not edge_col:
                return
            for r in range(1, ws.max_row + 1):
                ws.cell(row=r, column=edge_col).border = border_right_thick

        # infer group edges
        import_edge = None
        if import_fields:
            last = max((lower_map.get(f.lower(), 0) for f in import_fields), default=0)
            import_edge = last if last > 0 else None

        bu_edge = None
        if bu_cols:
            bu_edge = max(bu_cols)

        cost_edge = None
        if cost_fields:
            last = max((lower_map.get(f.lower(), 0) for f in cost_fields), default=0)
            cost_edge = last if last > 0 else None

        # apply group edges
        mark_group_right(import_edge)
        mark_group_right(bu_edge)
        mark_group_right(cost_edge)

        # Freeze + filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Fills
        pal = status_palette or STATUS_BG_HEX  # STATUS_BG_HEX you defined earlier

        def fill(hexc):
            return PatternFill(
                fill_type="solid",
                start_color=_hex_to_argb(hexc),
                end_color=_hex_to_argb(hexc),
            )

        status_fills = {k: fill(v) for k, v in pal.items()}
        prime_fill = fill("#e3f2fd")  # blue
        split_fill = fill("#e0f7fa")  # cyan
        bold_font = Font(bold=True)

        # 1) Color Status column
        if status_col:
            for r in range(2, ws.max_row + 1):
                raw = ws.cell(row=r, column=status_col).value
                key = str(raw).strip().upper() if raw is not None else ""
                if key.startswith("OK"):  # OK (Preferred)
                    key = "OK"
                f = status_fills.get(key)
                if f:
                    ws.cell(row=r, column=status_col).fill = f

        # 2) Preferred/Prime/Split (BOM pivot only)
        pref = (preferred_bu or "").upper()
        pref_idx = None
        if pref and pref in headers:
            pref_idx = headers.index(pref) + 1

        if rq_col and bu_cols:
            for r in range(2, ws.max_row + 1):
                # requested qty
                try:
                    req = int(ws.cell(row=r, column=rq_col).value or 0)
                except:
                    req = 0
                if req <= 0:
                    continue

                # BU values
                vals = []
                for c in bu_cols:
                    try:
                        v = int(ws.cell(row=r, column=c).value or 0)
                    except:
                        v = 0
                    vals.append(max(0, v))

                # Preferred wins if it meets req
                if pref_idx and pref_idx in bu_cols:
                    try:
                        v_pref = int(ws.cell(row=r, column=pref_idx).value or 0)
                    except:
                        v_pref = 0
                    if v_pref >= req:
                        cell = ws.cell(row=r, column=pref_idx)
                        cell.fill = prime_fill
                        cell.font = bold_font
                        continue

                # Prime: first BU meeting req
                prime_rel = next((i for i, v in enumerate(vals) if v >= req), None)
                if prime_rel is not None:
                    cell = ws.cell(row=r, column=bu_cols[prime_rel])
                    cell.fill = prime_fill
                    cell.font = bold_font
                else:
                    # Split: shade contributors until cumulative >= req
                    order = sorted(
                        range(len(vals)), key=lambda i: vals[i], reverse=True
                    )
                    cum = 0
                    for i in order:
                        if vals[i] <= 0:
                            continue
                        cell = ws.cell(row=r, column=bu_cols[i])
                        cell.fill = split_fill
                        cum += vals[i]
                        if cum >= req:
                            break

        # Simple width fit
        try:
            for col in ws.columns:
                width = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col[:400]
                )
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(10, width + 2), 48
                )
        except:
            pass

        # --- Number formats for costs (4 dp) ---
        cost_like = {"perpetual_avg_cost", "perpetual_avg_cost_used", "est_cost"}
        for idx, h in enumerate(headers, start=1):
            if h.strip().lower() in {n.lower() for n in cost_like}:
                col_letter = get_column_letter(idx)
                for r in range(2, ws.max_row + 1):
                    ws[f"{col_letter}{r}"].number_format = "0.0000"

        # --- If there is an old static TOTAL row, delete it ---
        last = ws.max_row
        first_col_val = str(ws.cell(row=last, column=1).value or "").strip().upper()
        if first_col_val == "TOTAL":
            ws.delete_rows(last, 1)

        # Build an index for columns we care about
        hdr_idx = {(h or "").strip().lower(): i for i, h in enumerate(headers, start=1)}

        est_col = hdr_idx.get("est_cost")
        req_col = hdr_idx.get("requested_qty")
        unit_col = hdr_idx.get("perpetual_avg_cost_used") or hdr_idx.get(
            "perpetual_avg_cost"
        )

        if est_col and req_col and unit_col:
            est_letter = get_column_letter(est_col)
            req_letter = get_column_letter(req_col)
            unit_letter = get_column_letter(unit_col)
            # Write the formula row-by-row (skip header)
            for r in range(2, ws.max_row + 1):
                ws[f"{est_letter}{r}"].value = (
                    f"=ROUND({req_letter}{r}*{unit_letter}{r},4)"
                )
                ws[f"{est_letter}{r}"].number_format = "0.0000"

        # --- TOTAL row with formula on est_cost ---
        est_col = next(
            (
                i
                for i, h in enumerate(headers, start=1)
                if h.strip().lower() == "est_cost"
            ),
            None,
        )
        if est_col:
            last = ws.max_row
            total_row = last + 1
            # label
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            _merge_border(ws.cell(row=total_row, column=1), top=thick)
            # formula (SUM of the est_cost column from row 2 to last)
            est_letter = get_column_letter(est_col)
            ws.cell(
                row=total_row,
                column=est_col,
                value=f"=SUM({est_letter}2:{est_letter}{last})",
            ).font = Font(bold=True)
            _merge_border(ws.cell(row=total_row, column=est_col), top=thick)
            # draw a top divider across the whole total row for clarity
            for c in range(1, ws.max_column + 1):
                _merge_border(ws.cell(row=total_row, column=c), top=thick)

        thick = Side(style="thick", color="FF808080")

        # Re-apply bottom border under EVERY header (some earlier ops may have replaced it)
        for c in range(1, ws.max_column + 1):
            _merge_border(ws.cell(row=1, column=c), bottom=thick)

        # Make doubly sure for these two headers (names are case-insensitive)
        hdrs_fix = {"requested_qty", "est_currency"}
        for idx, h in enumerate(headers, start=1):
            if h and h.strip().lower() in hdrs_fix:
                _merge_border(ws.cell(row=1, column=idx), bottom=thick)

        wb.save(path)
        return True, "Excel styling applied."
    except Exception as e:
        return False, f"Excel styling failed: {e}"


# ---------- File Search providers ----------


def human_size(n):
    try:
        n = float(n or 0)
    except Exception:
        n = 0.0
    for u in ("B", "KB", "MB", "GB", "TB", "PB"):
        if n < 1024 or u == "PB":
            return f"{n:,.1f} {u}"
        n /= 1024.0


class FileHit:
    __slots__ = ("name", "path", "is_dir", "size", "mtime", "source")

    def __init__(self, path, is_dir, size, mtime, source):
        self.path = path
        self.name = os.path.basename(path.rstrip("\\/"))
        self.is_dir = bool(is_dir)
        self.size = int(size or 0)
        self.mtime = float(mtime or time.time())
        self.source = str(source or "")


class QuickIndexProvider:
    """Search the single quick-index DB (if you’re still keeping one)."""

    def search(self, query: str, limit: int = 5000, ordered: bool = False):
        dbp = _qi_path()
        conn = ensure_quick_index_db(dbp)
        try:
            fts = _fts_build_query(query, ordered=ordered)
            rows = []
            # Check if FTS exists; otherwise fallback to LIKE
            have_fts = False
            try:
                conn.execute("SELECT 1 FROM files_fts LIMIT 1")
                have_fts = True
            except Exception:
                pass

            if have_fts and fts:
                sql = """
                  SELECT f.path, f.name, f.size, f.is_dir
                  FROM files f
                  JOIN files_fts ff ON ff.rowid = f.id
                  WHERE ff MATCH ?
                  ORDER BY f.is_dir DESC, f.name
                  LIMIT ?
                """
                for r in conn.execute(sql, (fts, limit)):
                    rows.append((r["path"], r["name"], r["size"], r["is_dir"]))
            else:
                # LIKE fallback (AND semantics across tokens)
                toks = [
                    t.strip('"')
                    for t in re.findall(r'"[^"]+"|\\S+', str(query or ""))
                    if t.strip()
                ]
                if toks:
                    where = []
                    args = []
                    for t in toks:
                        pat = f"%{t.replace('*','%').replace('?', '_')}%"
                        where.append(
                            "(name LIKE ? COLLATE NOCASE OR path LIKE ? COLLATE NOCASE)"
                        )
                        args.extend([pat, pat])
                    sql = f"SELECT path, name, size, is_dir FROM files WHERE {' AND '.join(where)} ORDER BY mtime DESC LIMIT ?"
                    args.append(limit)
                else:
                    sql = "SELECT path, name, size, is_dir FROM files ORDER BY mtime DESC LIMIT ?"
                    args = [limit]
                for r in conn.execute(sql, args):
                    rows.append((r["path"], r["name"], r["size"], r["is_dir"]))
            return rows
        finally:
            conn.close()

    def has_index(self) -> bool:
        return os.path.exists(_qi_path())


class WindowsSearchProvider:
    """Search via the Windows Search index (WDS/AQS) with a safe fallback."""

    def __init__(self):
        pass

    def available(self) -> bool:
        try:
            return True
        except Exception:
            return False

    def _contains_string(self, tokens):
        escaped = [t.replace('"', '""') for t in tokens if t]
        if not escaped:
            return "1=1"
        clauses = [f"CONTAINS(System.FileName, '\"{t}*\"' )" for t in escaped]
        return " AND ".join(clauses)

    def search(self, query, limit=2000, allowed_roots=None):
        tokens = [t for t in re.split(r"[\\s]+", str(query or "").strip()) if t]
        hits = []
        try:
            conn = win32.Dispatch("ADODB.Connection")
            rs = win32.Dispatch("ADODB.Recordset")
            conn.Open(
                "Provider=Search.CollatorDSO;Extended Properties='Application=Windows'"
            )

            scope = ""
            if allowed_roots:
                ors = " OR ".join([f"SCOPE='{root}'" for root in allowed_roots])
                scope = f" AND ({ors})"

            where = self._contains_string(tokens) + scope
            sql = (
                f"SELECT TOP {int(limit)} "
                "System.ItemNameDisplay, System.ItemTypeText, System.Size, System.DateModified, "
                "System.ItemFolderPathDisplay, System.ItemPathDisplay "
                f"FROM SYSTEMINDEX WHERE {where} ORDER BY System.DateModified DESC"
            )
            rs.Open(sql, conn, 1, 1)  # adOpenKeyset, adLockReadOnly
            while not rs.EOF:
                name = rs.Fields.Item(0).Value
                ftype = rs.Fields.Item(1).Value or ""
                size = rs.Fields.Item(2).Value or 0
                modified = rs.Fields.Item(3).Value
                folder = rs.Fields.Item(4).Value or ""
                path = rs.Fields.Item(5).Value or ""

                # best-effort timestamp
                try:
                    ts = modified.timestamp()
                except Exception:
                    try:
                        ts = time.mktime(modified.timetuple())
                    except Exception:
                        ts = time.time()

                is_dir = str(ftype).lower() in ("file folder", "folder")
                # For folders, prefer the folder path if the item path is blank
                p = path or folder or name
                hits.append(FileHit(p, is_dir, size, ts, "WinIndex"))
                rs.MoveNext()

            rs.Close()
            conn.Close()
            return hits
        except Exception:
            # fallback: crawl only the chosen roots (if any)
            return CrawlProvider(roots=allowed_roots).search(query, limit=limit)


class CrawlProvider:
    def __init__(self, roots=None):
        self.roots = roots or self._default_roots()

    def _default_roots(self):
        roots = [os.path.expanduser("~")]
        od = os.environ.get("OneDrive")
        if od and os.path.isdir(od):
            roots.append(od)
        sp = os.path.join(os.path.expanduser("~"), "SharePoint")
        if os.path.isdir(sp):
            roots.append(sp)
        # drives
        if os.name == "nt":
            try:
                import ctypes
                import string

                bitmask = ctypes.windll.kernel32.GetLogicalDrives()
                for i, l in enumerate(string.ascii_uppercase):
                    if bitmask & (1 << i):
                        d = f"{l}:\\"
                        if os.path.isdir(d):
                            roots.append(d)
            except Exception:
                pass
        # de-dupe
        uniq = []
        seen = set()
        for r in roots:
            if r not in seen and os.path.isdir(r):
                uniq.append(r)
                seen.add(r)
        return uniq

    def _walk(self, root, q, results, limit):
        try:
            dirs_scanned = files_scanned = 0
            for dirpath, dirnames, filenames in os.walk(root):
                # progress every ~200 items
                if hasattr(self, "progress_cb"):
                    if (dirs_scanned + files_scanned) % 200 == 0:
                        self.progress_cb(dirpath, dirs_scanned, files_scanned)

                # folder match
                try:
                    st = os.stat(dirpath)
                    if self._match(os.path.basename(dirpath), q):
                        results.append(FileHit(dirpath, True, 0, st.st_mtime, "Local"))
                        if len(results) >= limit:
                            return
                except Exception:
                    pass
                dirs_scanned += 1

                # file matches
                for fn in filenames:
                    if not self._match(fn, q):
                        continue
                    p = os.path.join(dirpath, fn)
                    try:
                        st = os.stat(p)
                        results.append(
                            FileHit(p, False, st.st_size, st.st_mtime, "Local")
                        )
                        if len(results) >= limit:
                            return
                    except Exception:
                        continue
                    files_scanned += 1
        except Exception:
            pass

    def _match(self, name, query):
        # any-order wildcards by default, supports quoted phrases
        toks = [t for t in re.findall(r'"[^"]+"|\S+', query) if t.strip()]
        if not toks:
            return True
        for t in toks:
            t = t.strip('"')
            if not fnmatch.fnmatch(name.lower(), (t + "*").lower()):
                return False
        return True

    def search(self, query, limit=3000):
        results = []
        threads = []
        for r in self.roots:
            t = threading.Thread(
                target=self._walk, args=(r, query, results, limit), daemon=True
            )
            t.start()
            threads.append(t)
        for t in threads:
            t.join()
        return results


class IndexedProvider:
    """Query local filename-only indexes built by the Index Builder pane."""

    def __init__(self, roots=None):
        self.roots = list(roots or [])

    def _iter_db_paths(self, allowed_roots=None):
        base = os.path.join(
            os.environ.get("LOCALAPPDATA", os.path.expanduser(r"~\AppData\Local")),
            "PartSearch",
            "index",
        )
        if not allowed_roots:
            # all known DBs
            for dbp in glob.glob(
                os.path.join(base, "**", "index_db.sqlite"), recursive=True
            ):
                yield dbp
            return

        # try hashed paths first
        missing = []
        for root in allowed_roots:
            dbp = index_db_path(root)
            if os.path.exists(dbp):
                yield dbp
            else:
                missing.append(root)

        if not missing:
            return

        # fallback: probe every DB and include those whose rows have matching files.root
        wanted = {os.path.normcase(os.path.normpath(r)).rstrip("\\/") for r in missing}
        for dbp in glob.glob(
            os.path.join(base, "**", "index_db.sqlite"), recursive=True
        ):
            try:
                con = open_sqlite_ro(dbp)
                con.row_factory = sqlite3.Row
                # sample just one row; DBs are small
                row = con.execute(
                    "SELECT root FROM files WHERE root IS NOT NULL LIMIT 1"
                ).fetchone()
                con.close()
                if row:
                    rkey = os.path.normcase(os.path.normpath(str(row["root"]))).rstrip(
                        "\\/"
                    )
                    if rkey in wanted:
                        yield dbp
            except Exception:
                try:
                    con.close()
                except:
                    pass
                continue

    def search(self, query: str, limit=5000, allowed_roots=None):
        tokens = [
            t.strip('"')
            for t in re.findall(r'"[^"]+"|\S+', str(query or ""))
            if t.strip()
        ]
        results = []
        for dbp in self._iter_db_paths(allowed_roots):
            if not os.path.exists(dbp):
                continue
            conn = open_sqlite_ro(dbp)
            conn.row_factory = sqlite3.Row
            try:
                if tokens:
                    where = []
                    args = []
                    for t in tokens:
                        pat = f"%{t.replace('*','%').replace('?', '_')}%"
                        where.append(
                            "(name LIKE ? COLLATE NOCASE OR path LIKE ? COLLATE NOCASE)"
                        )
                        args.extend([pat, pat])
                    sql = f"""
                        SELECT path, name, is_dir, size, mtime
                        FROM files
                        WHERE {" AND ".join(where)}
                        ORDER BY mtime DESC
                        LIMIT ?
                    """
                    args.append(max(0, int(limit) - len(results)))
                else:
                    sql = "SELECT path, name, is_dir, size, mtime FROM files ORDER BY mtime DESC LIMIT ?"
                    args = [max(0, int(limit) - len(results))]

                for row in conn.execute(sql, args):
                    results.append(
                        FileHit(
                            row["path"],
                            bool(row["is_dir"]),
                            row["size"],
                            row["mtime"],
                            "Index",
                        )
                    )
                    if len(results) >= limit:
                        break
            finally:
                conn.close()
            if len(results) >= limit:
                break
        return results


class EnoviaProvider:
    """Query locally indexed ENOVIA Online results (CSV-driven; permission-respecting)."""

    def __init__(self):
        self.dbp = enovia_db_path()

    def available(self) -> bool:
        return os.path.exists(self.dbp)

    def last_indexed(self) -> str:
        try:
            con = open_sqlite_ro(self.dbp)
            v = con.execute(
                "SELECT val FROM meta WHERE key='last_full_scan'"
            ).fetchone()
            con.close()
            return v[0] if v else "—"
        except Exception:
            return "—"

    def search(self, query: str, limit=2000):
        if not self.available():
            return []
        toks = [
            t.strip('"').lower()
            for t in re.findall(r'"[^"]+"|\S+', str(query or ""))
            if t.strip()
        ]
        where = " AND ".join(["LOWER(name) LIKE ?"] * len(toks)) if toks else "1=1"
        args = ["%" + t.replace("*", "%") + "%" for t in toks]
        con = open_sqlite_ro(self.dbp)
        con.row_factory = sqlite3.Row
        sql = (
            "SELECT name, type, rev, state, modified, url FROM enovia WHERE "
            + where
            + " ORDER BY last_seen DESC LIMIT ?"
        )
        rows = con.execute(sql, args + [int(limit)]).fetchall()
        con.close()
        hits = []
        for r in rows:
            name = r["name"] or "(unnamed)"
            # show as a virtual 'file' pointing to URL
            hits.append(FileHit(r["url"] or name, False, 0, time.time(), "ENOVIA"))
        return hits


class IndexWorker(QThread):
    # root, scanned, updated, seconds
    progress_root = pyqtSignal(str, int, int, float)
    progress_path = pyqtSignal(str)  # current path
    root_started = pyqtSignal(str)  # root started
    # root, scanned, updated, seconds
    root_done = pyqtSignal(str, int, int, float)
    all_done = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, roots: list[str], incremental: bool = False, excludes=None):
        super().__init__()
        self.roots = list(roots or [])
        self.incremental = bool(incremental)
        self.excludes = list(excludes or [])
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        SKIP_NAMES = {
            "System Volume Information",
            "$RECYCLE.BIN",
            "$Recycle.Bin",
            "Windows",
            "Windows.old",
            "AppData",
            "node_modules",
            ".git",
            ".svn",
        }
        BATCH = 800
        PROG_INTERVAL = 1.0

        try:
            for root in self.roots or []:
                if getattr(self, "_stop", False):
                    break

                # tell UI which root started
                try:
                    self.root_started.emit(str(root))
                except Exception:
                    pass

                db_path = index_db_path(root)  # canonical location
                # creates/migrates schema and enables WAL
                db = ensure_quick_index_db(db_path)
                try:
                    q = db.execute("PRAGMA query_only").fetchone()
                    if q and int(q[0]) == 1:
                        raise RuntimeError(
                            "Indexer opened a READ-ONLY connection. Use open_sqlite(...), not open_sqlite_ro(...)."
                        )
                except Exception:
                    pass
                cur = db.cursor()

                scanned_count = 0
                updated_count = 0
                started = time.time()
                last_prog = started
                current_pass = int(started)

                def _upsert(path, is_dir, name, ext, size, mtime, ctime, parent):
                    nonlocal updated_count
                    before = db.total_changes
                    cur.execute(
                        """
                        INSERT INTO files(root, path, name, ext, size, mtime, ctime, is_dir, parent, pass_id)
                        VALUES(?,?,?,?,?,?,?,?,?,?)
                        ON CONFLICT(root, path) DO UPDATE SET
                            name=excluded.name,
                            ext=excluded.ext,
                            size=excluded.size,
                            mtime=excluded.mtime,
                            ctime=excluded.ctime,
                            is_dir=excluded.is_dir,
                            parent=excluded.parent,
                            pass_id=excluded.pass_id
                    """,
                        (
                            root,
                            path,
                            name,
                            ext,
                            size,
                            mtime,
                            ctime,
                            1 if is_dir else 0,
                            parent,
                            current_pass,
                        ),
                    )
                    if db.total_changes > before:
                        updated_count += 1

                def _emit_progress(current_dir):
                    nonlocal last_prog
                    now = time.time()
                    if (now - last_prog) >= PROG_INTERVAL:
                        last_prog = now
                        try:
                            self.progress_path.emit(str(current_dir or root))
                        except Exception:
                            pass
                        try:
                            self.progress_root.emit(
                                str(root),
                                int(scanned_count),
                                int(updated_count),
                                float(now - started),
                            )
                        except Exception:
                            pass

                def _scan_dir(d):
                    nonlocal scanned_count
                    try:
                        with os.scandir(d) as it:
                            for entry in it:
                                if getattr(self, "_stop", False):
                                    return
                                name = entry.name
                                if name in SKIP_NAMES:
                                    continue
                                try:
                                    is_dir = entry.is_dir(follow_symlinks=False)
                                except Exception:
                                    is_dir = False

                                path = entry.path
                                parent = os.path.dirname(path)
                                if is_dir:
                                    try:
                                        st = entry.stat(follow_symlinks=False)
                                        size = None
                                        mtime = st.st_mtime
                                        ctime = st.st_ctime
                                    except Exception:
                                        size = None
                                        mtime = None
                                        ctime = None
                                    _upsert(
                                        path, True, name, "", size, mtime, ctime, parent
                                    )
                                    scanned_count += 1
                                    if scanned_count % BATCH == 0:
                                        db.commit()
                                        _emit_progress(path)
                                    _scan_dir(path)
                                else:
                                    try:
                                        st = entry.stat(follow_symlinks=False)
                                        size = st.st_size
                                        mtime = st.st_mtime
                                        ctime = st.st_ctime
                                    except Exception:
                                        size = None
                                        mtime = None
                                        ctime = None
                                    ext = os.path.splitext(name)[1][1:].lower()
                                    _upsert(
                                        path,
                                        False,
                                        name,
                                        ext,
                                        size,
                                        mtime,
                                        ctime,
                                        parent,
                                    )
                                    scanned_count += 1
                                    if scanned_count % BATCH == 0:
                                        db.commit()
                                        _emit_progress(parent or d)
                    except (PermissionError, FileNotFoundError, OSError):
                        return

                # ensure we index the root row itself
                try:
                    st_root = os.stat(root)
                    mtime_r = st_root.st_mtime
                    ctime_r = st_root.st_ctime
                except Exception:
                    mtime_r = None
                    ctime_r = None
                root_name = os.path.basename(root.rstrip("\\/")) or root
                _upsert(
                    root,
                    True,
                    root_name,
                    "",
                    None,
                    mtime_r,
                    ctime_r,
                    os.path.dirname(root.rstrip("\\/")),
                )
                scanned_count += 1
                _emit_progress(root)

                # walk
                _scan_dir(root)

                # finalize this root
                completed = not bool(getattr(self, "_stop", False))
                try:
                    if completed:
                        # purge stale rows from prior passes
                        cur.execute(
                            "DELETE FROM files WHERE pass_id <> ?", (current_pass,)
                        )
                        db.commit()  # commit before FTS maintenance

                        # try to keep FTS fresh, best-effort
                        try:
                            cur.execute(
                                "INSERT INTO files_fts(files_fts) VALUES('rebuild')"
                            )
                        except Exception:
                            pass
                        db.commit()

                        # stamp last scan using the canonical helper
                        try:
                            set_last_indexed_now_for_root(root)
                        except Exception:
                            pass
                finally:
                    try:
                        db.close()
                    except Exception:
                        pass

                elapsed_seconds = max(0.0, time.time() - started)
                try:
                    self.root_done.emit(
                        str(root),
                        int(scanned_count),
                        int(updated_count),
                        float(elapsed_seconds),
                    )
                except Exception:
                    pass

            try:
                self.all_done.emit()
            except Exception:
                pass

        except Exception as e:
            try:
                self.error.emit(str(e))
            except Exception:
                pass


class SearchWorker(QThread):
    chunk = pyqtSignal(list)  # emits a list of table rows (7 columns each)
    progress = pyqtSignal(str)  # status text, optional
    done = pyqtSignal(list)  # final rows (may be truncated by limit)
    error = pyqtSignal(str)

    def __init__(
        self,
        query: str,
        use_win=True,
        use_qi=True,
        use_crawl=False,
        allowed_roots=None,
        limit=5000,
    ):
        super().__init__()
        self.query = query or ""
        self.use_win = bool(use_win)
        self.use_qi = bool(use_qi)
        self.use_crawl = bool(use_crawl)
        self.allowed_roots = list(allowed_roots) if allowed_roots else None
        self.limit = int(limit)
        self._abort = False

    def abort(self):
        self._abort = True

    # --- helpers ---
    @staticmethod
    def _fmt_size(n):
        try:
            n = int(n)
        except Exception:
            return ""
        units = ("B", "KB", "MB", "GB", "TB")
        i = 0
        f = float(n)
        while f >= 1024 and i < len(units) - 1:
            f /= 1024.0
            i += 1
        return f"{f:.1f} {units[i]}"

    @staticmethod
    def _classify(path, is_dir):
        if is_dir:
            return "Folder"
        p = (path or "").lower()
        if p.endswith((".sldasm",)):
            return "Assembly"
        if p.endswith((".sldprt",)):
            return "Part"
        if p.endswith((".slddrw", ".pdf")):
            return "Drawing"
        return "File"

    def _hit_to_row(self, source, path, is_dir=False, size=None, mtime=None):
        name = os.path.basename(path) if path else ""
        ftype = self._classify(path, is_dir)
        loc = os.path.dirname(path) if path else ""
        mod = ""
        if mtime:
            try:
                mod = datetime.datetime.fromtimestamp(float(mtime)).strftime(
                    "%Y-%m-%d %H:%M"
                )
            except Exception:
                mod = ""
        return [name, ftype, self._fmt_size(size), mod, loc, source, path]

    def _search_windows(self, q):
        try:
            prov = WindowsSearchProvider()
            for hit in prov.search(q, allowed_roots=self.allowed_roots):
                if self._abort:
                    return
                row = self._hit_to_row(
                    "Windows", hit.path, hit.is_dir, hit.size, hit.mtime
                )
                yield row
        except Exception as e:
            self.error.emit(f"Windows index error: {e}")

    def _search_index(self, q):
        try:
            prov = IndexedProvider()
            for hit in prov.search(q, allowed_roots=self.allowed_roots):
                if self._abort:
                    return
                # hit is a FileHit object
                yield self._hit_to_row(
                    "Index", hit.path, hit.is_dir, hit.size, hit.mtime
                )
        except Exception as e:
            self.error.emit(f"Local index error: {e}")

    def _search_crawl(self, q):
        try:
            prov = CrawlProvider(roots=self.allowed_roots)
            if hasattr(self, "_crawl_progress"):
                prov.progress_cb = self._crawl_progress
            for hit in prov.search(q):
                if self._abort:
                    return
                row = self._hit_to_row(
                    "Crawl", hit.path, hit.is_dir, hit.size, hit.mtime
                )
                yield row
        except Exception as e:
            self.error.emit(f"Crawl error: {e}")

    def run(self):
        rows, seen = [], set()
        try:
            q = (self.query or "").strip()
            lim = max(0, int(self.limit))

            def add_row(r):
                key = os.path.normcase(os.path.normpath(r[6])).rstrip("\\/")
                if key in seen:
                    return False
                seen.add(key)
                rows.append(r)
                return True

            def feed(gen):
                batch = []
                for row in gen:
                    if self._abort:
                        return
                    if add_row(row):
                        batch.append(row)
                    if len(rows) >= lim:
                        break
                    if len(batch) >= 100:
                        self.chunk.emit(batch)
                        batch = []
                if batch:
                    self.chunk.emit(batch)

            # 1) Quick Index FIRST (if enabled)
            if getattr(self, "use_qi", False) and not self._abort:
                self.progress.emit("Searching Quick Index…")
                batch = []
                for r in self._search_index(q):
                    if self._abort:
                        break
                    if add_row(r):
                        batch.append(r)
                    if len(rows) >= lim:
                        break
                if rows:
                    # stream in chunks for UI smoothness
                    self.chunk.emit(rows[: min(len(rows), 200)])

            # 2) Windows Index (optional)
            if getattr(self, "use_win", False) and not self._abort and len(rows) < lim:
                self.progress.emit("Searching Windows Index…")
                feed(self._search_windows(q))

            # 3) Crawl (optional, slow)
            if (
                getattr(self, "use_crawl", False)
                and not self._abort
                and len(rows) < lim
            ):
                self.progress.emit("Crawling folders…")
                feed(self._search_crawl(q))

            self.done.emit(rows[:lim])
        except Exception as e:
            self.error.emit(str(e))


# ===== Sorting proxy for FileSearch (size/date aware) =====


class FileResultsProxy(QSortFilterProxyModel):
    # columns: ["Name","Type","Size","Modified","Location","Source","FullPath"]
    def lessThan(self, left, right):
        col = left.column()
        src = self.sourceModel()
        li = src.index(left.row(), col)
        ri = src.index(right.row(), col)
        lv = src.data(li, Qt.DisplayRole) or ""
        rv = src.data(ri, Qt.DisplayRole) or ""

        if col == 2:  # Size -> parse "1.2 MB"

            def to_bytes(s):
                try:
                    t = s.strip().split()
                    if not t:
                        return -1
                    num = float(t[0])
                    unit = (t[1] if len(t) > 1 else "B").upper()
                    mult = {
                        "B": 1,
                        "KB": 1024,
                        "MB": 1024**2,
                        "GB": 1024**3,
                        "TB": 1024**4,
                    }.get(unit, 1)
                    return num * mult
                except Exception:
                    return -1

            return to_bytes(lv) < to_bytes(rv)
        elif col == 3:  # Modified "YYYY-MM-DD HH:MM"

            def to_ts(s):
                try:
                    return QDateTime.fromString(
                        s, "yyyy-MM-dd HH:mm"
                    ).toSecsSinceEpoch()
                except Exception:
                    return -1

            return to_ts(lv) < to_ts(rv)
        else:
            return str(lv).lower() < str(rv).lower()


# ===== Scanner integration =====


class SerialScanListener(QThread):
    scanned = pyqtSignal(str)

    def __init__(self, port: str, baud: int = 9600, eol: str = "\r\n", parent=None):
        super().__init__(parent)
        self.port, self.baud, self.eol = port, baud, eol
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        try:
            import serial
        except Exception:
            return
        try:
            with serial.Serial(self.port, self.baud, timeout=0.2) as ser:
                buf = ""
                while not self._stop:
                    chunk = ser.read(1024).decode(errors="ignore")
                    if not chunk:
                        continue
                    buf += chunk
                    while True:
                        idx = buf.find(self.eol)
                        if idx < 0:
                            break
                        token = buf[:idx].strip()
                        buf = buf[idx + len(self.eol) :]
                        if token:
                            self.scanned.emit(token)
        except Exception:
            pass


# ----------------------------------
# GUI layer
# ----------------------------------


class PandasModel(QAbstractTableModel):
    def __init__(self, dataframe: pd.DataFrame):
        super().__init__()
        self._dataframe = dataframe

    def rowCount(self, parent=None):
        return self._dataframe.shape[0]

    def columnCount(self, parent=None):
        return self._dataframe.shape[1]

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None

        if role == Qt.DisplayRole:
            value = self._dataframe.iloc[index.row(), index.column()]
            return str(value)

        if role == Qt.BackgroundRole and "status" in self._dataframe.columns:
            try:
                key = normalize_status_key(self._dataframe.at[index.row(), "status"])
                brush = STATUS_BG_BRUSH.get(key)
                if brush:
                    return brush
            except Exception:
                pass

        if role == Qt.ForegroundRole and "status" in self._dataframe.columns:
            try:
                key = normalize_status_key(self._dataframe.at[index.row(), "status"])
                if key in STATUS_BG_BRUSH:
                    return QBrush(QColor("#111111"))
            except Exception:
                pass

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return str(self._dataframe.columns[section])
            elif orientation == Qt.Vertical:
                return str(self._dataframe.index[section])
        return None

    def flags(self, index):
        # Enables Selection (But keeps it as read only)
        return Qt.ItemIsSelectable | Qt.ItemIsEnabled


class DenodoQuery(QWidget):
    def __init__(self):
        super().__init__()
        self.bom_map = {}
        self.csv_desc_map = {}
        self.import_order = []
        self.bom_mode = None  # Set in init_UI
        self.target_currency = "USD"
        self.fx_rates = {"USD": 1.0, "EUR": 1.08, "NOK": 0.094, "GBP": 1.28}
        self.preferred_bu = ""  # empty = none selected
        self.reserve_across_bom = False  # default off
        self.init_UI()
        self.scan_thread = None

    def _set_last_indexed_label(self):
        try:
            if hasattr(self, "files_panel") and hasattr(
                self.files_panel, "_refresh_last_indexed_label"
            ):
                self.files_panel._refresh_last_indexed_label()
            if hasattr(self, "files_panel") and hasattr(
                self.files_panel, "_load_saved_locations"
            ):
                self.files_panel._load_saved_locations()
        except Exception:
            pass

    def _attach_scanner(self, port, baud=9600):
        if self.scan_thread and self.scan_thread.isRunning():
            self.scan_thread.stop()
            self.scan_thread.wait(200)
        self.scan_thread = SerialScanListener(port, baud, parent=self)
        self.scan_thread.scanned.connect(self._on_scan)
        self.scan_thread.start()

    def _open_scanner_dialog(self):
        try:
            import serial.tools.list_ports as lp

            ports = [p.device for p in lp.comports()]
        except Exception:
            ports = []
        if not ports:
            QMessageBox.information(self, "Scanner", "No serial scanners found.")
            return
        port, ok = QInputDialog.getItem(
            self, "Scanner", "Pick COM port:", ports, 0, False
        )
        if ok and port:
            self._attach_scanner(port)
            QMessageBox.information(
                self, "Scanner", f"Listening on {port}. Scan a barcode to search."
            )

    def _on_scan(self, text: str):
        idx = self.stack.currentIndex()
        if idx == 0:  # PSFT
            self.item_id.setText(text)
            self.display_data()
        elif idx == 1:  # File Search
            self.files_panel.q.setText(text)
            self.files_panel.run_search()

    def init_UI(self):
        root = QHBoxLayout(self)

        # left rail
        rail = QVBoxLayout()
        self.nav_index = QPushButton("Quick Index")
        self.nav_index.setCheckable(True)
        self.nav_psft = QPushButton("PSFT Search")
        self.nav_psft.setCheckable(True)
        self.nav_files = QPushButton("File Search")
        self.nav_files.setCheckable(True)
        self.nav_scan = QPushButton("Scanner…")
        self.nav_scan.setEnabled(False)
        rail.addWidget(self.nav_scan)
        self.nav_future = QPushButton("Future Options")
        self.nav_future.setCheckable(True)
        self.nav_future.setEnabled(False)

        for b in (self.nav_index, self.nav_psft, self.nav_files, self.nav_future):
            b.setAutoExclusive(True)
            rail.addWidget(b)
        rail.addStretch(1)
        root.addLayout(rail, 0)

        # stacked content
        self.stack = QStackedWidget(self)
        self.psft_panel = self._build_psft_panel()
        self.files_panel = FileSearchPane(
            self,
            run_psft_callback=lambda pn: self.run_psft_from_files(pn, strict=True),
            open_index_callback=lambda: self._switch_mode(2),
        )
        self.index_panel = IndexBuilderPane(self, policy=getattr(self, "policy", {}))

        self.stack.addWidget(self.psft_panel)  # 0
        self.stack.addWidget(self.files_panel)  # 1
        self.stack.addWidget(self.index_panel)  # 2

        root.addWidget(self.stack, 1)

        self.setLayout(root)
        self.setWindowTitle("That Search Tool")
        self.resize(1200, 750)

        self.nav_psft.clicked.connect(lambda: self._switch_mode(0))
        self.nav_files.clicked.connect(lambda: self._switch_mode(1))
        self.nav_index.clicked.connect(lambda: self._switch_mode(2))
        self.nav_scan.clicked.connect(self._open_scanner_dialog)
        self.nav_psft.setChecked(True)

        # detect once UI is shown
        QTimer.singleShot(200, self._detect_scanner)

        # Open on PSFT
        self.stack.setCurrentIndex(0)
        self.nav_psft.setChecked(True)
        self.nav_files.setChecked(False)
        self.nav_index.setChecked(False)
        self._load_user_prefs()
        self._set_last_indexed_label()  # (your existing UI refresh)

    def _detect_scanner(self):
        try:
            import serial.tools.list_ports as lp

            ports = [p.device for p in lp.comports()]
        except Exception:
            ports = []
        self.nav_scan.setEnabled(bool(ports))

    def _switch_mode(self, idx: int):
        if idx == 1 and not getattr(self, "files_enabled", True):
            QMessageBox.information(
                self, "File Search", "File search is disabled by policy."
            )
            return
        self.stack.setCurrentIndex(idx)
        if hasattr(self, "nav_psft"):
            self.nav_psft.setChecked(idx == 0)
        if hasattr(self, "nav_files"):
            self.nav_files.setChecked(idx == 1)
        if hasattr(self, "nav_index"):
            self.nav_index.setChecked(idx == 2)

    def run_psft_from_files(self, pn: str, strict: bool = True):
        # Go to PSFT search form
        self._switch_mode(0)
        # Keep checked BUs and qty; clear wildcards if strict
        self.item_id.setText(str(pn or "").strip())
        if strict:
            self.mfg_part.clear()
            self.and_wildcards.clear()
            self.either_or_wildcards.clear()
            self.or_wildcards.clear()
        # Auto-run
        self.display_data()

    def _get_checked_bus(self) -> list:
        """Return selected BU codes as uppercase strings."""
        out = []
        checks = getattr(self, "bu_checks", None)

        # Common patterns:
        if isinstance(checks, dict):
            for code, cb in checks.items():
                if cb.isChecked():
                    out.append(str(code).upper().strip())
        elif isinstance(checks, (list, tuple)):
            for item in checks:
                # list/tuple of (code, checkbox)
                if isinstance(item, (list, tuple)) and len(item) >= 2:
                    code, cb = item[0], item[1]
                    if hasattr(cb, "isChecked") and cb.isChecked():
                        out.append(str(code).upper().strip())
                # or it's directly a QCheckBox with its text as the code
                elif (
                    hasattr(item, "isChecked")
                    and item.isChecked()
                    and hasattr(item, "text")
                ):
                    out.append(item.text().upper().strip())

        return [b for b in out if b]

    """def _read_index_roots_json(self) -> list:
        p = os.path.join(PROGRAMDATA_DIR, "index_roots.json")
        try:
            with open(p, "r", encoding="utf-8") as f:
                return (json.load(f) or {}).get("roots", [])
        except Exception:
            return []"""

    def _collect_table_roots(self) -> list[dict]:
        rows = []
        seen = set()
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, self.COL_PATH)
            if not pit:
                continue
            path = (pit.text() or "").strip()
            if not path:
                continue
            k = path.lower()
            if k in seen:
                continue
            seen.add(k)

            chk = self.table.item(r, self.COL_CHECK)
            checked = (chk.checkState() == Qt.Checked) if chk else True

            def _num(col):
                it = self.table.item(r, col)
                try:
                    return int((it.text() or "0").replace(",", ""))
                except Exception:
                    return 0

            rows.append(
                {
                    "path": path,
                    "checked": checked,
                    "files_count": _num(self.COL_FILES),
                    "updated_count": _num(self.COL_UPDATED),
                    "last_full_scan": (
                        self.table.item(r, self.COL_LAST).text()
                        if self.table.item(r, self.COL_LAST)
                        else ""
                    ),
                }
            )
        return rows

    """def _write_index_roots_json(self):
        p = os.path.join(PROGRAMDATA_DIR, "index_roots.json")
        data = {"roots": self._collect_table_roots()}
        try:
            tmp = p + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            os.replace(tmp, p)
        except Exception:
            pass"""

    def _build_psft_panel(self) -> QWidget:
        w = QWidget(self)
        grid = QGridLayout(w)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 0)
        # Row 0 (Filters/Actions) should be content-height; row 1 (Business Units) can grow
        grid.setRowStretch(0, 0)
        grid.setRowStretch(1, 1)
        grid.setRowStretch(2, 0)  # Options strip stays compact

        # ------- Filters group -------
        g_filters = QGroupBox("Filters", w)
        gf = QGridLayout(g_filters)

        self.item_id = QLineEdit(w)
        self.item_id.setPlaceholderText("e.g., 12345; ABC-6789")
        self.item_id.textEdited.connect(lambda _: self.bom_mode.setChecked(False))
        self.item_id.returnPressed.connect(self.display_data)

        self.mfg_part = QLineEdit(w)
        self.mfg_part.setPlaceholderText("e.g., MFG123; 3M-7890")
        self.mfg_part.returnPressed.connect(self.display_data)

        self.and_wildcards = QLineEdit(w)
        self.and_wildcards.setPlaceholderText("semicolon ; separated")
        self.and_wildcards.returnPressed.connect(self.display_data)

        self.either_or_wildcards = QLineEdit(w)
        self.either_or_wildcards.setPlaceholderText(
            'groups with commas, groups separated by ;   e.g.  1/2",0.5; NPT,BSPP'
        )
        self.either_or_wildcards.returnPressed.connect(self.display_data)

        self.or_wildcards = QLineEdit(w)
        self.or_wildcards.setPlaceholderText(
            "semicolon ; separated (e.g., DIMR1; MTR1; NDT; WELR"
        )
        self.or_wildcards.returnPressed.connect(self.display_data)

        self.qty = QLineEdit(w)
        self.qty.setText("0")
        self.qty.setPlaceholderText(
            "leave blank to disregard filter; entered value ≥ BU available"
        )
        self.qty.returnPressed.connect(self.display_data)

        row = 0
        gf.addWidget(QLabel("Part No.:"), row, 0)
        gf.addWidget(self.item_id, row, 1)
        row += 1
        gf.addWidget(QLabel("MFG Part No.:"), row, 0)
        gf.addWidget(self.mfg_part, row, 1)
        row += 1
        gf.addWidget(QLabel("AND wildcards:"), row, 0)
        gf.addWidget(self.and_wildcards, row, 1)
        row += 1
        gf.addWidget(QLabel("Either/Or wildcards:"), row, 0)
        gf.addWidget(self.either_or_wildcards, row, 1)
        row += 1
        gf.addWidget(QLabel("Quality Code(s):"), row, 0)
        gf.addWidget(self.or_wildcards, row, 1)
        row += 1
        gf.addWidget(QLabel("Minimum Quantity:"), row, 0)
        gf.addWidget(self.qty, row, 1)
        row += 1

        # ------- Business Units (left) -------
        g_bus = QGroupBox("Business Units", w)
        g_bus.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        left = QWidget(w)
        grid_bu = QGridLayout(left)
        grid_bu.setContentsMargins(8, 8, 8, 8)
        grid_bu.setHorizontalSpacing(6)
        grid_bu.setVerticalSpacing(1)  # << tighter rows
        left.setStyleSheet(
            """
            QCheckBox { margin: 0px; padding: 0px; }
            QCheckBox::indicator { margin: 0px; }
        """
        )

        self.BUs = []
        rows_per_col = 9  # pack more per column; tweak to taste (9–11)
        r = c = 0
        for BU, desc in POSSIBLE_BUS.items():
            box = QCheckBox(BU, w)
            box.setToolTip(desc)
            if BU in DEFAULT_BUS:
                box.setChecked(True)
            grid_bu.addWidget(box, r, c)
            self.BUs.append(box)
            r += 1
            if r >= rows_per_col:
                r = 0
                c += 1

        # put the left checkbox grid into a light wrapper group to keep a boundary box
        g_bus_layout = QVBoxLayout(g_bus)
        g_bus_layout.setContentsMargins(8, 8, 8, 8)
        g_bus_layout.addWidget(left)

        # ------- BU Tools (right of Business Units) -------
        g_butools = QGroupBox("BU Tools", w)
        # expand vertically to match BU box
        g_butools.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
        g_butools.setMaximumWidth(200)

        vb_tools = QVBoxLayout(g_butools)
        vb_tools.setContentsMargins(8, 8, 8, 8)
        vb_tools.setSpacing(6)

        self.toggle_all = QPushButton("Check All", w)
        self.toggle_all.clicked.connect(self.toggle_BUs)
        vb_tools.addWidget(self.toggle_all)

        self.btn_save_defaults = QPushButton("Set Default BUs", w)
        self.btn_save_defaults.clicked.connect(self._save_user_prefs)
        vb_tools.addWidget(self.btn_save_defaults)

        self.sync_pref_btn = QPushButton("Use Checked BUs", w)
        self.sync_pref_btn.setToolTip(
            "Restrict Preferred BU choices to currently checked BUs."
        )
        self.sync_pref_btn.clicked.connect(self._refresh_preferred_choices)
        vb_tools.addWidget(self.sync_pref_btn)

        vb_tools.addWidget(QLabel("Preferred BU:", w))
        self.preferred_bu_combo = QComboBox(w)
        self.preferred_bu_combo.addItem("(None)")
        for bu in sorted(POSSIBLE_BUS.keys()):
            self.preferred_bu_combo.addItem(bu)
        self.preferred_bu_combo.currentTextChanged.connect(
            lambda s: setattr(
                self, "preferred_bu", "" if s == "(None)" else s.strip().upper()
            )
        )
        vb_tools.addWidget(self.preferred_bu_combo)
        """vb_tools.addStretch(1)"""

        # --- Include unassigned toggle (NULL / blank BUs) ---
        self.chk_include_unassigned = QCheckBox("Include unassigned (NULL/blank)")
        self.chk_include_unassigned.setChecked(True)

        # Prefer to keep it with the BU tools on the right
        try:
            # if your right-side layout is named vb_tools
            vb_tools.addWidget(self.chk_include_unassigned)
        except NameError:
            # otherwise add under the BU group
            g_bus_layout.addWidget(self.chk_include_unassigned)
        vb_tools.addStretch(1)
        # ------- Options strip (BOM + Reserve + Clear) -------
        g_opt = QGroupBox("Options", w)
        ho = QHBoxLayout(g_opt)
        ho.setContentsMargins(8, 6, 8, 6)
        ho.setSpacing(10)

        self.bom_mode = QCheckBox("BOM mode (pivot)", w)
        self.bom_mode.setToolTip("Use imported BOM and show per-BU pivot.")
        self.bom_mode.setChecked(False)
        self.item_id.textEdited.connect(lambda _: self.bom_mode.setChecked(False))

        self.reserve_chk = QCheckBox("Reserve across BOM", w)
        self.reserve_chk.setChecked(False)
        self.reserve_chk.toggled.connect(
            lambda v: setattr(self, "reserve_across_bom", bool(v))
        )

        clear_bom_btn = QPushButton("Clear BOM", w)
        clear_bom_btn.setToolTip(
            "Forget imported BOM and switch to normal search mode."
        )
        clear_bom_btn.clicked.connect(self.clear_bom)

        ho.addWidget(self.bom_mode)
        ho.addWidget(self.reserve_chk)
        ho.addStretch(1)
        ho.addWidget(clear_bom_btn)

        # keep it compact vertically
        g_opt.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        g_opt.setMaximumHeight(56)

        # ----- Actions (top-right) -----
        import os  # used below for factory-defaults option

        g_actions = QGroupBox("Actions", w)
        ga = QVBoxLayout(g_actions)
        ga.setContentsMargins(8, 8, 8, 8)
        ga.setSpacing(6)

        # Buttons
        self.button = QPushButton("Search", w)
        self.sql_button = QPushButton("Show SQL", w)
        self.import_button = QPushButton("Import CSV", w)
        self.auth_button = QPushButton("Generate Credentials", w)
        self.save_filters_btn = QPushButton("Save Default Filters", w)

        # Wire actions
        self.button.clicked.connect(self.display_data)
        self.sql_button.clicked.connect(self.show_sql)
        self.import_button.clicked.connect(self.import_csv)
        self.auth_button.clicked.connect(self.creds_window)
        self.save_filters_btn.clicked.connect(
            self._save_filter_defaults
        )  # <-- ensure this method exists

        # Reset toolbutton with dropdown
        self.reset_btn = QToolButton(w)
        self.reset_btn.setText("Reset…")
        self.reset_btn.setToolButtonStyle(Qt.ToolButtonTextOnly)
        self.reset_btn.setPopupMode(QToolButton.MenuButtonPopup)

        menu = QMenu(self.reset_btn)
        act_all = menu.addAction("Reset to Saved Defaults (Filters + BUs)")
        act_filters = menu.addAction("Reset Filters Only")
        act_bus = menu.addAction("Reset BUs Only")
        act_factory = None
        # show only if machine defaults exist
        if os.path.exists(self._machine_prefs_path()):
            act_factory = menu.addAction("Reset to Factory Defaults (machine)")
        self.reset_btn.setMenu(menu)

        # Default click = reset both filters and BUs to saved user defaults
        def _reset_all():
            # <-- uses the merged prefs
            self.apply_psft_defaults(filters=True, bus=True)

        self.reset_btn.clicked.connect(_reset_all)
        act_all.triggered.connect(_reset_all)
        act_filters.triggered.connect(
            lambda: self.apply_psft_defaults(filters=True, bus=False)
        )
        act_bus.triggered.connect(
            lambda: self.apply_psft_defaults(filters=False, bus=True)
        )
        if act_factory:
            act_factory.triggered.connect(
                lambda: self.apply_psft_defaults(
                    filters=True,
                    bus=True,
                    prefs_override=self._load_machine_defaults_only(),
                )
            )

        # Add buttons to the vertical layout
        for b in (
            self.button,
            self.sql_button,
            self.import_button,
            self.auth_button,
            self.save_filters_btn,
            self.reset_btn,
        ):
            b.setMinimumHeight(28)
            b.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            ga.addWidget(b)

        ga.addStretch(1)

        # Keep the Actions box slim and content-height, and roughly match Filters height
        g_actions.setMaximumWidth(180)
        g_actions.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Maximum)
        QTimer.singleShot(
            0,
            lambda: g_actions.setFixedHeight(
                max(g_filters.sizeHint().height(), g_actions.sizeHint().height())
            ),
        )

        # Helpful tooltips showing where defaults are saved
        tip = f"Saves to: {self._user_prefs_path()}"
        self.save_filters_btn.setToolTip(tip)
        if hasattr(
            self, "btn_save_defaults"
        ):  # BU Tools button lives in a different group
            self.btn_save_defaults.setToolTip(tip)

        # Finally, place Actions in the grid (top-right, aligned to top)
        # (Do this where you add other groups to 'grid')
        # grid.addWidget(g_actions, 0, 1, alignment=Qt.AlignTop)

        # grid is your main layout for PSFT panel
        grid.addWidget(g_filters, 0, 0)  # left, row 0
        # right, row 0, top-aligned, no vertical stretch
        grid.addWidget(g_actions, 0, 1, alignment=Qt.AlignTop)
        grid.addWidget(g_bus, 1, 0)  # row 1, col 0
        # row 1, col 1 (will match bu height)
        grid.addWidget(g_butools, 1, 1)
        # Options strip spans BOTH columns
        grid.addWidget(g_opt, 2, 0, 1, 2)
        self.apply_psft_defaults(filters=True, bus=True)

        w.setLayout(grid)

        return w

    # ----- user prefs IO (reuse your existing ones if present) -----
    # ---------- Paths ----------
    def _machine_dir(self):
        import os

        return os.path.join(
            os.environ.get("PROGRAMDATA", r"C:\ProgramData"), "PartSearch"
        )

    def _user_dir(self):
        import os

        return os.path.join(
            os.environ.get("LOCALAPPDATA", os.path.expanduser(r"~\AppData\Local")),
            "PartSearch",
        )

    def _machine_prefs_path(self):
        import os

        return os.path.join(self._machine_dir(), "user_prefs.json")

    def _user_prefs_path(self):
        import os

        return os.path.join(self._user_dir(), "user_prefs.json")

    # ---------- IO utils ----------

    def _read_json_safely(self, path):
        try:
            if not os.path.exists(path):
                return {}
            with io.open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}

    # ---------- Public helpers you call ----------

    def _load_user_prefs(self) -> dict:
        r"""
        Load prefs with layered fallback:
          1) machine-wide defaults (read-only): C:\ProgramData\PartSearch\user_prefs.json
          2) per-user overrides: %LOCALAPPDATA%\PartSearch\user_prefs.json
        User prefs override machine defaults.
        """
        base = self._read_json_safely(self._machine_prefs_path())
        user = self._read_json_safely(self._user_prefs_path())
        base.update(user)
        return base

    # ----- save defaults for PSFT filters -----
    def _write_user_prefs(self, updates: dict) -> None:
        _write_user_prefs(updates)  # Call module-level helper

    def _save_filter_defaults(self):
        # collect the PSFT inputs only (NOT BU checkboxes)
        def _nz(s):  # normalize blanks to ""
            return (s or "").strip()

        try:
            min_qty_txt = _nz(self.qty.text())
            # keep as text; your SQL builder treats None as 'ignore'
            min_qty = None if min_qty_txt == "" else min_qty_txt
            prefs = self._load_user_prefs()
            defaults = {
                "mfg_part": _nz(self.mfg_part.text()),
                "and_wildcards": _nz(self.and_wildcards.text()),
                "either_or_wildcards": _nz(self.either_or_wildcards.text()),
                "or_wildcards": _nz(self.or_wildcards.text()),
                "min_qty": min_qty,  # None means "don’t include qty filter"
                "preferred_bu": getattr(self, "preferred_bu", "") or "",
                "reserve_across_bom": bool(self.reserve_chk.isChecked()),
            }
            prefs.setdefault("psft_defaults", {})
            prefs["psft_defaults"] = defaults
            self._write_user_prefs(prefs)
            QMessageBox.information(self, "Defaults", "PSFT filter defaults saved.")
        except Exception as e:
            QMessageBox.warning(self, "Defaults", f"Could not save defaults:\n{e}")

    # ----- save pref bus and file search roots -----
    def _save_user_prefs(self):
        """Save BU checkbox selections, preferred BU, and file-search roots (if available)."""
        prefs = self._load_user_prefs()

        # 1) Business Units (checkboxes)
        try:
            bu_checked = [
                b.text().strip().upper()
                for b in getattr(self, "BUs", [])
                if b.isChecked()
            ]
        except Exception:
            bu_checked = []
        prefs["bus_defaults"] = bu_checked

        # 2) Preferred BU (optional)
        pb = getattr(self, "preferred_bu", "") or ""
        prefs.setdefault("psft_defaults", {})
        prefs["psft_defaults"]["preferred_bu"] = pb

        # 3) File Search: allowed roots (if pane exists)
        try:
            if hasattr(self, "files_panel"):
                roots = getattr(self.files_panel, "allowed_roots", None)
                if roots is None and hasattr(self.files_panel, "locations"):
                    # fallback if you store a list of checkable paths under a different name
                    roots = [
                        p
                        for p, checked in getattr(self.files_panel, "locations")
                        if checked
                    ]
                if roots is not None:
                    prefs.setdefault("file_inputs", {})
                    prefs["file_inputs"]["allowed_roots"] = roots
        except Exception:
            pass

        self._write_user_prefs(prefs)
        try:
            from PyQt5.QtWidgets import QMessageBox

            QMessageBox.information(self, "Defaults", "Defaults saved.")
        except Exception:
            pass

    # ----- apply defaults on startup / panel build -----
    def apply_psft_defaults(
        self, *, filters: bool = True, bus: bool = True, prefs_override: dict = None
    ) -> None:
        """
        Apply saved defaults to the PSFT pane in one pass.
        If prefs_override is provided, it should look like:
            {"psft_defaults": {...}, "bus_defaults": [...]}
        Otherwise we use the merged prefs (machine + user).
        """
        prefs = (
            prefs_override if prefs_override is not None else self._load_user_prefs()
        )
        d_psft = prefs.get("psft_defaults", {}) or {}
        d_bus = set(map(str.upper, prefs.get("bus_defaults", [])))

        # --- Apply BU defaults (checkboxes + preferred BU) ---
        if bus and hasattr(self, "BUs"):
            try:
                if d_bus:
                    for b in self.BUs:
                        label = b.text().strip().upper()
                        b.setChecked(label in d_bus)
                pb = (d_psft.get("preferred_bu") or "").strip().upper()
                if pb and hasattr(self, "preferred_bu_combo"):
                    self.preferred_bu_combo.blockSignals(True)
                    idx = self.preferred_bu_combo.findText(pb)
                    if idx >= 0:
                        self.preferred_bu_combo.setCurrentIndex(idx)
                    self.preferred_bu_combo.blockSignals(False)
                    self.preferred_bu = pb
            except Exception:
                pass

        # --- Apply filter defaults (wildcards, min qty, reserve) ---
        if filters:
            try:

                def _nz(v):
                    return (v or "").strip()

                if hasattr(self, "mfg_part"):
                    self.mfg_part.setText(_nz(d_psft.get("mfg_part")))
                if hasattr(self, "and_wildcards"):
                    self.and_wildcards.setText(_nz(d_psft.get("and_wildcards")))
                if hasattr(self, "either_or_wildcards"):
                    self.either_or_wildcards.setText(
                        _nz(d_psft.get("either_or_wildcards"))
                    )
                if hasattr(self, "or_wildcards"):
                    self.or_wildcards.setText(_nz(d_psft.get("or_wildcards")))
                if hasattr(self, "qty"):
                    mq = d_psft.get("min_qty", None)
                    (
                        self.qty.clear()
                        if (mq in (None, ""))
                        else self.qty.setText(str(mq).strip())
                    )
                if hasattr(self, "reserve_chk"):
                    self.reserve_chk.blockSignals(True)
                    self.reserve_chk.setChecked(
                        bool(d_psft.get("reserve_across_bom", False))
                    )
                    self.reserve_chk.blockSignals(False)
            except Exception:
                pass

    # Back-compat wrappers (if anything still calls the old names)
    def _apply_filter_defaults(self):
        self.apply_psft_defaults(filters=True, bus=False)

    def _apply_bu_defaults(self):
        self.apply_psft_defaults(filters=False, bus=True)

    def _load_machine_defaults_only(self) -> dict:
        """Read machine 'factory' defaults without merging user overrides."""
        return self._read_json_safely(self._machine_prefs_path())

    # --- Back-compat wrappers (keep existing callers working) ---
    def _refresh_preferred_choices(self):
        checked = [b.text().strip().upper() for b in self.BUs if b.isChecked()]
        cur = self.preferred_bu_combo.currentText().strip().upper()
        self.preferred_bu_combo.blockSignals(True)
        self.preferred_bu_combo.clear()
        self.preferred_bu_combo.addItem("(None)")
        for bu in checked:
            self.preferred_bu_combo.addItem(bu)
        # keep current if still valid
        if cur and cur in checked:
            idx = self.preferred_bu_combo.findText(cur)
            if idx >= 0:
                self.preferred_bu_combo.setCurrentIndex(idx)
        self.preferred_bu_combo.blockSignals(False)

    def clear_bom(self):
        self.bom_map = {}
        self.csv_desc_map = {}
        self.import_order = []
        if self.bom_mode:
            self.bom_mode.setChecked(False)
        QMessageBox.information(
            self,
            "BOM cleared",
            "BOM context cleared. You are now in normal search mode.",
        )

    def toggle_BUs(self):
        check_val = not all([box.isChecked() for box in self.BUs])
        for box in self.BUs:
            box.setChecked(check_val)
        self.toggle_all.setText("Uncheck All" if check_val else "Check All")
        self._refresh_preferred_choices()

    def display_data(self):
        import re

        import pandas as pd

        # ---------- helpers ----------
        def _safe_line_text(widget_name: str) -> str:
            w = getattr(self, widget_name, None)
            if w is None:
                return ""
            # QLineEdit / QTextEdit support
            try:
                return w.text()
            except Exception:
                try:
                    return w.toPlainText()
                except Exception:
                    return ""

        def _parse_list(s: str, sep=";"):
            return [t.strip() for t in (s or "").split(sep) if t and t.strip()]

        def _parse_int_or_none(s: str):
            s = (s or "").strip()
            if s == "":
                return None
            try:
                return int(s)
            except Exception:
                return None

        def _get_checked_bus():
            bus = []
            for box in getattr(self, "BUs", []):
                try:
                    if box.isChecked():
                        bus.append(box.text().strip().upper())
                except Exception:
                    pass
            return bus

        # ---------- read UI ----------
        checked_BUs = _get_checked_bus()

        bom_on = bool(getattr(self, "bom_mode", None) and self.bom_mode.isChecked())

        # include_unassigned checkbox (default True if absent)
        include_unassigned = True
        if hasattr(self, "chk_include_unassigned"):
            try:
                include_unassigned = bool(self.chk_include_unassigned.isChecked())
            except Exception:
                include_unassigned = True

        # Preferred BU (optional)
        preferred_bu = (getattr(self, "preferred_bu", "") or "").strip().upper()

        # Reserve-across-BOM (optional)
        reserve_across_bom = bool(
            getattr(self, "reserve_across_bom_chk", None)
            and self.reserve_across_bom_chk.isChecked()
        )

        # Wildcards & fields (ignored in BOM mode except for item_id list)
        if bom_on:
            and_wild = []
            or_wild = []
            either_or = []
            # Build bom_map from previously parsed CSV or from Part No field
            bom_map = {}
            if isinstance(getattr(self, "bom_map", None), dict) and self.bom_map:
                bom_map = dict(self.bom_map)
            else:
                # parse from a multiline/CSV field—prefer part_no_edit if present
                raw = _safe_line_text("part_no_edit")
                if not raw:
                    raw = _safe_line_text("item_id")
                toks = [t.strip() for t in re.split(r"[,\s]+", raw or "") if t.strip()]
                bom_map = {t.upper(): 1 for t in toks}
            item_ids = list(bom_map.keys())
            min_qty = None  # do NOT apply min-qty filter in BOM mode
        else:
            # normal mode
            and_wild = _parse_list(_safe_line_text("and_wildcards"))
            or_wild = _parse_list(_safe_line_text("or_wildcards"))
            # either_or: "a,b;c,d" -> [["a","b"],["c","d"]]
            eo_raw = _parse_list(_safe_line_text("either_or_wildcards"))
            either_or = (
                [[e.strip() for e in g.split(",") if e.strip()] for g in eo_raw]
                if eo_raw
                else []
            )

            item_ids = _parse_list(_safe_line_text("item_id"))
            # Allow mfg_part if you expose it
            mfg_parts = _parse_list(_safe_line_text("mfg_part"))
            min_qty = _parse_int_or_none(_safe_line_text("qty"))
            bom_map = {}  # none in normal mode

        # ---------- credentials check ----------
        try:
            _ = get_basic_auth_header("Denodo")  # raises if creds missing
        except Exception as e:
            QMessageBox.warning(
                self,
                "Credentials",
                "Denodo credentials are not available.\n\n"
                "Click 'Generate Credentials' and save your OII username/password.\n\n"
                f"Details: {e}",
            )
            return

        # ---------- data fetch ----------
        try:
            result = items_lookup(
                BUs=checked_BUs,
                item_id=item_ids,
                # mfg_part only when not in BOM mode
                mfg_part=(
                    _parse_list(_safe_line_text("mfg_part")) if not bom_on else []
                ),
                and_wildcards=(and_wild if not bom_on else []),
                either_or_wildcards=(either_or if not bom_on else []),
                or_wildcards=(or_wild if not bom_on else []),
                min_qty=min_qty,
                include_unassigned=include_unassigned,
            )
        except Exception as e:
            QMessageBox.critical(self, "PSFT Lookup Failed", str(e))
            return

        # ---------- normalize result ----------
        out, per_loc, inv_raw = (pd.DataFrame(), pd.DataFrame(), pd.DataFrame())
        if isinstance(result, tuple) and len(result) >= 3:
            out, per_loc, inv_raw = result
        elif isinstance(result, pd.DataFrame):
            out = result

        if not isinstance(out, pd.DataFrame):
            out = pd.DataFrame()

        # Friendly notice on empty
        if out.empty:
            QMessageBox.information(
                self,
                "PSFT Search",
                "No results were returned.\n\nTips:\n"
                "• Verify Business Units selection (or check 'Include unassigned')\n"
                "• Adjust wildcards (AND / Either-Or / OR)\n"
                "• Try clearing Min Qty if set\n"
                "• Re-save Denodo creds if this persists",
            )

        # ---------- build and show DataWindow ----------
        # Clean up any previous results panel
        try:
            if getattr(self, "results_panel", None):
                try:
                    self.stack.removeWidget(self.results_panel)
                except Exception:
                    pass
                self.results_panel.deleteLater()
        except Exception:
            pass

        # Currency config if you have it
        target_currency = (getattr(self, "target_currency", "USD") or "USD").upper()
        fx_rates = getattr(self, "fx_rates", {"USD": 1.0})

        if bom_on and bom_map:
            dw = DataWindow(
                df=out,
                per_loc=per_loc,
                inv_raw=inv_raw,
                bom_map=bom_map,
                selected_bus=checked_BUs,
                import_order=list(bom_map.keys()),
                csv_desc_map=getattr(self, "csv_desc_map", {}),
                bom_mode=True,
                target_currency=target_currency,
                fx_rates=fx_rates,
                preferred_bu=preferred_bu,
                reserve_across_bom=reserve_across_bom,
                parent=self,
            )
        else:
            dw = DataWindow(
                df=out,
                per_loc=per_loc,
                inv_raw=inv_raw,
                bom_map={},  # no pivot
                selected_bus=checked_BUs,
                import_order=None,
                csv_desc_map={},
                bom_mode=False,
                target_currency=target_currency,
                fx_rates=fx_rates,
                preferred_bu=preferred_bu,
                reserve_across_bom=False,
                parent=self,
            )

        # callbacks for navigation and PSFT↔File bridge
        dw.on_back = lambda: self._switch_mode(0)  # back to PSFT input pane
        dw.on_open_files = lambda pn: self.open_files_for_pn(pn)

        # mount into the stacked widget
        self.results_panel = dw
        try:
            self.stack.addWidget(self.results_panel)
        except Exception:
            # ignore if already added
            pass
        self.stack.setCurrentWidget(self.results_panel)

        # keep left-rail toggle visually on PSFT
        try:
            if hasattr(self, "nav_psft"):
                self.nav_psft.setChecked(True)
        except Exception:
            pass

    def creds_window(self):
        self.auth_window = AuthWindow()
        self.auth_window.show()

    def show_sql(self):
        checked_BUs = [box.text() for box in self.BUs if box.isChecked()]
        item_ids = (
            [i.strip() for i in self.item_id.text().split(";")]
            if self.item_id.text()
            else []
        )
        mfg_parts = (
            [m.strip() for m in self.mfg_part.text().split(";")]
            if self.mfg_part.text()
            else []
        )
        and_wild = (
            [a.strip() for a in self.and_wildcards.text().split(";")]
            if self.and_wildcards.text()
            else []
        )
        either_or = (
            [e.split(",") for e in self.either_or_wildcards.text().split(";")]
            if self.either_or_wildcards.text()
            else []
        )
        or_wild = (
            [o.strip() for o in self.or_wildcards.text().split(";")]
            if self.or_wildcards.text()
            else []
        )
        min_qty = int(self.qty.text()) if self.qty.text().strip() != "" else None
        sql = build_equivalent_sql(
            item_ids,
            mfg_parts,
            and_wild,
            either_or,
            or_wild,
            list(checked_BUs),
            min_qty,
        )
        self.sql_window = SqlWindow(sql)
        self.sql_window.show()

    """def open_fx_window(self):
        try:
            self.fx_win = FxWindow(self, self.fx_rates.copy())
            
            self.fx_win.setWindowModality(Qt.ApplicationModal)
            self.fx_win.show()
            self.fx_win.raise_()
            self.fx_win.activateWindow()
        except Exception as e:
            QMessageBox.critical(self, 'Rates', f'Failed to open Rates window:\n{e}')"""

    def open_file_search(self):
        self._switch_mode(1)
        self.files_panel.q.setFocus()

    def open_files_for_pn(self, pn: str):
        if not getattr(self, "files_enabled", True):
            QMessageBox.information(
                self, "File Search", "File search is disabled by policy."
            )
            return
        # switch to File pane, show "Back to PSFT Results" if a results pane exists
        self._switch_mode(1)
        res_idx = self.stack.indexOf(getattr(self, "results_panel", None))
        self.files_panel.enable_back_to_results(
            res_idx != -1, cb=lambda: self._switch_mode(res_idx)
        )
        # auto-populate and auto-run
        self.files_panel.q.setText(str(pn or "").strip())
        self.files_panel.run_search()

    def import_csv(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Import CSV", "", "CSV Files (*.csv)"
        )
        if not path:
            return

        # Try common encodings (some CAD/BOM exports include a UTF-8 BOM)
        df = None
        for enc in (None, "utf-8-sig", "latin1"):
            try:
                df = (
                    pd.read_csv(path, dtype=str, keep_default_na=False, encoding=enc)
                    if enc
                    else pd.read_csv(path, dtype=str, keep_default_na=False)
                )
                break
            except Exception:
                df = None
        if df is None:
            QMessageBox.critical(
                self, "Import failed", "Could not read CSV with common encodings."
            )
            return
        if df.empty:
            QMessageBox.information(self, "Import CSV", "CSV is empty.")
            return

        # --- header normalization ---
        norm = {c: c.lower().replace(" ", "").replace("_", "") for c in df.columns}
        candidates = {k: v for k, v in norm.items()}

        # part number column
        pn_col = None
        pn_wants = [
            "itemid",
            "item",
            "partnumber",
            "pn",
            "part",
            "partno",
            "partid",
            "itemno",
            "pn#",
            "filename",
            "drawingno",
            "drawingnumber",
            "drawing",
            "dwgno",
            "dwg",
        ]
        for c, n in candidates.items():
            if n in pn_wants:
                pn_col = c
                break
        if pn_col is None:
            # last-resort heuristic: column name containing a useful token
            for c in df.columns:
                lc = c.lower()
                if any(tok in lc for tok in ("drawing", "dwg", "part", "item")):
                    pn_col = c
                    break
        if pn_col is None:
            QMessageBox.warning(
                self,
                "Import CSV",
                "Couldn't find a part-number column.\n"
                "Looked for: " + ", ".join(pn_wants),
            )
            return

        # qty column (optional)
        qty_wants = {
            "qty",
            "quantity",
            "qtyrequired",
            "requiredqty",
            "need",
            "needed",
            "qty_need",
            "qtypurchased",
            "qtyreq",
            "required",
            "orderqty",
            "qtyordered",
            "qtyreqd",
        }
        qty_col = next((c for c, n in candidates.items() if n in qty_wants), None)

        # description column (optional, for OOS rows)
        desc_wants = {
            "description",
            "partdescription",
            "desc",
            "partdesc",
            "itemdescription",
            "part_description",
            "item_desc",
            "descr",
        }
        desc_col = next((c for c, n in candidates.items() if n in desc_wants), None)

        # --- normalize parts (strip SolidWorks extensions, uppercase) ---
        def _normalize_part(p: str) -> str:
            s = str(p).strip()
            s = re.sub(r"\.(sldprt|sldasm|slddrw)$", "", s, flags=re.IGNORECASE)
            return s.upper()

        parts_series = df[pn_col].astype(str).map(_normalize_part)
        parts_series = parts_series.replace("", pd.NA).dropna()
        parts = parts_series.tolist()

        # Debug aid (now correct)
        print("Headers seen:", list(df.columns))
        print("Matched part column:", pn_col, "Matched qty column:", qty_col)

        # If no parts, do not auto-run anything
        if not parts:
            self.bom_map = {}
            self.csv_desc_map = {}
            self.import_order = []
            self.item_id.clear()
            QMessageBox.information(
                self,
                "Import CSV",
                f"No parts found in '{os.path.basename(path)}' using column '{pn_col}'.",
            )
            return

        # Build BOM map (sum qty by PN)
        bom_map = {}
        if qty_col:
            tmp = df[[pn_col, qty_col]].copy()
            tmp[pn_col] = tmp[pn_col].astype(str).map(_normalize_part)
            tmp[qty_col] = pd.to_numeric(tmp[qty_col], errors="coerce")
            tmp = tmp.dropna(subset=[pn_col])
            grouped = tmp.groupby(pn_col)[qty_col].sum(min_count=1)
            grouped = grouped.fillna(0).astype(int).clip(lower=0)
            bom_map = grouped.to_dict()

        # Optional CSV descriptions (PN -> desc) for OOS rows
        csv_desc_map = {}
        if desc_col:
            dtmp = df[[pn_col, desc_col]].copy()
            dtmp[pn_col] = dtmp[pn_col].astype(str).map(_normalize_part)
            dtmp[desc_col] = dtmp[desc_col].astype(str).str.strip()
            csv_desc_map = {pn: d for pn, d in zip(dtmp[pn_col], dtmp[desc_col]) if pn}

        # persist on the form
        self.bom_map = bom_map
        self.csv_desc_map = csv_desc_map
        self.import_order = list(dict.fromkeys(parts_series.tolist()))
        if self.import_button:
            try:
                self.bom_mode.setChecked(True)
            except Exception:
                pass

        # Limit visible text but keep full BOM internally
        max_show = 2000
        self.item_id.setText(";".join(self.import_order[:max_show]))

        msg = f"Loaded {len(parts)} parts from '{os.path.basename(path)}'."
        if qty_col:
            msg += f" Using '{pn_col}' and '{qty_col}' as the part and quantity columns, respectively."
        QMessageBox.information(self, "Import OK", msg)

        # Only auto-run when we actually have parts
        # (keep your existing trigger here, e.g., self._on_submit_clicked() or similar)
        # If you prefer not to auto-run, just remove/comment the line below:
        # self._on_submit_clicked()


####


class SimpleTableModel(QAbstractTableModel):
    HEADERS = ["Name", "Type", "Size", "Modified", "Location", "Source", "FullPath"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.rows = []

    def rowCount(self, *a):
        return len(self.rows)

    def columnCount(self, *a):
        return len(self.HEADERS)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None
        if role in (Qt.DisplayRole, Qt.ToolTipRole):
            return self.rows[index.row()][index.column()]
        return None

    def headerData(self, sec, orient, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        return self.HEADERS[sec] if orient == Qt.Horizontal else str(sec + 1)

    def clear(self):
        self.beginResetModel()
        self.rows = []
        self.endResetModel()

    def append_chunk(self, chunk):
        if not chunk:
            return
        s = len(self.rows)
        self.beginInsertRows(QModelIndex(), s, s + len(chunk) - 1)
        self.rows.extend(chunk)
        self.endInsertRows()


class IndexBuilderPane(QWidget):
    """
    Manages the list of indexable roots (drives/folders) and builds the Quick Index.
    This is the source of truth for roots. FileSearchPane only filters which of
    these roots to use when searching.
    """

    root_started = pyqtSignal(str)
    root_progress = pyqtSignal(str, int, int)
    root_done = pyqtSignal(str, int, int, float)
    all_done = pyqtSignal()
    error = pyqtSignal(str)

    COL_CHECK = 0
    COL_PATH = 1
    COL_FILES = 2
    COL_UPDATED = 3
    COL_LAST = 4
    COL_STATUS = 5

    def __init__(self, parent=None, policy=None):
        super().__init__(parent)
        self.policy = policy or {}
        self.worker = None
        self._build_ui()
        self._load_roots_into_table()
        self._indexing_roots = getattr(self, "_indexing_roots", set())

    # ---------- JSON paths & I/O (inside IndexBuilderPane) ----------
    def _json_dir(self):
        return os.path.dirname(self._locations_path())

    def _prefill_from_db(self, row: int, path: str):
        # read actual counts from DB and stamp the row
        files = indexed_file_count_for_root(path)
        last = last_full_scan_for_root(path)
        self.table.setItem(row, self.COL_FILES, QTableWidgetItem(f"{files:,}"))
        self.table.setItem(row, self.COL_UPDATED, QTableWidgetItem("0"))
        self.table.setItem(row, self.COL_LAST, QTableWidgetItem(last))

    def _load_index_roots(self):
        try:
            with open(self._index_roots_path(), "r", encoding="utf-8") as f:
                obj = json.load(f)
                # expected shape: {"roots":[{"path": "...", "last_full_scan": "...", "files_count": 0}]}
                if isinstance(obj, dict) and isinstance(obj.get("roots"), list):
                    return obj["roots"]
        except Exception:
            pass
        return []  # empty list means no roots yet

    def _save_index_roots(self, roots):
        path = self._index_roots_path()
        tmp = path + ".tmp"
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({"roots": roots}, f, indent=2)
        os.replace(tmp, path)

    def _locations_path(self):
        return str(app_paths()["locations"])

    def _save_locations_checked(self, checked_paths: list[str]):
        # Persist "which locations should File Search use" as a simple list
        path = self._locations_path()
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump({"checked_roots": list(checked_paths or [])}, f, indent=2)
        os.replace(tmp, path)

    def _refresh_totals(self):
        files = 0
        updated = 0

        def _to_int(it):
            try:
                return int((it.text() or "0").replace(",", ""))
            except Exception:
                return 0

        for r in range(self.table.rowCount()):
            files += _to_int(self.table.item(r, self.COL_FILES))
            updated += _to_int(self.table.item(r, self.COL_UPDATED))

        if hasattr(self, "lbl_totals") and self.lbl_totals:
            self.lbl_totals.setText(
                f"Indexed files: {files:,}   •   Updated last run: {updated:,}"
            )

    def _index_roots_path(self):
        return str(app_paths()["index_roots"])

    def _read_index_roots_json(self) -> list[dict]:
        """
        Read saved roots. Supports either:
          - {"roots": [ {...}, {...} ]}
          - [ {...}, {...} ]   (legacy)
        Filters out blank/duplicate paths.
        """
        p = self._index_roots_path()
        if not os.path.exists(p):
            return []

        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return []

        roots = []
        if isinstance(data, dict) and isinstance(data.get("roots"), list):
            roots = data["roots"]
        elif isinstance(data, list):
            roots = data
        else:
            return []

        # normalize: drop blanks + dedupe by lowercased path
        out, seen = [], set()
        for r in roots:
            path = (r.get("path") or "").strip()
            if not path:
                continue
            key = path.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(
                {
                    "path": path,
                    "checked": bool(r.get("checked", True)),
                    "files": int(r.get("files", 0) or 0),
                    "updated": int(r.get("updated", 0) or 0),
                    "last_full_scan": r.get("last_full_scan", "") or r.get("last", ""),
                    "status": r.get("status", "Idle"),
                }
            )
        return out

    def _collect_table_roots(self) -> list[dict]:
        """
        Harvest the current table rows for saving. Skips blank paths.
        Requires your COL_* constants and table items to be set.
        """
        rows, seen = [], set()
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, self.COL_PATH)
            if not pit:
                continue
            path = (pit.text() or "").strip()
            if not path:
                continue

            key = path.lower()
            if key in seen:
                continue
            seen.add(key)

            chk = self.table.item(r, self.COL_CHECK)
            checked = (chk.checkState() == Qt.Checked) if chk else True

            def _get(col, default=""):
                it = self.table.item(r, col)
                return it.text() if it else default

            files = int((_get(self.COL_FILES, "0") or "0").replace(",", "") or 0)
            updated = int((_get(self.COL_UPDATED, "0") or "0").replace(",", "") or 0)
            last = _get(self.COL_LAST, "")
            status = _get(self.COL_STATUS, "Idle")

            rows.append(
                {
                    "path": path,
                    "checked": checked,
                    "files": files,
                    "updated": updated,
                    "last_full_scan": last,
                    "status": status,
                }
            )
        return rows

    def _write_index_roots_json(self) -> None:
        """Atomically write the current roots to disk."""
        p = self._index_roots_path()
        data = {"roots": self._collect_table_roots()}
        try:
            tmp = p + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            os.replace(tmp, p)
        except Exception as e:
            # Non-fatal: we don't block UI; print to console for debugging
            print("index_roots write error:", e)

    def _load_roots_into_table(self):
        """
        Load roots into the table, sanitize & de-dupe, then ALWAYS prefill
        Files/Last from the per-root DB. Finally, persist the refreshed view
        back to index_roots.json so the next session shows the correct data.
        """
        # 1) Read saved roots; shape: {"roots":[ {...} ]}  (module-level helper)
        try:
            data = read_index_roots_json()  # returns {"roots":[...]}
        except Exception:
            data = {"roots": []}

        raw_roots = list(data.get("roots", []) or [])

        # If nothing saved yet, fall back to sensible defaults (drives, etc.)
        if not raw_roots and hasattr(self, "_discover_default_roots"):
            try:
                for r in self._discover_default_roots() or []:
                    raw_roots.append(
                        {
                            "path": (r.get("path") or "").strip(),
                            "checked": bool(r.get("checked", True)),
                            "files_count": 0,
                            "updated_count": 0,
                            "last_full_scan": "",
                        }
                    )
            except Exception:
                pass

        # 2) Sanitize & de-dupe by normalized path (case-insensitive on Windows)
        cleaned = []
        seen = set()
        for r in raw_roots:
            path = str(r.get("path", "")).strip()
            if not path:
                continue
            key = os.path.normcase(os.path.normpath(path))
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(
                {
                    "path": path,
                    "checked": bool(r.get("checked", True)),
                    # accept either old ("files"/"updated") or new ("files_count"/"updated_count") keys
                    "files_count": int(r.get("files_count", r.get("files", 0)) or 0),
                    "updated_count": int(
                        r.get("updated_count", r.get("updated", 0)) or 0
                    ),
                    "last_full_scan": (r.get("last_full_scan") or r.get("last") or ""),
                }
            )

        # 3) Reset table & add rows
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)

        for rec in cleaned:
            # Paint the row with whatever was saved
            self._append_row(
                path=rec["path"],
                checked=rec["checked"],
                files_count=rec["files_count"],
                updated_count=rec["updated_count"],
                last=rec["last_full_scan"],
                status="Idle",
            )

            # ALWAYS refresh from DB to show authoritative values on load
            # (uses indexed_file_count_for_root / last_full_scan_for_root)
            row_ix = self.table.rowCount() - 1
            try:
                self._prefill_from_db(row_ix, rec["path"])  # stamps Files/Last
            except Exception:
                # Don't crash UI if a path is temporarily unavailable
                pass

        self.table.setSortingEnabled(True)
        if hasattr(self, "_refresh_totals"):
            self._refresh_totals()

        # 4) Persist the refreshed rows back to disk so the next session opens clean
        try:
            self._write_index_roots_json()
        except Exception:
            # Non-fatal; UI already shows the refreshed state
            pass

    def _discover_default_roots(self) -> list:
        out = []
        # Physical drives A:..Z: that actually exist
        import string

        for d in string.ascii_uppercase:
            root = f"{d}:\\"
            if os.path.isdir(root):
                out.append({"path": root, "checked": d in ("C", "J", "S")})
        # Add any policy-required vault folders here if you have them
        return out

    def _append_row(self, path: str, **kw):
        """
        Adds one row to the table.

        Accepts either:
          files_count/updated_count   (new)
          files/updated               (old)
        """
        if not path:
            return

        # tolerate both key styles
        checked = bool(kw.get("checked", True))
        files = int(kw.get("files_count", kw.get("files", 0)) or 0)
        updated = int(kw.get("updated_count", kw.get("updated", 0)) or 0)
        last = kw.get("last", kw.get("last_full_scan", "")) or ""
        status = kw.get("status", "Idle")

        r = self.table.rowCount()
        self.table.insertRow(r)

        chk = QTableWidgetItem()
        chk.setCheckState(Qt.Checked if checked else Qt.Unchecked)
        self.table.setItem(r, self.COL_CHECK, chk)

        self.table.setItem(r, self.COL_PATH, QTableWidgetItem(path))
        self.table.setItem(r, self.COL_FILES, QTableWidgetItem(f"{files:,}"))
        self.table.setItem(r, self.COL_UPDATED, QTableWidgetItem(f"{updated:,}"))
        self.table.setItem(r, self.COL_LAST, QTableWidgetItem(last))
        self.table.setItem(r, self.COL_STATUS, QTableWidgetItem(status))

    def _gather_roots_from_table(self) -> list[dict]:
        out = []
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, self.COL_PATH)
            if not pit:
                continue
            path = (pit.text() or "").strip()
            if not path:
                continue
            chk = self.table.item(r, self.COL_CHECK)
            files_it = self.table.item(r, self.COL_FILES)
            last_it = self.table.item(r, self.COL_LAST)

            out.append(
                {
                    "path": path,
                    "checked": bool(chk and chk.checkState() == Qt.Checked),
                    "files_count": int(
                        ((files_it.text() if files_it else "0").replace(",", "")) or 0
                    ),
                    "last_full_scan": (last_it.text() if last_it else ""),
                }
            )
        return out

    def _selected_paths_to_index(self) -> list[str]:
        paths = []
        for r in range(self.table.rowCount()):
            it = self.table.item(r, self.COL_CHECK)
            pit = self.table.item(r, self.COL_PATH)
            if not pit or not pit.text():
                continue
            if it and it.checkState() == Qt.Checked:
                paths.append(pit.text().strip())
        return paths

    def _build_totals_footer(self):
        """Footer with live totals."""
        bar = QHBoxLayout()
        self.lbl_totals = QLabel("Indexed files: 0   •   Updated last run: 0", self)
        self.lbl_totals.setMinimumHeight(22)
        self.lbl_totals.setTextInteractionFlags(Qt.TextSelectableByMouse)
        bar.addWidget(self.lbl_totals)
        bar.addStretch(1)
        return bar

    # ---------- UI ----------

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # Row of controls
        top = QHBoxLayout()
        self.btn_add = QPushButton("Add folders…")
        self.btn_remove = QPushButton("Remove selected")
        self.chk_incremental = QCheckBox("Incremental updates")
        self.btn_start = QPushButton("Start")
        self.btn_cancel = QPushButton("Cancel")
        self.btn_cancel.setEnabled(False)

        self.btn_add.clicked.connect(self._add_folders)
        self.btn_remove.clicked.connect(self._remove_selected)
        self.btn_start.clicked.connect(self._start_index)
        self.btn_cancel.clicked.connect(self._cancel_index)

        top.addWidget(self.btn_add)
        top.addWidget(self.btn_remove)
        top.addStretch(1)
        top.addWidget(self.chk_incremental)
        top.addWidget(self.btn_start)
        top.addWidget(self.btn_cancel)
        layout.addLayout(top)

        # Table
        self.table = QTableWidget(0, 6, self)
        self.table.setHorizontalHeaderLabels(
            ["", "Path", "Files", "Updated", "Last Indexed", "Status"]
        )
        self.table.horizontalHeader().setSectionResizeMode(
            self.COL_CHECK, QHeaderView.ResizeToContents
        )
        self.table.horizontalHeader().setSectionResizeMode(
            self.COL_PATH, QHeaderView.Stretch
        )
        self.table.horizontalHeader().setSectionResizeMode(
            self.COL_FILES, QHeaderView.ResizeToContents
        )
        self.table.horizontalHeader().setSectionResizeMode(
            self.COL_UPDATED, QHeaderView.ResizeToContents
        )
        self.table.horizontalHeader().setSectionResizeMode(
            self.COL_LAST, QHeaderView.ResizeToContents
        )
        self.table.horizontalHeader().setSectionResizeMode(
            self.COL_STATUS, QHeaderView.ResizeToContents
        )
        self.table.setSortingEnabled(True)
        layout.addWidget(self.table)

        layout.addLayout(self._build_totals_footer())

        # tip
        tip = QLabel(
            "• Use ‘Add folders…’ to register vaults/drives.\n"
            "• Check the rows you want to index, then click Start.\n"
            "• ‘Incremental updates’ reuses existing DB and only scans changes."
        )
        tip.setWordWrap(True)
        layout.addWidget(tip)

    # ---------- actions ----------
    def _add_folders(self):
        paths = QFileDialog.getExistingDirectory(
            self,
            "Add folder to index",
            "",
            QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks,
        )
        if not paths:
            return
        # Single select dialog; wrap to list
        to_add = [paths]

        # merge with saved
        current = read_index_roots_json().get("roots", [])
        byk = {_normkey(r["path"]): r for r in current}

        added_any = False
        for p in to_add:
            p = (p or "").strip()
            if not p:
                continue
            k = _normkey(p)
            if k not in byk:
                byk[k] = {
                    "path": p,
                    "checked": True,
                    "files_count": 0,
                    "updated_count": 0,
                    "last_full_scan": None,
                }
                added_any = True

        if added_any:
            write_index_roots_json({"roots": list(byk.values())})
            self._load_roots_into_table()

    """def _add_folders(self):
        """
    """Add one or more folders to the index list.
        - Dedupes paths already present (case-insensitive, normalized).
        - Inserts rows with default status.
        - Prefills Files / Last Indexed if an index DB already exists for that path.
        """
    """# Multi-select directory dialog (use non-native to allow multi-select on Windows)
        dlg = QFileDialog(self, "Add folders…")
        dlg.setFileMode(QFileDialog.Directory)
        dlg.setOption(QFileDialog.ShowDirsOnly, True)
        dlg.setOption(QFileDialog.DontUseNativeDialog, True)
        dlg.setDirectory("C:\\")  # start here; tweak if you want
     
        if dlg.exec_() != QFileDialog.Accepted:
            return
     
        candidates = dlg.selectedFiles() or []
        if not candidates:
            return
     
        # Build a set of existing normalized paths in the table
        def _norm(p: str) -> str:
            return os.path.normcase(os.path.normpath((p or "").strip()))
     
        existing = set()
        for r in range(self.table.rowCount()):
            pit = self.table.item(r, self.COL_PATH)
            if pit and pit.text():
                existing.add(_norm(pit.text()))
     
        # Disable sorting while inserting, so rows land where we expect
        was_sorting = self.table.isSortingEnabled()
        if was_sorting:
            self.table.setSortingEnabled(False)
     
        try:
            added = 0
            for raw in candidates:
                p = (raw or "").strip()
                if not p:
                    continue
                key = _norm(p)
                if key in existing:
                    # already listed -> skip
                    continue
                existing.add(key)
     
                # Append a new row with defaults
                self._append_row(p, files=0, updated=0, last="", status="Idle", checked=True)
                added += 1
     
                # Try to pre-fill stats from an existing per-root index DB
                try:
                    dbp = index_db_path(p)
                    if os.path.exists(dbp):
                        con = open_sqlite(dbp); cur = con.cursor()
     
                        # Files count
                        try:
                            cnt = cur.execute("SELECT COUNT(*) FROM files").fetchone()[0] or 0
                        except Exception:
                            cnt = 0
     
                        # Last full scan (prefer meta.val, fall back to meta.value if present)
                        last = ""
                        try:
                            row = cur.execute("SELECT val FROM meta WHERE key='last_full_scan'").fetchone()
                            last = (row[0] if row else "") or ""
                            if not last:
                                row = cur.execute("SELECT value FROM meta WHERE key='last_full_scan'").fetchone()
                                last = (row[0] if row else "") or ""
                        except Exception:
                            pass
     
                        con.close()
     
                        # Update the just-inserted row (last row)
                        r = self.table.rowCount() - 1
                        self.table.setItem(r, self.COL_FILES,   QTableWidgetItem(f"{int(cnt):,}"))
                        self.table.setItem(r, self.COL_UPDATED, QTableWidgetItem("0"))
                        self.table.setItem(r, self.COL_LAST,    QTableWidgetItem(last))
                except Exception:
                    # Prefill is best-effort; keep the row even if stats fail
                    pass
     
            if added and hasattr(self, "_refresh_totals"):
                self._refresh_totals()
     
            # Persist immediately (only non-blank rows are written by _collect_table_roots)
            if hasattr(self, "_write_index_roots_json"):
                self._write_index_roots_json()
     
        finally:
            if was_sorting:
                self.table.setSortingEnabled(True)"""

    def _remove_selected(self):
        rows = sorted({ix.row() for ix in self.table.selectedIndexes()}, reverse=True)
        for r in rows:
            self.table.removeRow(r)
        self._save_index_roots(self._gather_roots_from_table())
        self._save_locations_checked(self._selected_paths_to_index())

    def _start_index(self):
        """Start indexing the checked rows."""
        # which paths are checked?
        checked_paths = self._selected_paths_to_index()
        if not checked_paths:
            QMessageBox.information(
                self, "Index", "Select at least one folder to index."
            )
            return

        dupes = [p for p in checked_paths if p in self._indexing_roots]
        if dupes:
            QMessageBox.warning(
                self, "Index", "Already indexing:\n" + "\n".join(map(str, dupes))
            )
            checked_paths = [p for p in checked_paths if p not in self._indexing_roots]
        if not checked_paths:
            return

        self._indexing_roots.update(checked_paths)

        # UI state
        if hasattr(self, "btn_start"):
            self.btn_start.setEnabled(False)
        if hasattr(self, "btn_cancel"):
            self.btn_cancel.setEnabled(True)
        for p in checked_paths:
            r = self._row_for_path(p)
            if r is not None:
                self._set_row_status(r, "Scanning…")
                self._set_num(r, self.COL_UPDATED, 0)

        # create worker and wire signals
        incr = bool(
            getattr(self, "chk_incremental", None) and self.chk_incremental.isChecked()
        )
        self.worker = IndexWorker(roots=checked_paths, incremental=incr)
        self.worker.root_started.connect(self._on_root_started)
        self.worker.progress_root.connect(self._on_root_progress)
        self.worker.progress_path.connect(self._on_path_progress)
        self.worker.root_done.connect(self._on_root_done)
        self.worker.root_done.connect(
            lambda root, *_: self._indexing_roots.discard(root)
        )
        self.worker.all_done.connect(self._on_all_done)
        self.worker.all_done.connect(lambda: self._indexing_roots.clear())
        self.worker.error.connect(lambda e: QMessageBox.warning(self, "Indexing", e))

        if hasattr(self, "lbl_status"):
            self.lbl_status.setText("Indexing…")
        self.worker.start()

    def _cancel_index(self):
        if getattr(self, "worker", None):
            try:
                self.worker._stop = True
            except Exception:
                pass
        if hasattr(self, "btn_start"):
            self.btn_start.setEnabled(True)
        if hasattr(self, "btn_cancel"):
            self.btn_cancel.setEnabled(False)
        if hasattr(self, "lbl_status"):
            self.lbl_status.setText("Cancelling…")

    # ---------- signal handlers from worker ----------
    def _row_of_path(self, path: str) -> int:
        p = (path or "").strip().lower()
        for r in range(self.table.rowCount()):
            if self.table.item(r, self.COL_PATH).text().strip().lower() == p:
                return r
        return -1

    def _row_for_path(self, path: str):
        want = (path or "").strip().lower()
        for r in range(self.table.rowCount()):
            it = self.table.item(r, self.COL_PATH)
            if it and (it.text() or "").strip().lower() == want:
                return r
        return None

    def _set_num(self, row: int, col: int, val: int):
        it = self.table.item(row, col)
        if it is None:
            it = QTableWidgetItem()
            self.table.setItem(row, col, it)
        it.setText(f"{int(val):,}")

    def _set_text(self, row: int, col: int, txt: str):
        it = self.table.item(row, col)
        if it is None:
            it = QTableWidgetItem()
            self.table.setItem(row, col, it)
        it.setText(str(txt or ""))

    def _set_status(self, text: str):
        try:
            if hasattr(self, "status_label") and self.status_label:
                self.status_label.setText(str(text))
        except Exception:
            pass
        QApplication.processEvents()

    def _set_row_status(self, row: int, txt: str):
        """Write status text into the table's Status column for a given row."""
        self._set_text(row, self.COL_STATUS, txt)

    def _recalc_footer(self):
        files = 0
        updated = 0
        for r in range(self.table.rowCount()):
            try:
                f = self.table.item(r, self.COL_FILES)
                u = self.table.item(r, self.COL_UPDATED)
                files += int((f.text() or "0").replace(",", "")) if f else 0
                updated += int((u.text() or "0").replace(",", "")) if u else 0
            except Exception:
                pass
        if hasattr(self, "lbl_totals"):
            self.lbl_totals.setText(
                f"Indexed files: {files:,}   |   Updated this run: {updated:,}"
            )

    @pyqtSlot(str)
    def _on_root_started(self, root: str):
        r = self._row_for_path(root)
        if r is not None:
            self._set_row_status(r, "Scanning…")
            # reset updated shown for this run
            self._set_num(r, self.COL_UPDATED, 0)
        if hasattr(self, "lbl_status"):
            self.lbl_status.setText(f"Indexing {root}…")

    @pyqtSlot(str, int, int, float)
    def _on_root_progress(self, root: str, scanned: int, updated: int, secs: float):
        r = self._row_for_path(root)
        if r is not None:
            # live tick while scanning
            self._set_num(r, self.COL_FILES, scanned)
            self._set_num(r, self.COL_UPDATED, updated)
        if hasattr(self, "lbl_status"):
            self.lbl_status.setText(
                f"Indexing {root} — {scanned:,} scanned, {updated:,} upserts ({int(secs)}s)"
            )
        self._recalc_footer()

    @pyqtSlot(str)
    def _on_path_progress(self, path: str):
        if hasattr(self, "lbl_status"):
            self.lbl_status.setText(f"Scanning: {path}")

    @pyqtSlot(str, int, int, float)
    def _on_root_done(self, root: str, scanned: int, updated: int, secs: float):
        r = self._row_for_path(root)
        total = indexed_file_count_for_root(root)
        last = local_now_short()
        if r is not None:
            self.table.setItem(r, self.COL_FILES, QTableWidgetItem(f"{total:,}"))
            self.table.setItem(r, self.COL_UPDATED, QTableWidgetItem(f"{updated:,}"))
            self.table.setItem(r, self.COL_LAST, QTableWidgetItem(last))
            self.table.setItem(r, self.COL_STATUS, QTableWidgetItem("Idle"))
        # write JSON so restart shows real numbers
        if hasattr(self, "_write_index_roots_json"):
            self._write_index_roots_json()
        # footer
        if hasattr(self, "_refresh_totals"):
            self._refresh_totals()

    @pyqtSlot()
    def _on_all_done(self):
        if hasattr(self, "btn_start"):
            self.btn_start.setEnabled(True)
        if hasattr(self, "btn_cancel"):
            self.btn_cancel.setEnabled(False)
        if hasattr(self, "lbl_status"):
            self.lbl_status.setText("Indexing complete.")


class FileSearchPane(QWidget):
    """
    Drop-in replacement.
    Fixes:
      - duplicate 'use_qi' checkbox (caused overlap over 'Query:')
      - wrong arg name to SearchWorker (use_qi vs use_index)
      - finished/cancel UI not resetting (btn vs btn_search)
      - consistent status updates + enter-to-search
    """

    def __init__(self, parent=None, run_psft_callback=None, open_index_callback=None):
        super().__init__(parent)
        self.run_psft_callback = run_psft_callback
        self.open_index_cb = open_index_callback

        g = QGridLayout(self)

        # --- top row ---
        self.q = QLineEdit(self)
        self.q.setPlaceholderText('e.g. motor* hous* 316* or "motor housing"')
        self.q.returnPressed.connect(self.run_search)

        # keep original names to avoid ripples
        self.btn = QPushButton("Search", self)
        self.cancel = QPushButton("Cancel", self)  # keep original names
        self.cancel.setEnabled(False)

        self.back_btn = QPushButton("Back to PSFT Results", self)
        self.back_btn.setVisible(False)

        self.loc_btn = QPushButton("Locations…", self)
        self.loc_btn.clicked.connect(self.choose_locations)
        self.allowed_roots = read_search_locations().get("checked_roots", [])
        self._update_locations_label()

        self.tips_btn = QPushButton("Search Tips", self)
        self.tips_btn.clicked.connect(self.show_tips)

        g.addWidget(QLabel("Query:"), 0, 0)
        g.addWidget(self.q, 0, 1)
        g.addWidget(self.btn, 0, 2)
        g.addWidget(self.cancel, 0, 3)
        g.addWidget(self.back_btn, 0, 4)

        # --- right-hand options stack (tighter, vertical) ---
        ops_box = QGroupBox("Options", self)
        vb_ops = QVBoxLayout(ops_box)

        # IMPORTANT: define 'use_qi' ONCE and add it to a layout
        """self.use_qi    = QCheckBox("Use local Index", self); self.use_qi.setChecked(True)
        self.use_win   = QCheckBox("Windows Search (WDS)",  self); self.use_win.setChecked(False)
        self.use_crawl = QCheckBox("Direct Crawl (slow)",   self); self.use_crawl.setChecked(False)"""
        self.order_matters = QCheckBox("Order matters", self)
        self.order_matters.setChecked(False)

        self.manage_idx = QPushButton("Build/Manage Index…", self)
        if callable(self.open_index_cb):
            self.manage_idx.clicked.connect(self.open_index_cb)

        """vb_ops.addWidget(self.use_qi)
        vb_ops.addWidget(self.use_win)
        vb_ops.addWidget(self.use_crawl)"""
        vb_ops.addWidget(self.manage_idx)
        vb_ops.addSpacing(6)
        vb_ops.addWidget(self.order_matters)
        vb_ops.addSpacing(6)

        """# ENOVIA controls (left intact; enable/disable stays policy-driven)
        self.use_enovia       = QCheckBox("Use ENOVIA Online", self)
        self.enovia_open      = QPushButton("Open ENOVIA…", self)
        self.enovia_import    = QPushButton("Import ENOVIA CSV…", self)
        self.enovia_autowatch = QCheckBox("Auto-import ENOVIA CSV", self)
        self.enovia_timer     = QTimer(self); self.enovia_seen = set()
 
        try:
            cfg = enovia_config()
            enovia_ok = cfg["enabled"] and bool(cfg.get("search_url"))
        except Exception:
            cfg = {"index_enabled": False}
            enovia_ok = False
 
        self.use_enovia.setChecked(True)
        self.enovia_open.setEnabled(enovia_ok)
        self.enovia_import.setEnabled(bool(cfg.get("index_enabled")))
        self.enovia_autowatch.setEnabled(bool(cfg.get("index_enabled")))
        try:
            self.enovia_open.clicked.connect(self._enovia_open)
            self.enovia_import.clicked.connect(self._enovia_import_dialog)
            self.enovia_autowatch.toggled.connect(self._enovia_autowatch_toggled)
        except Exception:
            pass
 
        vb_ops.addWidget(self.use_enovia)
        vb_ops.addWidget(self.enovia_open)
        vb_ops.addWidget(self.enovia_import)
        vb_ops.addWidget(self.enovia_autowatch)"""
        vb_ops.addStretch(1)
        self.last_idx = QLabel("", self)
        self.last_idx.setWordWrap(True)
        self.last_idx.setStyleSheet("QLabel { font-size: 11px; color: #666; }")
        vb_ops.addSpacing(6)
        vb_ops.addWidget(self.last_idx)

        # status_label line
        self.status_label = QLabel("Ready", self)
        self.status_label.setObjectName("fileSearchStatus")
        self.status_label.setMinimumHeight(18)
        g.addWidget(self.status_label)

        # results table
        self.tbl = QTableView(self)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tbl.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tbl.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tbl.customContextMenuRequested.connect(self._menu)
        self.tbl.doubleClicked.connect(self._open_location_default)

        # place widgets
        g.addWidget(ops_box, 1, 2, 2, 3)  # right side, spans two rows
        g.addWidget(self.tips_btn, 1, 0, 1, 1)
        g.addWidget(self.loc_btn, 1, 1, 1, 1)
        g.addWidget(self.tbl, 2, 0, 1, 2)
        g.addWidget(self.status_label, 3, 0, 1, 5)

        self.model = SimpleTableModel(self)
        self.proxy = FileResultsProxy(self)
        self.proxy.setSourceModel(self.model)
        self.tbl.setModel(self.proxy)
        self.tbl.setSortingEnabled(True)
        self.tbl.sortByColumn(0, Qt.AscendingOrder)
        self.allowed_roots = []  # roots chosen for searching
        self.lbl_roots = QLabel("")  # optional small hint label

        vb_ops.addWidget(self.lbl_roots)
        self._load_saved_locations()  # populate from locations.json
        if hasattr(self, "_refresh_last_indexed_label"):
            self._refresh_last_indexed_label()

        # wire actions
        self.worker = None
        self.btn.clicked.connect(self.run_search)
        self.cancel.clicked.connect(self.cancel_search)

    # ----- core search flow -----
    def _locations_path(self):
        return str(app_paths()["locations"])

    def _update_locations_label(self):
        n = len(self.allowed_roots or [])
        self.loc_btn.setText(f"Locations… ({n} selected)")

    def _load_saved_locations(self):
        try:
            with open(self._locations_path(), "r", encoding="utf-8") as f:
                obj = json.load(f)
                self.allowed_roots = list(obj.get("checked_roots", []))
        except Exception:
            self.allowed_roots = []
        # small hint text
        if hasattr(self, "lbl_roots") and isinstance(self.lbl_roots, QLabel):
            n = len(self.allowed_roots)
            self.lbl_roots.setText(
                f"Using {n} location{'s' if n!=1 else ''}"
                if n
                else "No locations selected"
            )

    def _src_row(self, idx: QModelIndex) -> int:
        # Map proxy index to source row if needed
        try:
            if (
                idx
                and idx.isValid()
                and getattr(self, "proxy", None) is not None
                and idx.model() is self.proxy
            ):
                return self.proxy.mapToSource(idx).row()
        except Exception:
            pass
        return idx.row() if idx and idx.isValid() else -1

    def _open_locations_dialog(self):
        dlg = SearchLocationsDialog(self)
        if dlg.exec_() == QDialog.Accepted:
            self._load_saved_locations()  # refresh allowed_roots and hint

    def _set_status(self, text: str):
        try:
            self.status_label.setText(str(text or ""))
        except Exception:
            pass

    def run_search(self):
        q = (self.q.text() or "").strip()
        cfg = read_search_locations()
        allowed = cfg.get("checked_roots") or None
        if not q:
            QMessageBox.information(self, "File Search", "Enter a query.")
            return

        # 1) Load roots from the per-user locations file; None => search ALL DBs
        try:
            data = read_search_locations()  # now reads %APPDATA%
            allowed = data.get("checked_roots") or None
        except Exception:
            allowed = None

        # 2) Clear previous rows
        try:
            self.model.clear()
        except Exception:
            pass

        # 3) Spin up the worker (Index -> Windows -> Crawl)
        limit = 5000
        self._set_status(
            f"Searching across: {', '.join(allowed) if allowed else 'ALL indexed roots'}"
        )
        w = SearchWorker(
            q,
            use_win=True,
            use_qi=True,
            use_crawl=True,
            allowed_roots=allowed,
            limit=limit,
        )

        # 4) Wire signals to UI
        def _on_chunk(rows):
            try:
                self.model.append_chunk(rows)
            except Exception:
                # fallback, row-by-row if your model lacks append_chunk
                for r in rows:
                    try:
                        self.model.append_chunk([r])
                    except Exception:
                        pass

        w.chunk.connect(_on_chunk)
        w.progress.connect(self._set_status)
        w.error.connect(lambda e: self._set_status(f"Error: {e}"))
        w.done.connect(lambda rows: self._set_status(f"Done ({len(rows)} results)"))

        # Keep a ref so it isn’t GC’d
        self._worker = w
        w.start()

    def _populate_results(self, hits):
        """
        Normalize a list of hits (FileHit objects, dicts, or 7-col rows) into the
        table shape: ["Name","Type","Size","Modified","Location","Source","FullPath"].
        """
        import os
        import time

        from PyQt5.QtCore import QTimer
        from PyQt5.QtWidgets import QHeaderView

        def fmt_size(n):
            try:
                if n is None:
                    return ""
                n = int(n)
                units = ["B", "KB", "MB", "GB", "TB"]
                i = 0
                val = float(n)
                while val >= 1024 and i < len(units) - 1:
                    val /= 1024.0
                    i += 1
                # keep 0/1 decimals for human readability
                return (
                    f"{val:.0f} {units[i]}"
                    if val >= 10 or i == 0
                    else f"{val:.1f} {units[i]}"
                )
            except Exception:
                return ""

        def fmt_time(ts):
            try:
                return time.strftime("%Y-%m-%d %H:%M", time.localtime(float(ts)))
            except Exception:
                return ""

        rows = []
        for h in hits or []:
            # already a 7-col row?
            if isinstance(h, (list, tuple)) and len(h) >= 7:
                rows.append(list(h)[:7])
                continue

            # object/dict normalization
            get = (
                (lambda k, d=None: getattr(h, k, d))
                if not isinstance(h, dict)
                else (lambda k, d=None: h.get(k, d))
            )
            path = get("path", "") or get("fullpath", "") or get("FullPath", "")
            if not path:
                continue

            name = os.path.basename(path) or "(unnamed)"
            is_dir = bool(get("is_dir", False))
            size = get("size", None)
            mtime = get("mtime", None)
            src = get("source", "") or get("Source", "") or ""

            ext = os.path.splitext(name)[1][1:].lower()
            ftype = "Folder" if is_dir else (f"{ext.upper()} file" if ext else "File")

            rows.append(
                [
                    name,  # Name
                    ftype,  # Type
                    fmt_size(size),  # Size
                    fmt_time(mtime),  # Modified
                    os.path.dirname(path),  # Location
                    src or ("Index" if not is_dir else "Index"),  # Source
                    path,  # FullPath
                ]
            )

        # feed the model
        try:
            self.model.append_chunk(rows)
        except Exception:
            pass

        # present nicely: auto-size once, then allow free resizing; keep FullPath hidden but present
        try:
            hh = self.tbl.horizontalHeader()
            hh.setSectionResizeMode(QHeaderView.ResizeToContents)
            self.tbl.resizeColumnsToContents()
            QTimer.singleShot(
                0, lambda: hh.setSectionResizeMode(QHeaderView.Interactive)
            )
            # Hide FullPath (col 6) from view, keep it in the model for open/copy actions
            self.tbl.setColumnHidden(6, True)
        except Exception:
            pass

        # make sure context menu & double-click are wired (your helpers already exist)
        try:
            self.tbl.setContextMenuPolicy(Qt.CustomContextMenu)
            self.tbl.customContextMenuRequested.connect(self._menu)
        except Exception:
            pass
        try:
            self.tbl.doubleClicked.connect(self._open_location_default)
        except Exception:
            pass

    def cancel_search(self):
        if self.worker and self.worker.isRunning():
            self.status.setText("Cancelling…")
            try:
                self.worker.abort()
            except Exception:
                pass
        # UI resets in _on_worker_finished()

    # worker signal handlers
    def _on_worker_chunk(self, rows):
        self.model.append_chunk(rows)

    def _on_worker_done(self, rows):
        if self.worker and getattr(self.worker, "_abort", False):
            self.status.setText("Cancelled.")
        else:
            self.status.setText(f"Found {len(rows)} item(s).")

    def _on_worker_finished(self):
        # ALWAYS reset UI even if cancelled/errored
        self.btn.setEnabled(True)
        self.cancel.setEnabled(False)
        self.worker = None

    def _prefs_get(self):
        p = self.parent()
        if p and hasattr(p, "_load_user_prefs"):
            return p._load_user_prefs()
        return _prefs_read()

    def _prefs_set(self, data: dict):
        p = self.parent()
        if p and hasattr(p, "_write_user_prefs"):
            p._write_user_prefs(data)
        else:
            _prefs_write(data)

    def _load_saved_file_roots(self):
        sel = get_saved_roots("file_search_roots")
        self.allowed_roots = sel or None  # None = all indexed roots
        if hasattr(self, "_refresh_last_indexed_label"):
            self._refresh_last_indexed_label()

    def _save_file_roots(self):
        set_saved_roots("file_search_roots", list(self.allowed_roots or []))
        if hasattr(self, "_refresh_last_indexed_label"):
            self._refresh_last_indexed_label()

    def choose_locations(self):
        """
        Lets the user pick which *indexed* roots to search against.
        This reads from index_roots.json (set in IndexBuilderPane) and
        saves the user's selection into search_locations.json.
        """
        roots = read_index_roots_json().get("roots", [])
        if not roots:
            QMessageBox.information(
                self,
                "Locations",
                "No indexed folders found yet. Add folders and build the index first in the Quick Index tab.",
            )
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Choose search locations")
        lay = QVBoxLayout(dlg)

        info = QLabel("Select which registered locations you want to search.")
        info.setWordWrap(True)
        lay.addWidget(info)

        # List with checkboxes
        lst = QListWidget(dlg)
        lst.setSelectionMode(QAbstractItemView.NoSelection)
        lay.addWidget(lst, 1)

        # preselect from saved SEARCH_LOCATIONS_JSON
        saved = set(
            _normkey(p) for p in read_search_locations().get("checked_roots", [])
        )
        for r in roots:
            p = r["path"]
            it = QListWidgetItem(p)
            it.setFlags(it.flags() | Qt.ItemIsUserCheckable)
            it.setCheckState(
                Qt.Checked
                if _normkey(p) in saved or r.get("checked", True)
                else Qt.Unchecked
            )
            lst.addItem(it)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(btns)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)

        if dlg.exec_() != QDialog.Accepted:
            return

        chosen = []
        for i in range(lst.count()):
            it = lst.item(i)
            if it.checkState() == Qt.Checked:
                chosen.append(it.text())

        write_search_locations(chosen)
        # reflect in the pane
        self.allowed_roots = chosen[:]  # used by run_search / worker
        self._update_locations_label()

    def _load_saved_roots(self):
        try:
            prefs = self._prefs_get()
            roots = (prefs.get("file_search", {}) or {}).get("roots", [])
            return [os.path.normpath(r) for r in roots] if roots else None
        except Exception:
            return None

    def _save_roots(self):
        try:
            prefs = self._prefs_get()
            fs = prefs.setdefault("file_search", {})
            fs["roots"] = list(self.allowed_roots or [])
            self._prefs_set(prefs)
        except Exception:
            pass

    def _refresh_last_indexed_label(self):
        if getattr(self, "last_idx_label", None) is None:
            return
        roots = getattr(self, "allowed_roots", []) or []
        if len(roots) == 1:
            self.last_idx_label.setText(
                f"Last indexed: {get_last_indexed_text_for_root(roots[0])}"
            )
        else:
            self.last_idx_label.setText("Last indexed: —")

    # ----- Back to PSFT results toggle -----
    def enable_back_to_results(self, enabled: bool, cb=None):
        self.back_btn.setVisible(bool(enabled))
        try:
            self.back_btn.clicked.disconnect()
        except Exception:
            pass
        if enabled and cb:
            self.back_btn.clicked.connect(cb)

    # ----- context menu & helpers -----
    def _menu(self, pos):
        idx = self.tbl.indexAt(pos)
        if not idx.isValid():
            return
        r = self._src_row(idx)
        if r < 0:
            return
        # columns: ["Name","Type","Size","Modified","Location","Source","FullPath"]
        path = self.model.rows[r][6]
        name = self.model.rows[r][0]

        m = QMenu(self)
        m.addAction("Open", lambda: self._open_file(path))
        m.addAction("Open folder", lambda: self._open_folder(path))
        m.addSeparator()
        m.addAction("Copy full path", lambda: self._copy_text(path))
        m.addAction("Copy folder path", lambda: self._copy_text(os.path.dirname(path)))
        m.addSeparator()
        m.addAction("Search PSFT for this PN", lambda: self._psft_for_name(name))
        m.exec_(self.tbl.viewport().mapToGlobal(pos))

    def _psft_for_name(self, name):
        if callable(self.run_psft_callback):
            # best-effort: extract PN-like token from filename
            token = os.path.splitext(name)[0]
            self.run_psft_callback(token)

    def _open_location_default(self, idx):
        if not idx.isValid():
            return
        r = self._src_row(idx)
        if r < 0:
            return
        self._open_file(self.model.rows[r][6])

    def _copy_text(self, text):
        cb = QApplication.clipboard()
        cb.setText(text or "")

    def _open_folder(self, path):
        try:
            p = str(path or "")
            folder = p if os.path.isdir(p) else os.path.dirname(p)
            if folder and os.path.isdir(folder):
                os.startfile(folder)
            else:
                raise FileNotFoundError(folder)
        except Exception as e:
            QMessageBox.warning(
                self, "Open Folder", f"Could not open folder:\n{path}\n\n{e}"
            )

    def _open_file(self, path: str):
        try:
            p = str(path or "")
            # ENOVIA URL?
            if p.startswith("http://") or p.startswith("https://"):
                webbrowser.open(p)
                return
            # local folder or file
            if os.path.isdir(p):
                os.startfile(p)
            else:
                if os.path.exists(p):
                    os.startfile(p)
                else:
                    parent = os.path.dirname(p)
                    if parent and os.path.isdir(parent):
                        # highlight in Explorer if possible
                        try:
                            subprocess.Popen(["explorer", "/select,", p])
                        except Exception:
                            os.startfile(parent)
                    else:
                        raise FileNotFoundError(p)
        except Exception as e:
            QMessageBox.warning(self, "Open", f"Could not open:\n{path}\n\n{e}")

    def show_tips(self):
        tips = (
            "Search basics:\n"
            "• Tokens are ANDed:  motor 316  -> matches items containing both\n"
            '• Quotes for phrases:  "motor housing"\n'
            "• Wildcards:  foo* (prefix)   *bar (suffix)   *mid* (substring)\n"
            "• Order matters: enable the checkbox if sequence must match\n"
            "• Case is ignored for Quick Index & Windows index\n"
        )
        QMessageBox.information(self, "Search Tips", tips)


class AuthWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.init_UI()

    def init_UI(self):
        layout = QGridLayout()

        self.username = QLineEdit(self)
        self.username.setPlaceholderText("OII Username NOT email")
        layout.addWidget(QLabel("Username:"), 0, 0)
        layout.addWidget(self.username, 0, 1)

        self.password = QLineEdit(self)
        self.password.setPlaceholderText("OII Password")
        self.password.setEchoMode(QLineEdit.Password)
        layout.addWidget(QLabel("Password:"), 1, 0)
        layout.addWidget(self.password, 1, 1)

        self.button = QPushButton("Save", self)
        self.button.clicked.connect(self.get_creds)
        layout.addWidget(self.button, 2, 0)

        self.setLayout(layout)
        self.setWindowTitle("Authentication Info")
        self.resize(360, 120)

    def get_creds(self):
        user = self.username.text().strip()
        pwd = self.password.text()
        if not user or not pwd:
            QMessageBox.warning(self, "Missing", "Username and password are required.")
            return
        keyring.set_password("Denodo", user, pwd)
        QMessageBox.information(self, "Saved", f"Credentials stored for user {user}.")
        self.close()


class DataWindow(QWidget):
    """
    Results window.

    - If bom_map provided: PIVOT summary
        part_number | status | part_description | requested_qty | <BU...> | total_available | cost cols...
      (CSV import order preserved; row coloring can use model.status_col == 'status')

    - Otherwise: long view (one row per BU).

    Compatibility for existing callers:
      • self.table_view (QTableView)
      • self.table (alias to table_view)
      • self.view (alias to table_view)
      • self.model (PandasModel)
      • self.proxy / self._proxy (QSortFilterProxyModel)
      • self.on_back (callable the parent may set)
    """

    def __init__(
        self,
        df: pd.DataFrame,
        per_loc: pd.DataFrame,
        inv_raw: pd.DataFrame,
        bom_map: Optional[dict] = None,
        selected_bus: Optional[List[str]] = None,
        import_order: Optional[List[str]] = None,
        # PN -> description from CSV (optional)
        csv_desc_map: Optional[dict] = None,
        bom_mode: bool = False,
        target_currency: str = "USD",
        fx_rates: Optional[dict] = None,
        preferred_bu: str = "",
        reserve_across_bom: bool = False,
        parent=None,
    ):
        super().__init__(parent)
        self.df = df if isinstance(df, pd.DataFrame) else pd.DataFrame()
        self.per_loc = per_loc if isinstance(per_loc, pd.DataFrame) else pd.DataFrame()
        self.inv_raw = inv_raw if isinstance(inv_raw, pd.DataFrame) else pd.DataFrame()
        self.bom_map = bom_map or {}
        self.selected_bus = list(selected_bus or [])
        self.import_order = list(import_order or [])
        self.csv_desc_map = csv_desc_map or {}
        self.bom_mode = bool(bom_mode)
        self.target_currency = (target_currency or "USD").upper()
        self.fx_rates = fx_rates or {"USD": 1.0}
        self.current_view_df: Optional[pd.DataFrame] = None
        self.preferred_bu = (preferred_bu or "").upper()
        self.reserve_across_bom = bool(reserve_across_bom)

        self.on_back = None  # parent can set later
        self._build_ui()  # real UI builder

    # --- legacy alias so any straggler calls won't break
    def initUI(self):
        if not hasattr(self, "table_view"):
            self._build_ui()

    # --------------------- UI builder ---------------------
    def _build_ui(self):
        """
        Build the results UI:
          • Top ribbon (Back, Search Files, Export, Copy, Help)
          • Full-width table (sorted, proxy-backed)
          • Builds either BOM/PIVOT or LONG view, with status/cost helpers in BOM mode
        Relies on:
          self.df                (pd.DataFrame)
          self.per_loc           (pd.DataFrame) [optional]
          self.inv_raw           (pd.DataFrame) [optional]
          self.bom_map           (dict PN->requested_qty) [optional]
          self.selected_bus      (list[str]) [optional]
          self.import_order      (list[str]) [optional]
          self.csv_desc_map      (dict PN->desc) [optional]
          self.bom_mode          (bool)
          self.target_currency   (str)
          self.fx_rates          (dict like {'USD':1.0, 'EUR':1.08}) [optional]
          self.preferred_bu      (str) [optional]
          self.reserve_across_bom(bool) [optional]
        """

        # -------- local helpers used by this builder --------
        def _host_with(method_name: str):
            p = self.parent()
            while p is not None and not hasattr(p, method_name):
                p = p.parent()
            return p

        def _fx_convert_local(val, src_ccy, tgt_ccy, rates: dict):
            try:
                v = float(val or 0)
                s = (src_ccy or "").upper()
                t = (tgt_ccy or "").upper()
                if not s or not t or s == t:
                    return v
                rs = float(rates.get(s, 1.0))
                rt = float(rates.get(t, 1.0))
                if rt == 0:
                    return v
                return v * (rs / rt)
            except Exception:
                return float(val or 0)

        def _add(col: str, used: list, space: pd.DataFrame):
            if col in space.columns and col not in used:
                used.append(col)

        # -------- ribbon (single row) --------
        ribbon = QHBoxLayout()

        back = QPushButton("← Back to PSFT Search", self)
        back.clicked.connect(
            lambda: self.on_back() if callable(getattr(self, "on_back", None)) else None
        )

        open_files = QPushButton("Search Files for Selected PN", self)
        open_files.setToolTip(
            "Right-click a row for the same action. Uses the currently selected row."
        )

        btn_export = QPushButton("Export", self)
        btn_export.clicked.connect(self.save_data)

        btn_copy = QPushButton("Copy Selection", self)
        btn_copy.clicked.connect(self.copy_selection)

        help_btn = QPushButton("?", self)
        help_btn.setFixedWidth(28)
        help_btn.setToolTip(
            "Tips:\n"
            "• Click a column header to sort.\n"
            "• Right-click a row for actions.\n"
            "• Double-click a BU to see all parts in that BU.\n"
            "• Double-click a part for a BU summary.\n"
            "• Ctrl/Cmd+C to copy."
        )

        ribbon.addWidget(back)
        ribbon.addSpacing(8)
        ribbon.addWidget(open_files)
        ribbon.addStretch(1)
        ribbon.addWidget(btn_export)
        ribbon.addWidget(btn_copy)
        ribbon.addWidget(help_btn)

        # -------- build the DataFrame for this view --------
        base_df = (
            self.df.copy()
            if isinstance(getattr(self, "df", None), pd.DataFrame)
            else pd.DataFrame()
        )
        # normalize core columns if needed
        rename_map = {}
        if "item_id" in base_df.columns and "part_number" not in base_df.columns:
            rename_map["item_id"] = "part_number"
        if (
            "item_description" in base_df.columns
            and "part_description" not in base_df.columns
        ):
            rename_map["item_description"] = "part_description"
        if "business_unit" in base_df.columns and "BU" not in base_df.columns:
            rename_map["business_unit"] = "BU"
        if rename_map:
            base_df = base_df.rename(columns=rename_map)

        # coerce numeric qty/costs where present
        for c in (
            "available_quantity",
            "onhand_quantity",
            "reserved_quantity",
            "min_qty_reorder_point",
            "perpetual_avg_cost",
        ):
            if c in base_df.columns:
                base_df[c] = pd.to_numeric(base_df[c], errors="coerce")

        # --- choose path: BOM/PIVOT vs LONG ---
        if (
            bool(getattr(self, "bom_mode", False))
            and isinstance(getattr(self, "bom_map", None), dict)
            and self.bom_map
        ):
            # ----- BOM / PIVOT -----
            work = base_df.copy()
            # sanitize types
            if "available_quantity" not in work.columns:
                work["available_quantity"] = 0
            work["available_quantity"] = (
                pd.to_numeric(work["available_quantity"], errors="coerce")
                .fillna(0)
                .clip(lower=0)
            )

            # collapse dupes at (part, BU)
            grp = work.groupby(["part_number", "BU"], as_index=False).agg(
                {"available_quantity": "sum", "part_description": "first"}
            )

            # pivot into BU columns
            pivot = grp.pivot_table(
                index="part_number",
                columns="BU",
                values="available_quantity",
                aggfunc="sum",
                fill_value=0,
            )

            # reindex to csv import order if provided
            if (
                isinstance(getattr(self, "import_order", None), list)
                and self.import_order
            ):
                pivot = pivot.reindex(self.import_order, fill_value=0)

            # select BU columns
            bu_cols = list(getattr(self, "selected_bus", []) or [])
            if not bu_cols:
                bu_cols = sorted(pivot.columns.tolist())

            # ensure all BU columns exist & are ints
            for b in bu_cols:
                if b not in pivot.columns:
                    pivot[b] = 0
            pivot = pivot[bu_cols] if bu_cols else pivot
            for b in bu_cols:
                pivot[b] = (
                    pd.to_numeric(pivot[b], errors="coerce").fillna(0).astype(int)
                )

            # assemble summary
            df_sum = pivot.copy()
            df_sum["part_number"] = df_sum.index.astype(str)
            df_sum.reset_index(drop=True, inplace=True)

            # requested qty from bom_map
            rq = pd.Series(self.bom_map, name="requested_qty")
            df_sum = df_sum.join(rq, on="part_number", how="left")
            df_sum["requested_qty"] = (
                pd.to_numeric(df_sum["requested_qty"], errors="coerce")
                .fillna(0)
                .astype(int)
            )

            # total available across selected BUs
            df_sum["total_available"] = (
                df_sum[bu_cols].sum(axis=1).astype(int) if bu_cols else 0
            )

            # availability map for allocation
            avail_map = {}
            for _, r in base_df.iterrows():
                pn = str(r.get("part_number") or "").strip()
                bu = str(r.get("BU") or "").strip().upper()
                if not pn or not bu:
                    continue
                a = int(
                    pd.to_numeric(r.get("available_quantity"), errors="coerce") or 0
                )
                avail_map.setdefault(pn, {}).setdefault(bu, 0)
                avail_map[pn][bu] += max(0, a)

            tgt = (getattr(self, "target_currency", "USD") or "USD").upper()
            rates = getattr(self, "fx_rates", None) or {"USD": 1.0}
            pref_bu = (getattr(self, "preferred_bu", "") or "").upper()
            reserve_on = bool(getattr(self, "reserve_across_bom", False))

            # cost base from base_df
            cost_base = base_df.copy()
            for need in (
                "part_number",
                "BU",
                "available_quantity",
                "perpetual_avg_cost",
                "currency_cd",
            ):
                if need not in cost_base.columns:
                    cost_base[need] = (
                        0
                        if need not in ("part_number", "BU", "currency_cd")
                        else ("USD" if need == "currency_cd" else "")
                    )
            cost_base["available_quantity"] = (
                pd.to_numeric(cost_base["available_quantity"], errors="coerce")
                .fillna(0)
                .clip(lower=0)
            )
            cost_base["perpetual_avg_cost"] = pd.to_numeric(
                cost_base["perpetual_avg_cost"], errors="coerce"
            ).fillna(0.0)
            cost_base["BU"] = cost_base["BU"].astype(str).str.upper()
            cost_base["part_number"] = cost_base["part_number"].astype(str)

            def allocate_cost_for_row(pn: str, req: int) -> tuple[float, float, str]:
                """Preferred BU first, then cheapest of the rest (in target currency)."""
                if req <= 0:
                    return 0.0, 0.0, tgt
                rows = cost_base[cost_base["part_number"] == pn].copy()
                if rows.empty:
                    return 0.0, 0.0, tgt
                rows["unit_cost_tgt"] = rows.apply(
                    lambda r: _fx_convert_local(
                        r["perpetual_avg_cost"], r.get("currency_cd", tgt), tgt, rates
                    ),
                    axis=1,
                )
                rows["available_quantity"] = (
                    rows["available_quantity"].clip(lower=0).astype(int)
                )

                remaining = int(req)
                total = 0.0
                taken = 0

                # 1) preferred BU
                if pref_bu:
                    rp = rows[rows["BU"] == pref_bu].sort_values("unit_cost_tgt")
                    for _, rr in rp.iterrows():
                        if remaining <= 0:
                            break
                        take = min(int(rr["available_quantity"]), remaining)
                        if take > 0:
                            total += take * float(rr["unit_cost_tgt"])
                            remaining -= take
                            taken += take

                # 2) cheapest of the rest
                if remaining > 0:
                    rr = rows[rows["BU"] != pref_bu].sort_values("unit_cost_tgt")
                    for _, r2 in rr.iterrows():
                        if remaining <= 0:
                            break
                        take = min(int(r2["available_quantity"]), remaining)
                        if take > 0:
                            total += take * float(r2["unit_cost_tgt"])
                            remaining -= take
                            taken += take

                unit_used = (total / taken) if taken > 0 else 0.0

                # optional: consume avail_map if reserve toggled
                if reserve_on and pn in avail_map:
                    # simulate the same preference order
                    order = []
                    if pref_bu:
                        order.append(pref_bu)
                    order += [b for b in avail_map[pn].keys() if b != pref_bu]
                    need = int(req)
                    for b in order:
                        if need <= 0:
                            break
                        have = int(avail_map[pn].get(b, 0))
                        if have <= 0:
                            continue
                        take = min(have, need)
                        if take > 0:
                            avail_map[pn][b] = have - take
                            need -= take

                return float(total), float(unit_used), tgt

            def status_row(row) -> str:
                req = int(row.get("requested_qty", 0))
                if req <= 0:
                    return ""
                tot = int(row.get("total_available", 0))
                if tot <= 0:
                    return "OOS"
                # preferred only?
                if pref_bu and pref_bu in row.index and int(row.get(pref_bu, 0)) >= req:
                    return "OK (Preferred)"
                # any single BU can satisfy?
                if any(int(row.get(b, 0)) >= req for b in bu_cols):
                    return "OK"
                # split across BUs?
                if tot >= req:
                    return "SPLIT"
                return "SHORT"

            # description mapping: prefer Denodo then CSV
            denodo_desc = (
                base_df.dropna(subset=["part_number"])
                .groupby("part_number")["part_description"]
                .first()
            )
            pn_series = df_sum["part_number"].astype(str)
            desc_denodo = pn_series.map(denodo_desc).fillna("")
            desc_csv = pn_series.map(getattr(self, "csv_desc_map", {}) or {}).fillna("")
            df_sum["part_description"] = desc_denodo.mask(
                desc_denodo.str.strip().eq(""), desc_csv
            ).fillna("")

            # Bring through single-value attributes per PN (first non-null from base_df)
            def _first_map(col):
                if col in base_df.columns:
                    return pn_series.map(
                        base_df.dropna(subset=["part_number"])
                        .groupby("part_number")[col]
                        .first()
                    ).fillna("")
                return ""

            df_sum["lot"] = _first_map("lot")
            df_sum["serial"] = _first_map("serial")
            df_sum["q_codes"] = _first_map("q_codes")
            df_sum["mfg_id"] = _first_map("mfg_id")
            df_sum["mfg_part"] = _first_map("mfg_part")
            df_sum["item_status"] = _first_map("item_status")  # after agg_map fix
            df_sum["ctrl_in"] = _first_map("ctrl_in")
            df_sum["group_description"] = _first_map("group_description")

            # status + costs
            df_sum["status"] = df_sum.apply(status_row, axis=1)
            cost_res = [
                allocate_cost_for_row(pn, int(req))
                for pn, req in zip(df_sum["part_number"], df_sum["requested_qty"])
            ]
            df_sum["est_cost"] = [t[0] for t in cost_res]
            df_sum["perpetual_avg_cost_used"] = [t[1] for t in cost_res]
            df_sum["unit_currency_used"] = tgt
            df_sum["est_currency"] = tgt

            # final column order
            ordered = []
            for c in (
                [
                    "part_number",
                    "status",
                    "part_description",
                    "requested_qty",
                    "lot",
                    "serial",
                    "q_codes",
                    "mfg_id",
                    "mfg_part",
                    "item_status",
                    "ctrl_in",
                    "group_description",
                ]
                + bu_cols
                + [
                    "total_available",
                    "perpetual_avg_cost_used",
                    "unit_currency_used",
                    "est_cost",
                    "est_currency",
                ]
            ):
                _add(c, ordered, df_sum)
            view_df = df_sum[ordered] if ordered else df_sum

            self.current_view_df = view_df

        else:
            # ----- LONG VIEW -----
            view_df = base_df.copy()

            # normalize text/numeric
            if "perpetual_avg_cost" in view_df.columns:
                view_df["perpetual_avg_cost"] = (
                    pd.to_numeric(view_df["perpetual_avg_cost"], errors="coerce")
                    .fillna(0)
                    .round(4)
                )

            ordered = []
            for c in [
                "BU",
                "part_number",
                "part_description",
                "revision",
                "UOM",
                "q_codes",
                "lot",
                "serial",
                "available_quantity",
                "onhand_quantity",
                "reserved_quantity",
                "mfg_id",
                "mfg_part",
                "perpetual_avg_cost",
                "currency_cd",
                "family",
                "source",
                "item_status",
                "ctrl_in",
                "group_description",
            ]:
                _add(c, ordered, view_df)
            view_df = view_df[ordered] if ordered else view_df

            self.current_view_df = view_df

        # -------- table & proxy --------
        self.table_view = QTableView(self)
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table_view.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table_view.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_view.setSortingEnabled(True)

        self.model = PandasModel(
            self.current_view_df
            if isinstance(self.current_view_df, pd.DataFrame)
            else pd.DataFrame()
        )
        # optional: tell your model which column is 'status' for coloring
        if (
            isinstance(self.current_view_df, pd.DataFrame)
            and "status" in self.current_view_df.columns
        ):
            try:
                self.model.status_col = "status"
            except Exception:
                pass

        self.proxy = QSortFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)
        self.proxy.setDynamicSortFilter(True)
        self.table_view.setModel(self.proxy)
        self.table_view.sortByColumn(0, Qt.AscendingOrder)

        # --- header: restore widths if we have them; else auto-size once, then Interactive ---
        hh = self.table_view.horizontalHeader()
        hh.setSectionsClickable(True)
        hh.setSortIndicatorShown(True)
        hh.setHighlightSections(False)
        hh.setMinimumSectionSize(40)

        if not self._restore_header_state():  # try to restore saved widths
            # no saved state -> pick a good initial sizing
            self._init_column_sizes()  # see helper below

        # after initial sizing, allow the user to drag/rescale freely
        QTimer.singleShot(0, lambda: hh.setSectionResizeMode(QHeaderView.Interactive))

        # context menu + double-click (these should already exist on the class)
        self.table_view.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self._row_menu)
        self.table_view.doubleClicked.connect(self.on_double_click)

        # keyboard copy
        self.copy_shortcut = QShortcut(QKeySequence.Copy, self)
        self.copy_shortcut.activated.connect(self.copy_selection)

        # wire the "Search Files for Selected PN" button with proxy->source mapping
        def _open_sel_for_files():
            sm = self.table_view.selectionModel()
            if not sm:
                return
            rows = sm.selectedRows()
            if not rows:
                return
            view_idx = rows[0]
            src_idx = (
                self.proxy.mapToSource(view_idx)
                if isinstance(self.proxy, QSortFilterProxyModel)
                else view_idx
            )
            r = src_idx.row()
            dfv = getattr(self.model, "_dataframe", None)
            if dfv is None or "part_number" not in dfv.columns:
                return
            pn = str(dfv.iloc[r]["part_number"]).strip().upper()
            host = _host_with("open_files_for_pn")
            if pn and host and callable(getattr(host, "open_files_for_pn", None)):
                host.open_files_for_pn(pn)

        open_files.clicked.connect(_open_sel_for_files)

        # -------- root layout (vertical) --------
        root = QVBoxLayout(self)
        root.addLayout(ribbon)
        root.addWidget(self.table_view, 1)
        self.setLayout(root)

        self.setWindowTitle("Data Table")
        self.resize(1100, 650)

    def _col_index(self, name: str) -> int:
        try:
            return self.model._dataframe.columns.get_loc(name)
        except Exception:
            return -1

    def _init_column_sizes(self):
        """One-time, fast and sane auto sizing: contents for narrow cols; stretch a key column."""
        from PyQt5.QtWidgets import QHeaderView

        df = getattr(self.model, "_dataframe", None)
        if df is None or df.empty:
            return

        hh = self.table_view.horizontalHeader()
        # 1) size to contents (one shot; much faster than keeping ResizeToContents on)
        hh.setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table_view.resizeColumnsToContents()

        # 2) optionally stretch a “main” text column to fill the extra space
        #    long view: prefer 'part_description'; BOM: still stretch description if present
        stretch_candidate = None
        for cand in ("part_description", "status", "part_number"):
            i = self._col_index(cand)
            if i >= 0:
                stretch_candidate = i
                break
        if stretch_candidate is not None:
            hh.setSectionResizeMode(stretch_candidate, QHeaderView.Stretch)

        # cap any overly-wide columns
        max_w = 600
        for i in range(df.shape[1]):
            w = self.table_view.columnWidth(i)
            if w > max_w:
                self.table_view.setColumnWidth(i, max_w)

    def _save_header_state(self):
        """Persist user-resized widths to prefs JSON."""
        try:
            state = self.table_view.horizontalHeader().saveState()
            prefs = load_json_or_default(app_paths()["prefs"], DEFAULT_PREFS)
            prefs.setdefault("ui", {}).setdefault("tables", {})["datawindow_header"] = (
                bytes(state).hex()
            )
            save_json_atomic(app_paths()["prefs"], prefs)
            return True
        except Exception:
            return False

    def _restore_header_state(self) -> bool:
        """Restore widths if we have them. Returns True if restored."""
        try:
            prefs = load_json_or_default(app_paths()["prefs"], DEFAULT_PREFS)
            hhex = prefs.get("ui", {}).get("tables", {}).get("datawindow_header")
            if not hhex:
                return False
            ba = QByteArray.fromHex(hhex.encode("ascii"))
            return bool(self.table_view.horizontalHeader().restoreState(ba))
        except Exception:
            return False

    def closeEvent(self, e):
        """Save widths when this window/pane is closed/destroyed."""
        try:
            self._save_header_state()
        finally:
            super().closeEvent(e)

    # ------------------- Public helpers -------------------
    def set_back_enabled(self, enabled: bool, callback=None):
        self.on_back = callback
        # find the back button (first in top row)
        # kept simple; can add handle if you expose it as self.btn_back
        # callers typically show a back button in the parent, so most leave this hidden
        # No-op here unless you want to store a handle to the back button

    def set_dataframe(self, df: pd.DataFrame):
        """Swap data after creation; keeps model/proxy/table intact."""
        self.current_view_df = df if isinstance(df, pd.DataFrame) else pd.DataFrame()
        if hasattr(self.model, "update_dataframe"):
            self.model.update_dataframe(self.current_view_df)
        else:
            # fallback: replace model and rebind proxy
            self.model = PandasModel(self.current_view_df)
            self.proxy.setSourceModel(self.model)
            self.table_view.setModel(self.proxy)

    # ------------------- Actions -------------------
    def save_data(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Save File", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        df_out = (
            self.current_view_df.copy()
            if isinstance(self.current_view_df, pd.DataFrame)
            else self.df.copy()
        )
        if "est_cost" in df_out.columns and not df_out.empty:
            total = pd.DataFrame(
                [
                    {
                        df_out.columns[0]: "TOTAL",
                        "est_cost": df_out["est_cost"].sum(),
                        "est_currency": self.target_currency,
                    }
                ]
            )
            df_out = pd.concat([df_out, total], ignore_index=True)

        df_out.to_excel(path, index=False)

        ok, msg = apply_excel_status_colors(
            path,
            status_header="status",
            preferred_bu=getattr(self, "preferred_bu", ""),
            bu_headers=getattr(self, "selected_bus", None),
            status_palette=STATUS_BG_HEX,
            import_fields=["part_number", "part_description", "requested_qty"],
            cost_fields=[
                "perpetual_avg_cost_used",
                "unit_currency_used",
                "est_cost",
                "est_currency",
            ],
        )
        QMessageBox.information(self, "Export", msg if ok else f"Export note: {msg}")

    def copy_selection(self):
        sel_model = self.table_view.selectionModel()
        if sel_model is None:
            return
        idxs = sel_model.selectedIndexes()

        def to_src(ix):
            m = self.table_view.model()
            return m.mapToSource(ix) if isinstance(m, QSortFilterProxyModel) else ix

        if not idxs:
            ix = self.table_view.currentIndex()
            if ix.isValid():
                s = to_src(ix)
                val = str(self.model._dataframe.iloc[s.row(), s.column()])
                QApplication.clipboard().setText(val)
            return

        # Map all to source, then sort by source coords
        sidxs = sorted((to_src(ix) for ix in idxs), key=lambda i: (i.row(), i.column()))
        if len(sidxs) == 1:
            s = sidxs[0]
            val = str(self.model._dataframe.iloc[s.row(), s.column()])
            QApplication.clipboard().setText(val)
            return

        rows = {}
        cols_used = set()
        for s in sidxs:
            r, c = s.row(), s.column()
            rows.setdefault(r, {})[c] = str(self.model._dataframe.iloc[r, c])
            cols_used.add(c)

        cols_used = sorted(cols_used)
        headers = [str(self.model._dataframe.columns[c]) for c in cols_used]
        lines = ["\t".join(headers)]
        for r in sorted(rows):
            values = [rows[r].get(c, "") for c in cols_used]
            lines.append("\t".join(values))
        QApplication.clipboard().setText("\n".join(lines))

    def on_double_click(self, index: QModelIndex):
        try:
            if not index.isValid():
                return

            # If a proxy is present, map the clicked cell back to the source model
            src_index = index
            src_model = self.table_view.model()
            if isinstance(src_model, QSortFilterProxyModel):
                src_index = src_model.mapToSource(index)

            row = src_index.row()
            col = src_index.column()

            df_view = (
                self.model._dataframe
            )  # <- always source dataframe behind the proxy
            colname = str(df_view.columns[col])

            # 1) BU cell in LONG view -> all parts in that BU
            if colname == "BU":
                bu_val = str(df_view.iloc[row]["BU"]).strip()
                long = self.df.copy()

                if "BU" not in long.columns and "business_unit" in long.columns:
                    long = long.rename(columns={"business_unit": "BU"})

                # Create mask that treats '(UNASSIGNED)' as NULL/blank in the original data
                if bu_val == "(UNASSIGNED)":
                    mask = long["BU"].isna() | (
                        long["BU"].astype(str).str.strip() == ""
                    )
                else:
                    mask = long["BU"].astype(str).str.strip().eq(bu_val)

                slice_df = long[mask].copy()
                # Display the same '(UNASSIGNED)' label in the slice for clarity
                slice_df["BU"] = slice_df["BU"].astype(str)
                slice_df["BU"] = slice_df["BU"].where(
                    slice_df["BU"].str.strip() != "", "(UNASSIGNED)"
                )

                self.slice_win = SliceWindow(
                    f"Business Unit: {bu_val or '(UNASSIGNED)'}",
                    slice_df.reset_index(drop=True),
                )
                self.slice_win.show()
                return

            # 2) Part number -> aggregate across BUs
            if colname == "part_number":
                pn = str(df_view.iloc[row]["part_number"]).strip().upper()
                long = self.df.copy()

                # normalize columns
                if "part_number" not in long.columns and "item_id" in long.columns:
                    long = long.rename(columns={"item_id": "part_number"})
                if "BU" not in long.columns and "business_unit" in long.columns:
                    long = long.rename(columns={"business_unit": "BU"})

                long["part_number"] = long["part_number"].astype(str).str.upper()
                long = long[long["part_number"] == pn]
                if long.empty:
                    QMessageBox.information(self, "Drilldown", "No rows for this part.")
                    return

                for c in ["available_quantity", "onhand_quantity", "reserved_quantity"]:
                    if c not in long.columns:
                        long[c] = 0
                    long[c] = (
                        pd.to_numeric(long[c], errors="coerce").fillna(0).clip(lower=0)
                    )

                g = (
                    long.groupby("BU", as_index=False)[
                        ["available_quantity", "onhand_quantity", "reserved_quantity"]
                    ]
                    .sum()
                    .sort_values("BU")
                )
                self.slice_win = SliceWindow(
                    f"Part {pn}: All BUs", g.reset_index(drop=True)
                )
                self.slice_win.show()
                return

            # 3) Quick-copy convenience
            if colname in ("part_number", "part_description"):
                val = str(df_view.iloc[row][colname])
                QApplication.clipboard().setText(val)
                QMessageBox.information(
                    self, "Copied", f"{colname.replace('_',' ').title()} copied."
                )
                return

        except Exception as e:
            QMessageBox.warning(self, "Drilldown error", str(e))

    def _row_menu(self, pos):
        view = self.table_view
        sm = view.selectionModel()
        if not sm:
            return
        idxs = sm.selectedRows()
        if not idxs:
            return

        idx = idxs[0]
        model = view.model()
        if isinstance(model, QSortFilterProxyModel):
            src = model.mapToSource(idx)
        else:
            src = idx

        row = src.row()
        dfv = self.model._dataframe
        pn = None
        if "part_number" in dfv.columns:
            pn = str(dfv.iloc[row]["part_number"]).strip().upper()
        if not pn:
            return

        def _go():
            p = self.parent()
            while p is not None and not hasattr(p, "open_files_for_pn"):
                p = p.parent()
            if p and callable(getattr(p, "open_files_for_pn", None)):
                if file_search_enabled():
                    p.open_files_for_pn(pn)
                else:
                    QMessageBox.information(
                        self, "File Search", "File search is disabled by policy."
                    )

        m = QMenu(self)
        m.addAction("Find files for this PN").triggered.connect(_go)
        m.exec_(view.viewport().mapToGlobal(pos))
        m.addSeparator()
        m.addAction("Auto size → by contents").triggered.connect(
            self._init_column_sizes
        )

        def _auto_header():
            # width = header text + padding, then Interactive
            fm = self.table_view.fontMetrics()
            df = self.model._dataframe
            for i, col in enumerate(df.columns):
                w = fm.horizontalAdvance(str(col)) + 24
                self.table_view.setColumnWidth(i, max(60, min(600, w)))

        m.addAction("Auto size → by header").triggered.connect(_auto_header)

    # ---- optional fallback if you don't have a custom exporter wired ----
    def _export_excel_fallback(self):
        try:
            path, _ = QFileDialog.getSaveFileName(
                self, "Export to Excel", "results.xlsx", "Excel Files (*.xlsx)"
            )
            if not path:
                return
            rows = []
            cols = self.model.columnCount()
            for r in range(self.proxy.rowCount()):
                row_vals = []
                for c in range(cols):
                    idx = self.proxy.index(r, c)
                    row_vals.append(self.proxy.data(idx, Qt.DisplayRole))
                rows.append(row_vals)
            headers = [
                self.model.headerData(c, Qt.Horizontal, Qt.DisplayRole)
                for c in range(cols)
            ]
            pd.DataFrame(rows, columns=headers).to_excel(path, index=False)
        except Exception as e:
            QMessageBox.warning(self, "Export", f"Could not export:\n{e}")


class SliceWindow(QWidget):
    """
    Simple read-only table view for drilldowns (BU slice or PN slice).
    Adds Export / Copy buttons and a '?' helper.
    """

    def __init__(self, title: str, df: pd.DataFrame, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title or "Slice")
        self.df = df if isinstance(df, pd.DataFrame) else pd.DataFrame()

        # --- layout
        root = QVBoxLayout(self)

        # --- top bar (actions)
        bar = QHBoxLayout()
        self.btn_export = QPushButton("Export")
        self.btn_copy = QPushButton("Copy")
        self.btn_help = QPushButton("?")
        self.btn_help.setFixedWidth(28)
        self.btn_help.setToolTip(
            "• Click a column header to sort.\n"
            "• Select cells and press Ctrl/Cmd+C, or use 'Copy'.\n"
            "• 'Export' saves the visible slice to Excel."
        )
        self.btn_export.clicked.connect(self._export_excel)
        self.btn_copy.clicked.connect(self._copy_selection)
        bar.addWidget(self.btn_export)
        bar.addWidget(self.btn_copy)
        bar.addStretch(1)
        bar.addWidget(self.btn_help)
        root.addLayout(bar)

        # --- table
        self.table = QTableView(self)
        self.table.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.model = PandasModel(self.df.copy())
        self.proxy = QSortFilterProxyModel(self)
        self.proxy.setSourceModel(self.model)
        self.table.setModel(self.proxy)
        self.table.setSortingEnabled(True)
        self.table.sortByColumn(0, Qt.AscendingOrder)
        root.addWidget(self.table, 1)

        # keyboard copy
        self.copy_sc = QShortcut(QKeySequence.Copy, self)
        self.copy_sc.activated.connect(self._copy_selection)

        self.resize(900, 550)

    # ---------- actions ----------
    def _export_excel(self):
        if self.df is None or self.df.empty:
            QMessageBox.information(self, "Export", "Nothing to export.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Slice", "", "Excel Files (*.xlsx)"
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # Export the data in current proxy order
        src = self.model._dataframe
        if src is None or src.empty:
            QMessageBox.information(self, "Export", "Nothing to export.")
            return

        # Rebuild a DataFrame using the proxy's current row order
        rows = []
        for r in range(self.proxy.rowCount()):
            ir = self.proxy.index(r, 0)
            sr = self.proxy.mapToSource(ir).row()
            rows.append(src.iloc[sr].copy())
        out = pd.DataFrame(rows)
        try:
            out.to_excel(path, index=False)
            QMessageBox.information(self, "Export", f"Saved:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Export", f"Could not save:\n{e}")

    def _copy_selection(self):
        sel = self.table.selectionModel()
        if sel is None:
            return
        idxs = sel.selectedIndexes()
        if not idxs:
            # copy current cell if any
            ix = self.table.currentIndex()
            if ix.isValid():
                src_ix = self.proxy.mapToSource(ix)
                val = str(self.model._dataframe.iat[src_ix.row(), src_ix.column()])
                QApplication.clipboard().setText(val)
            return

        # Normalize into rectangular block
        idxs = sorted(idxs, key=lambda x: (x.row(), x.column()))
        min_r = min(i.row() for i in idxs)
        max_r = max(i.row() for i in idxs)
        min_c = min(i.column() for i in idxs)
        max_c = max(i.column() for i in idxs)

        lines = []
        headers = [
            str(
                self.model._dataframe.columns[
                    self.proxy.mapToSource(self.proxy.index(0, c)).column()
                ]
            )
            for c in range(min_c, max_c + 1)
        ]
        lines.append("\t".join(headers))

        for r in range(min_r, max_r + 1):
            row_vals = []
            for c in range(min_c, max_c + 1):
                p = self.proxy.index(r, c)
                s = self.proxy.mapToSource(p)
                try:
                    row_vals.append(str(self.model._dataframe.iat[s.row(), s.column()]))
                except Exception:
                    row_vals.append("")
            lines.append("\t".join(row_vals))
        QApplication.clipboard().setText("\n".join(lines))


class SQLHighlighter(QSyntaxHighlighter):
    def __init__(self, document):
        super().__init__(document)
        self.rules = []

        def fmt(
            color: str, bold: bool = False, italic: bool = False
        ) -> QTextCharFormat:
            f = QTextCharFormat()
            f.setForeground(QColor(color))
            if bold:
                f.setFontWeight(QFont.Bold)
            if italic:
                f.setFontItalic(True)
            return f

        # Styles
        self.kw_format = fmt("#0077aa", True)  # keywords
        self.fn_format = fmt("#6f42c1")  # functions
        self.num_format = fmt("#1a7f37")  # numbers
        self.str_format = fmt("#d14")  # strings
        self.cmt_format = fmt("#6a737d", italic=True)  # comments

        # Keywords & functions
        keywords = (
            "SELECT",
            "DISTINCT",
            "FROM",
            "WHERE",
            "GROUP",
            "BY",
            "HAVING",
            "ORDER",
            "LIMIT",
            "OFFSET",
            "JOIN",
            "LEFT",
            "RIGHT",
            "INNER",
            "OUTER",
            "FULL",
            "ON",
            "AS",
            "AND",
            "OR",
            "NOT",
            "IN",
            "LIKE",
            "IS",
            "NULL",
            "CASE",
            "WHEN",
            "THEN",
            "ELSE",
            "END",
            "UNION",
            "ALL",
            "EXISTS",
        )
        for kw in keywords:
            self.rules.append(
                (
                    re.compile(r"\b" + re.escape(kw) + r"\b", re.IGNORECASE),
                    self.kw_format,
                )
            )

        self.rules.append(
            (
                re.compile(r"\b(SUM|COUNT|AVG|MIN|MAX|COALESCE|NVL)\b", re.IGNORECASE),
                self.fn_format,
            )
        )
        self.rules.append((re.compile(r"\b\d+(?:\.\d+)?\b"), self.num_format))
        self.rules.append((re.compile(r"'([^']|'')*'"), self.str_format))

        # Comments
        self.sl_comment = re.compile(r"--[^\n]*")
        self.ml_start = re.compile(r"/\*")
        self.ml_end = re.compile(r"\*/")

    def highlightBlock(self, text: str):
        # token rules
        for pattern, fmt in self.rules:
            for m in pattern.finditer(text):
                self.setFormat(m.start(), m.end() - m.start(), fmt)

        # single-line comments
        for m in self.sl_comment.finditer(text):
            self.setFormat(m.start(), len(text) - m.start(), self.cmt_format)

        # multi-line comments
        self.setCurrentBlockState(0)
        if self.previousBlockState() != 1:
            m = self.ml_start.search(text)
            if m:
                start = m.start()
                endm = self.ml_end.search(text, m.end())
                if endm:
                    self.setFormat(start, endm.end() - start, self.cmt_format)
                else:
                    self.setFormat(start, len(text) - start, self.cmt_format)
                    self.setCurrentBlockState(1)
        else:
            endm = self.ml_end.search(text)
            if endm:
                self.setFormat(0, endm.end(), self.cmt_format)
                self.setCurrentBlockState(0)
            else:
                self.setFormat(0, len(text), self.cmt_format)
                self.setCurrentBlockState(1)


class SqlWindow(QWidget):
    def __init__(self, sql: str):
        super().__init__()
        layout = QGridLayout()
        self.text = QPlainTextEdit()
        self.highlighter = SQLHighlighter(self.text.document())
        # Monospace + no wrapping for readability
        try:
            from PyQt5.QtGui import QFont

            self.text.setFont(QFont("Consolas"))
        except Exception:
            pass
        self.text.setLineWrapMode(QPlainTextEdit.NoWrap)
        try:
            self.text.setPlainText(pretty_sql(sql))
        except Exception:
            self.text.setPlainText(sql)
        self.text.setReadOnly(True)
        layout.addWidget(self.text, 0, 0, 1, 3)

        # Copy (pretty, single line collapsed)
        self.btn_copy = QPushButton("Copy")
        self.btn_copy.clicked.connect(
            lambda: QApplication.clipboard().setText(
                " ".join(self.text.toPlainText().split())
            )
        )
        layout.addWidget(self.btn_copy, 1, 0)

        # Copy (exact, multi-line)
        self.btn_copy_oneline = QPushButton("Copy (exact)")
        self.btn_copy_oneline.clicked.connect(
            lambda: QApplication.clipboard().setText(self.text.toPlainText())
        )
        layout.addWidget(self.btn_copy_oneline, 1, 1)

        self.btn_close = QPushButton("Close")
        self.btn_close.clicked.connect(self.close)
        layout.addWidget(self.btn_close, 1, 2)
        self.setLayout(layout)
        self.setWindowTitle("Equivalent SQL")
        self.resize(900, 600)


# ----------------------------------
# Main
# ----------------------------------
if __name__ == "__main__":
    try:
        bootstrap_files()  # make sure JSONs exist before app reads them
    except Exception as e:
        print("Bootstrap warning:", e)

    app = QApplication(sys.argv)

    # load theme pref
    prefs = load_json_or_default(app_paths()["prefs"], DEFAULT_PREFS)
    ui_prefs = prefs.get("ui", {})
    theme_mode = (ui_prefs.get("theme") or "system").lower()
    apply_theme(app, theme_mode)  # <-- apply before creating windows

    win = DenodoQuery()
    win.show()
    sys.exit(app.exec_())
